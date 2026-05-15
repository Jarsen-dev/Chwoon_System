"""
Router para Plan de Inyección:
- CRUD del plan con máquina, prioridad y secuencia de partes
- Importación Excel (soporta celdas combinadas)
- Avance de producción, paros, reanudación, finalización y auto-siguiente
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from typing import List, Optional
from datetime import datetime, timedelta, timezone
import pandas as pd
import io
import uuid

from app.database import AsyncSessionLocal
from app.models.plan_inyeccion import PlanInyeccion
from app.models.registro_avance_inyeccion import RegistroAvanceInyeccion
from app.models.producto import Producto
from app.models.plan_produccion import PlanProduccion
from app.models.lote_inventario import LoteInventario, MovimientoLote
from app.models.ubicacion import Ubicacion
from app.core.deps import get_current_user
from app.models.usuario import Usuario
from app.schemas.ordenes_produccion import (
    PlanInyeccionCreate,
    PlanInyeccionBatchCreate,
    PlanInyeccionResponse,
    PlanInyeccionAvanceRequest,
    PlanInyeccionParoRequest,
    PlanInyeccionReanudarRequest,
    PlanInyeccionFinalizarResponse,
)

router = APIRouter(prefix="/plan-inyeccion", tags=["plan-inyeccion"])


async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


def _ahora_utc() -> datetime:
    """Retorna datetime UTC naive compatible con asyncpg/DateTime"""
    return datetime.utcnow()


def _require_prod_role(user: Usuario):
    if user.rol.value not in ("admin", "supervisor", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos")

# ══════════════════════════════════════════════════════════════════
# FIX: Constante de zona horaria GMT-6 (igual que en turno_service.py)
# ══════════════════════════════════════════════════════════════════
TZ_LOCAL = timezone(timedelta(hours=-6))


def _utc_to_local(dt: datetime) -> datetime:
    """Convierte UTC naive a GMT-6"""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_LOCAL)


def _local_naive_to_utc_naive(dt: datetime) -> datetime:
    """Convierte datetime naive GMT-6 a UTC naive (suma 6 horas)"""
    return dt + timedelta(hours=6)


def _get_turno_real(hora_inicio: Optional[datetime]) -> str:
    """
    Determina el turno real basado en la hora de inicio convertida a GMT-6.
    DIA: 07:30 a 19:29:59
    NOCHE: 19:30 a 07:29:59 (del día siguiente)
    """
    if not hora_inicio:
        return "NOCHE"
    
    local = _utc_to_local(hora_inicio)
    total_minutes = local.hour * 60 + local.minute
    
    # DIA: 07:30 = 450 min, hasta 19:29:59 = 1169 min
    if 450 <= total_minutes < 1170:
        return "DIA"
    return "NOCHE"


import re

async def _construir_lote(
    db: AsyncSession,
    numero_parte: str,
    turno_real: str,
    fecha: str,
    maquina_orden: str,
) -> str:
    """
    Construye el lote: ULTIMOS4_TURNO(D/N)_FECHA_NUMMAQUINA
    La máquina se extrae de PlanProduccion (prioridad) o de la orden.
    Solo se toman los dígitos finales (MAQ-01 → 01).
    """
    # Buscar máquina en PlanProd
    plan_result = await db.execute(
        select(PlanProduccion).where(PlanProduccion.numero_parte == numero_parte)
    )
    plan_item = plan_result.scalar_one_or_none()

    maquina_raw = (
        (plan_item.maquina if plan_item and plan_item.maquina else "")
        or maquina_orden
        or "SINMAQUINA"
    )

    # Extraer solo los dígitos finales
    match = re.search(r'(\d+)$', maquina_raw)
    num_maquina = match.group(1) if match else "00"

    lote_parte = numero_parte[-4:] if len(numero_parte) >= 4 else numero_parte
    lote_turno = "D" if turno_real == "DIA" else "N"
    lote_fecha = fecha.replace("-", "")

    return f"{lote_parte}_{lote_turno}_{lote_fecha}_{num_maquina}"


# ══════════════════════════════════════════════════════════════════
# CRUD
# ══════════════════════════════════════════════════════════════════

@router.get("/", response_model=List[PlanInyeccionResponse])
@router.get("", response_model=List[PlanInyeccionResponse])
async def listar_plan(
    status: Optional[str] = None,
    maquina: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _require_prod_role(user)
    q = select(PlanInyeccion).order_by(
        PlanInyeccion.prioridad.asc(),
        PlanInyeccion.maquina.asc(),
        PlanInyeccion.orden_secuencia.asc(),
    )
    if status:
        q = q.where(PlanInyeccion.status == status)
    if maquina:
        q = q.where(PlanInyeccion.maquina == maquina.upper())
    result = await db.execute(q)
    ordenes = result.scalars().all()

    if ordenes:
        plan_ids = [p.id for p in ordenes]
        ultimos_avances = await db.execute(
            select(
                RegistroAvanceInyeccion.plan_inyeccion_id,
                func.max(RegistroAvanceInyeccion.timestamp).label("ultimo_ts"),
            )
            .where(RegistroAvanceInyeccion.plan_inyeccion_id.in_(plan_ids))
            .group_by(RegistroAvanceInyeccion.plan_inyeccion_id)
        )
        avance_map = {row.plan_inyeccion_id: row.ultimo_ts for row in ultimos_avances.all()}
    else:
        avance_map = {}

    return [
        PlanInyeccionResponse.model_validate({
            **{c.name: getattr(p, c.name) for c in p.__table__.columns},
            "hora_ultimo_avance": avance_map.get(p.id),
        })
        for p in ordenes
    ]


@router.post("/", response_model=PlanInyeccionResponse)
@router.post("", response_model=PlanInyeccionResponse)
async def crear_item(
    data: PlanInyeccionCreate,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    data.maquina = data.maquina.strip().upper()
    data.numero_parte = data.numero_parte.strip().upper()

    res = await db.execute(
        select(PlanInyeccion).where(
            and_(
                PlanInyeccion.prioridad == data.prioridad,
                PlanInyeccion.maquina != data.maquina,
                PlanInyeccion.status != "Finalizado",
            )
        )
    )
    if res.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"La prioridad {data.prioridad} ya está asignada a otra máquina activa",
        )

    item = PlanInyeccion(**data.model_dump())
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return item


@router.post("/batch", response_model=List[PlanInyeccionResponse])
async def crear_batch(
    data: PlanInyeccionBatchCreate,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    if not data.items:
        raise HTTPException(status_code=400, detail="Lista vacía")

    maquina = data.items[0].maquina.strip().upper()
    prioridad = data.items[0].prioridad

    for it in data.items:
        if it.maquina.strip().upper() != maquina or it.prioridad != prioridad:
            raise HTTPException(status_code=400, detail="Todos los items deben tener la misma máquina y prioridad")

    res = await db.execute(
        select(PlanInyeccion).where(
            and_(
                PlanInyeccion.prioridad == prioridad,
                PlanInyeccion.maquina != maquina,
                PlanInyeccion.status != "Finalizado",
            )
        )
    )
    if res.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"La prioridad {prioridad} ya está asignada a otra máquina activa",
        )

    res_seq = await db.execute(
        select(func.coalesce(func.max(PlanInyeccion.orden_secuencia), -1)).where(
            and_(
                PlanInyeccion.maquina == maquina,
                PlanInyeccion.prioridad == prioridad,
            )
        )
    )
    max_seq = res_seq.scalar() or -1

    creados = []
    for idx, it in enumerate(data.items):
        it.maquina = maquina
        it.numero_parte = it.numero_parte.strip().upper()
        if it.orden_secuencia == 0:
            it.orden_secuencia = max_seq + 1 + idx
        item = PlanInyeccion(**it.model_dump())
        db.add(item)
        creados.append(item)

    await db.commit()
    for c in creados:
        await db.refresh(c)
    return creados


@router.delete("/{id}")
async def eliminar_item(
    id: int,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    await db.delete(item)
    await db.commit()
    return {"message": "Eliminado"}


@router.patch("/{id}/aux-silo")
async def asignar_aux_silo(
    id: int,
    aux_silo: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Asigna o cambia el Silo Aux de una orden de plan de inyección."""
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")

    # Validar que el AUX exista como ubicación
    if aux_silo and aux_silo.strip():
        aux_nombre = aux_silo.strip()
        res_ub = await db.execute(select(Ubicacion).where(Ubicacion.nombre == aux_nombre))
        ub = res_ub.scalar_one_or_none()
        if not ub:
            raise HTTPException(status_code=400, detail=f"La ubicación '{aux_nombre}' no existe")
        item.aux_silo = aux_nombre
    else:
        item.aux_silo = None

    await db.commit()
    await db.refresh(item)
    return {"message": f"Silo Aux actualizado: {item.aux_silo or 'Ninguno'}", "aux_silo": item.aux_silo}


# ══════════════════════════════════════════════════════════════════
# IMPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════

@router.post("/importar-excel")
async def importar_excel(
    file: UploadFile = File(...),
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Archivo debe ser Excel o CSV")

    contents = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo: {e}")

    df.dropna(how='all', inplace=True)

    def find_col(cols, keys):
        for c in cols:
            low = str(c).lower().strip()
            if any(k in low for k in keys):
                return c
        return None

    col_maquina = find_col(df.columns, ['maquina', 'machine', 'maq'])
    col_prioridad = find_col(df.columns, ['prioridad', 'prio', 'priority'])
    col_parte = find_col(df.columns, ['parte', 'part', 'numero', 'no.', 'no '])
    col_plan = find_col(df.columns, ['plan', 'meta', 'cantidad', 'qty', 'piezas', 'total'])

    if not col_maquina:
        raise HTTPException(status_code=400, detail="No se encontró columna 'Maquina'")
    if not col_prioridad:
        raise HTTPException(status_code=400, detail="No se encontró columna 'Prioridad'")
    if not col_parte:
        raise HTTPException(status_code=400, detail="No se encontró columna 'Numero de Parte'")
    if not col_plan:
        raise HTTPException(status_code=400, detail="No se encontró columna 'Plan'")

    df[col_maquina] = df[col_maquina].ffill()
    df[col_prioridad] = df[col_prioridad].ffill()

    errores = []
    creados = 0

    for idx, row in df.iterrows():
        maquina = str(row[col_maquina]).strip().upper()
        prioridad_str = str(row[col_prioridad]).strip()
        parte = str(row[col_parte]).strip().upper()
        plan_str = str(row[col_plan]).strip()

        if not maquina or maquina == 'NAN' or not parte or parte == 'NAN':
            continue

        try:
            prioridad = int(float(prioridad_str))
        except (ValueError, TypeError):
            errores.append(f"Fila {idx+2}: Prioridad inválida")
            continue

        try:
            plan = int(float(plan_str))
            if plan <= 0:
                continue
        except (ValueError, TypeError):
            errores.append(f"Fila {idx+2}: Plan inválido para {parte}")
            continue

        res = await db.execute(
            select(PlanInyeccion).where(
                and_(
                    PlanInyeccion.prioridad == prioridad,
                    PlanInyeccion.maquina != maquina,
                    PlanInyeccion.status != "Finalizado",  # ← Solo bloquear si NO está finalizada
                )
            )
        )
        if res.scalar_one_or_none():
            errores.append(f"Fila {idx+2}: Prioridad {prioridad} ya usada por otra máquina activa")
            continue

        res_seq = await db.execute(
            select(func.coalesce(func.max(PlanInyeccion.orden_secuencia), -1)).where(
                and_(
                    PlanInyeccion.maquina == maquina,
                    PlanInyeccion.prioridad == prioridad,
                )
            )
        )
        next_seq = (res_seq.scalar() or -1) + 1

        item = PlanInyeccion(
            maquina=maquina,
            prioridad=prioridad,
            numero_parte=parte,
            plan_piezas=plan,
            orden_secuencia=next_seq,
            status="Pendiente",
        )
        db.add(item)
        creados += 1

    await db.commit()
    return {
        "message": "Importación completada",
        "creados": creados,
        "errores": errores,
    }


# ══════════════════════════════════════════════════════════════════
# PROCESO: INICIAR / AVANZAR / PARO / REANUDAR / FINALIZAR
# ══════════════════════════════════════════════════════════════════

@router.post("/{id}/iniciar", response_model=PlanInyeccionResponse)
async def iniciar_orden(
    id: int,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    if item.status != "Pendiente":
        raise HTTPException(status_code=400, detail="La orden no está pendiente")

    ahora = _ahora_utc()
    item.status = "En Proceso"
    item.hora_inicio = ahora
    item.hora_ultimo_inicio = ahora
    item.en_paro = False  # FIX: Inicia produciendo directamente

    await db.commit()
    await db.refresh(item)
    return item


@router.post("/{id}/avanzar", response_model=PlanInyeccionResponse)
async def avanzar_produccion(
    id: int,
    data: PlanInyeccionAvanceRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    if item.status != "En Proceso":
        raise HTTPException(status_code=400, detail="La orden no está en proceso")
    if item.en_paro:
        raise HTTPException(status_code=400, detail="La máquina está enparo. Reanude primero desde la sub-tab Paros.")

    # ══════════════════════════════════════════════════════════════════
    # FIX: Validar que no haya avance en la misma franja horaria
    # ══════════════════════════════════════════════════════════════════
    from datetime import timedelta, timezone
    
    TZ_LOCAL = timezone(timedelta(hours=-6))
    
    def get_franja_index(dt: datetime) -> int:
        """Retorna índice de franja horaria en GMT-6"""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        local = dt.astimezone(TZ_LOCAL)
        total_minutes = local.hour * 60 + local.minute
        
        if 450 <= total_minutes < 1170:  # Día
            return (total_minutes - 450) // 60
        elif total_minutes >= 1170:  # Noche primera parte
            return (total_minutes - 1170) // 60 + 12
        else:  # Noche segunda parte (madrugada)
            return (total_minutes + 1440 - 1170) // 60 + 12

    # Verificar último avance
    ultimo_avance = await db.execute(
        select(RegistroAvanceInyeccion)
        .where(RegistroAvanceInyeccion.plan_inyeccion_id == id)
        .order_by(RegistroAvanceInyeccion.timestamp.desc())
        .limit(1)
    )
    ultimo = ultimo_avance.scalar_one_or_none()
    
    if ultimo:
        franja_ultimo = get_franja_index(ultimo.timestamp)
        franja_actual = get_franja_index(datetime.utcnow())
        
        if franja_ultimo == franja_actual:
            raise HTTPException(
                status_code=400, 
                detail="Ya se registró un avance en esta franja horaria. Espere la siguiente hora."
            )
        elif franja_actual < franja_ultimo:
            # Caso especial: cambio de turno o inconsistencia
            pass  # Permitir, podría ser nuevo turno

    # FIX: Recalcular piezas desde contador_hora × cav para evitar inconsistencias
    piezas_reales = data.contador_hora * (item.cav or 1)
    item.piezas_producidas += piezas_reales

    # ══════════════════════════════════════════════════════════════════
    # CONSUMO DE MATERIAL DESDE AUX SILO
    # ══════════════════════════════════════════════════════════════════
    if item.aux_silo:
        # Obtener peso seco del producto
        res_prod = await db.execute(select(Producto).where(Producto.sku == item.numero_parte))
        producto = res_prod.scalar_one_or_none()
        peso_seco = 0.0
        if producto and producto.caracteristicas_inyeccion:
            peso_seco = float(producto.caracteristicas_inyeccion.get("peso_seco") or 0)

        if peso_seco > 0 and piezas_reales > 0:
            kg_necesarios = piezas_reales * peso_seco

            # Buscar ubicación AUX
            res_ub = await db.execute(select(Ubicacion).where(Ubicacion.nombre == item.aux_silo))
            ub = res_ub.scalar_one_or_none()
            if ub:
                # Buscar lotes con stock en el AUX
                res_lotes = await db.execute(
                    select(LoteInventario).where(
                        and_(
                            LoteInventario.ubicacion_id == ub.id,
                            LoteInventario.cantidad_actual > 0,
                            LoteInventario.estado_calidad == "Aprobado",
                        )
                    ).order_by(LoteInventario.fecha_recepcion)
                )
                lotes = res_lotes.scalars().all()
                stock_disponible = sum(l.cantidad_actual for l in lotes)

                if stock_disponible < kg_necesarios:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Stock insuficiente en {item.aux_silo}. "
                            f"Disponible: {stock_disponible:.2f} kg, "
                            f"Requerido: {kg_necesarios:.2f} kg "
                            f"({piezas_reales} pz × {peso_seco} kg/pz)"
                        ),
                    )

                # Consumir FIFO del AUX
                restante = kg_necesarios
                for lote in lotes:
                    if restante <= 0:
                        break
                    tomar = min(lote.cantidad_actual, restante)
                    lote.cantidad_actual -= tomar
                    db.add(MovimientoLote(
                        lote_id=lote.lote_id,
                        fecha=datetime.utcnow(),
                        tipo="CONSUMO_INYECCION",
                        cantidad=-tomar,
                        detalles={
                            "plan_inyeccion_id": item.id,
                            "numero_parte": item.numero_parte,
                            "maquina": item.maquina,
                            "piezas_producidas": piezas_reales,
                            "peso_seco": peso_seco,
                            "aux_silo": item.aux_silo,
                        },
                    ))
                    restante -= tomar

    avance = RegistroAvanceInyeccion(
        plan_inyeccion_id=item.id,
        tiempo_ciclo=data.tiempo_ciclo,
        contador_hora=data.contador_hora,
        produccion_total=piezas_reales,
    )
    db.add(avance)

    await db.commit()
    await db.refresh(item)
    return item


@router.post("/{id}/paro", response_model=PlanInyeccionResponse)
async def registrar_paro(
    id: int,
    data: PlanInyeccionParoRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    if item.status != "En Proceso":
        raise HTTPException(status_code=400, detail="La orden no está en proceso")
    if item.en_paro:
        raise HTTPException(status_code=400, detail="Ya está en paro")

    ahora = _ahora_utc()
    # Acumular tiempo productivo transcurrido
    if item.hora_ultimo_inicio:
        delta = (ahora - item.hora_ultimo_inicio).total_seconds()
        item.tiempo_acumulado_seg += int(delta)

    item.en_paro = True
    item.hora_ultimo_inicio = None

    # FIX: Crear NUEVA lista para que SQLAlchemy detecte cambios en JSON
    paro_nuevo = {
        "id": str(uuid.uuid4())[:8],
        "motivo": data.motivo or "",
        "motivo_mantenimiento": data.motivo_mantenimiento,
        "comentarios": data.comentarios or "",
        "inicio": ahora.isoformat(),
        "fin": None,
        "duracion_segundos": 0,
        "status": "Activo",
    }
    item.paros = (item.paros or []) + [paro_nuevo]

    await db.commit()
    await db.refresh(item)
    return item


@router.post("/{id}/reanudar", response_model=PlanInyeccionResponse)
async def reanudar_orden(
    id: int,
    data: PlanInyeccionReanudarRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    if not item.en_paro:
        raise HTTPException(status_code=400, detail="No está en paro")

    ahora = _ahora_utc()
    item.en_paro = False
    item.hora_ultimo_inicio = ahora

    # FIX CRÍTICO: Crear NUEVA lista de paros para que SQLAlchemy detecte cambios en JSON
    paros_actualizados = []
    for p in (item.paros or []):
        # Crear copia profunda del paro
        paro_copia = dict(p)

        if paro_copia.get("status") == "Activo":
            inicio = datetime.fromisoformat(paro_copia["inicio"])
            duracion = int((ahora - inicio).total_seconds())
            paro_copia["fin"] = ahora.isoformat()
            paro_copia["duracion_segundos"] = duracion
            paro_copia["status"] = "Finalizado"
            # Actualizar con los datos enviados al reanudar
            paro_copia["motivo"] = data.motivo
            paro_copia["motivo_mantenimiento"] = data.motivo_mantenimiento
            paro_copia["comentarios"] = data.comentarios

        paros_actualizados.append(paro_copia)

    # Asignar nueva lista para triggerar cambio en SQLAlchemy
    item.paros = paros_actualizados

    await db.commit()
    await db.refresh(item)
    return item


@router.post("/{id}/finalizar", response_model=PlanInyeccionFinalizarResponse)
async def finalizar_orden(
    id: int,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    if item.status != "En Proceso":
        raise HTTPException(status_code=400, detail="La orden no está en proceso")

    ahora = _ahora_utc()

    # Acumular tiempo final si no está en paro
    if not item.en_paro and item.hora_ultimo_inicio:
        delta = (ahora - item.hora_ultimo_inicio).total_seconds()
        item.tiempo_acumulado_seg += int(delta)
    # Si está en paro, cerrar el paro activo primero
    elif item.en_paro:
        # FIX: Crear NUEVA lista de paros
        paros_actualizados = []
        for p in (item.paros or []):
            paro_copia = dict(p)
            if paro_copia.get("status") == "Activo":
                inicio = datetime.fromisoformat(paro_copia["inicio"])
                duracion = int((ahora - inicio).total_seconds())
                paro_copia["fin"] = ahora.isoformat()
                paro_copia["duracion_segundos"] = duracion
                paro_copia["status"] = "Finalizado"
                # Si no tiene motivo, marcar como finalización
                if not paro_copia.get("motivo"):
                    paro_copia["motivo"] = "Finalización de Orden"
                    paro_copia["comentarios"] = "Orden finalizada estando en paro"
            paros_actualizados.append(paro_copia)
        item.paros = paros_actualizados

    item.status = "Finalizado"
    item.hora_fin = ahora
    item.hora_ultimo_inicio = None
    item.en_paro = False

    # FIX: Commit antes de buscar siguiente
    await db.commit()
    await db.refresh(item)

    # Auto-iniciar siguiente
    siguiente = None
    res_next = await db.execute(
        select(PlanInyeccion).where(
            and_(
                PlanInyeccion.maquina == item.maquina,
                PlanInyeccion.prioridad == item.prioridad,
                PlanInyeccion.orden_secuencia > item.orden_secuencia,
                PlanInyeccion.status == "Pendiente",
            )
        ).order_by(PlanInyeccion.orden_secuencia.asc()).limit(1)
    )
    next_item = res_next.scalar_one_or_none()

    if next_item:
        next_item.status = "En Proceso"
        next_item.hora_inicio = ahora
        next_item.hora_ultimo_inicio = ahora
        next_item.en_paro = False  # FIX: Siguiente inicia produciendo
        await db.commit()
        await db.refresh(next_item)
        siguiente = PlanInyeccionResponse.model_validate(next_item)

    return PlanInyeccionFinalizarResponse(
        message=f"Orden {item.numero_parte} finalizada",
        finalizado_id=item.id,
        siguiente_iniciado=siguiente,
    )


# ══════════════════════════════════════════════════════════════════
# REPORTES
# ══════════════════════════════════════════════════════════════════

def _calc_tiempo_trabajo(orden):
    """Calcula tiempo de trabajo incluyendo tiempo actual si está activa"""
    total = orden.tiempo_acumulado_seg or 0
    if not orden.en_paro and orden.hora_ultimo_inicio:
        ahora = datetime.utcnow()
        try:
            inicio = orden.hora_ultimo_inicio
            if isinstance(inicio, str):
                inicio = datetime.fromisoformat(inicio.replace('Z', '+00:00').replace('+00:00', ''))
            delta = (ahora - inicio).total_seconds()
            if delta > 0:
                total += int(delta)
        except:
            pass
    return total


async def _reporte_general_data(fecha: str, turno: Optional[str], db: AsyncSession):
    """Helper interno para obtener datos del reporte general"""
    from collections import Counter

    fecha_date = datetime.strptime(fecha, "%Y-%m-%d").date()

    if turno == "DIA":
        hora_inicio = datetime.combine(fecha_date, datetime.min.time().replace(hour=7, minute=30))
        hora_fin = datetime.combine(fecha_date, datetime.min.time().replace(hour=19, minute=30))
    elif turno == "NOCHE":
        hora_inicio = datetime.combine(fecha_date, datetime.min.time().replace(hour=19, minute=30))
        hora_fin = datetime.combine(fecha_date + timedelta(days=1), datetime.min.time().replace(hour=7, minute=30))
    else:
        hora_inicio = datetime.combine(fecha_date, datetime.min.time().replace(hour=7, minute=30))
        hora_fin = datetime.combine(fecha_date + timedelta(days=1), datetime.min.time().replace(hour=7, minute=30))

    # Convertir rangos del filtro de GMT-6 a UTC (BD almacena UTC naive)
    hora_inicio = _local_naive_to_utc_naive(hora_inicio)
    hora_fin = _local_naive_to_utc_naive(hora_fin)

    q = select(PlanInyeccion).where(
        PlanInyeccion.hora_inicio >= hora_inicio,
        PlanInyeccion.hora_inicio < hora_fin,
    ).order_by(PlanInyeccion.maquina, PlanInyeccion.prioridad)

    result = await db.execute(q)
    ordenes = result.scalars().all()

    reporte = []
    for orden in ordenes:
        avances = await db.execute(
            select(RegistroAvanceInyeccion).where(
                RegistroAvanceInyeccion.plan_inyeccion_id == orden.id
            ).order_by(RegistroAvanceInyeccion.timestamp)
        )
        avances_list = avances.scalars().all()

        if avances_list:
            tiempos_ciclo = [a.tiempo_ciclo for a in avances_list]
            ciclo_moda = Counter(tiempos_ciclo).most_common(1)[0][0]
        else:
            ciclo_moda = None

        produccion_total_orden = sum(a.produccion_total for a in avances_list)

        paros_finalizados = [p for p in (orden.paros or []) 
                            if p.get("status") == "Finalizado" 
                            and p.get("duracion_segundos", 0) > 0]

        motivo_paro = ", ".join([p.get("motivo") for p in paros_finalizados if p.get("motivo")]) if paros_finalizados else None
        motivo_mantenimiento = ", ".join([p.get("motivo_mantenimiento") for p in paros_finalizados if p.get("motivo_mantenimiento")]) if paros_finalizados else None
        tiempo_paro = sum(p.get("duracion_segundos", 0) for p in paros_finalizados)
        comentarios_paro = ", ".join([p.get("comentarios") for p in paros_finalizados if p.get("comentarios")]) if paros_finalizados else None

        q_anterior = select(PlanInyeccion).where(
            PlanInyeccion.maquina == orden.maquina,
            PlanInyeccion.hora_inicio < orden.hora_inicio,
            PlanInyeccion.status == "Finalizado",
        ).order_by(PlanInyeccion.hora_fin.desc()).limit(1)
        res_anterior = await db.execute(q_anterior)
        orden_anterior = res_anterior.scalar_one_or_none()
        parte_anterior = orden_anterior.numero_parte if orden_anterior else None

        percent_prod = round((orden.piezas_producidas / orden.plan_piezas * 100), 1) if orden.plan_piezas > 0 else 0

        tiempo_trabajo = _calc_tiempo_trabajo(orden)

        # FIX: Calcular turno real usando GMT-6
        turno_real = _get_turno_real(orden.hora_inicio)
        lote = await _construir_lote(db, orden.numero_parte, turno_real, fecha, orden.maquina)

        reporte.append({
             "lote": lote,
             "fecha": fecha,
             "turno": turno or "AMBOS",
             "turno_real": turno_real,
             "numero_parte": orden.numero_parte,
             "maquina": orden.maquina,
             "cav": orden.cav,
             "ciclo": ciclo_moda,
             "tiempo_trabajo": tiempo_trabajo,
             "meta_plan": orden.plan_piezas,
             "produccion_total": produccion_total_orden,
             "percent_prod": percent_prod,
             "motivo_paro": motivo_paro,
             "motivo_mantenimiento": motivo_mantenimiento,
             "tiempo_paro": tiempo_paro,
             "comentarios": comentarios_paro,
             "parte_anterior": parte_anterior,
             "orden_id": orden.id,
         })

    return reporte


@router.get("/reporte-general")
async def reporte_general(
    fecha: str,
    turno: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """
    Reporte general de inyección por fecha y turno.
    Retorna lista de órdenes con datos para el reporte.
    """
    return await _reporte_general_data(fecha, turno, db)


@router.get("/reporte-general/excel")
async def reporte_general_excel(
    fecha: str,
    turno: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """
    Descarga Excel del reporte general de inyección por fecha y turno.
    Estilo profesional con colores azules.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    reporte_data = await _reporte_general_data(fecha, turno, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Inyección"

    # Colores del tema
    AZUL_OSCURO = "1F4E78"
    AZUL_MEDIO = "2E75B6"
    AZUL_CLARO = "D6E3F8"
    GRIS_CLARO = "F2F2F2"
    BLANCO = "FFFFFF"
    NEGRO = "000000"
    VERDE = "27AE60"
    ROJO = "E74C3C"
    AMARILLO = "F1C40F"

    # Bordes
    thin_border = Border(
        left=Side(style='thin', color=AZUL_OSCURO),
        right=Side(style='thin', color=AZUL_OSCURO),
        top=Side(style='thin', color=AZUL_OSCURO),
        bottom=Side(style='thin', color=AZUL_OSCURO)
    )

    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:P1')
    titulo_cell = ws['A1']
    titulo_cell.value = "REPORTE DE INYECCIÓN"
    titulo_cell.font = Font(name='Calibri', size=18, bold=True, color=BLANCO)
    titulo_cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # SUBTÍTULO (Fecha y Turno)
    ws.merge_cells('A2:P2')
    subtitulo_cell = ws['A2']
    subtitulo_cell.value = f"Fecha: {fecha}"
    subtitulo_cell.font = Font(name='Calibri', size=11, bold=True, color=AZUL_OSCURO)
    subtitulo_cell.fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
    subtitulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # ENCABEZADOS (fila 4)
    headers = [
        "Lote", "Fecha", "Turno", "No. Parte", "Máquina", "Cav", "Ciclo",
        "Tiempo Trabajo", "Meta Plan", "Producción Total", "% Prod",
        "Motivo Paro", "Motivo Mantenimiento", "Tiempo Paro", "Comentarios", "C/M"
    ]

    header_font = Font(name='Calibri', size=10, bold=True, color=BLANCO)
    header_fill = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment

    ws.row_dimensions[4].height = 30

    # DATOS
    data_font = Font(name='Calibri', size=10, color=NEGRO)
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    def fmt_time(segundos):
        if not segundos:
            return "00:00:00"
        h = segundos // 3600
        m = (segundos % 3600) // 60
        s = segundos % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    for row_idx, item in enumerate(reporte_data, start=5):
        # Alternar colores de fondo
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
        else:
            row_fill = PatternFill(start_color=BLANCO, end_color=BLANCO, fill_type="solid")

        # FIX: Usar turno_real en lugar de turno
        turno_valor = item.get("turno_real") or item.get("turno") or ""

        # Lote
        cell = ws.cell(row=row_idx, column=1, value=item.get("lote", ""))
        cell.font = Font(name='Calibri', size=9, color=NEGRO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_left

        # Fecha
        cell = ws.cell(row=row_idx, column=2, value=item.get("fecha", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Turno (FIX: usar turno_real)
        cell = ws.cell(row=row_idx, column=3, value=turno_valor)
        cell.font = Font(name='Calibri', size=10, bold=True, color=AZUL_OSCURO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # No. Parte
        cell = ws.cell(row=row_idx, column=4, value=item.get("numero_parte", ""))
        cell.font = Font(name='Calibri', size=10, bold=True, color=NEGRO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Máquina
        cell = ws.cell(row=row_idx, column=5, value=item.get("maquina", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Cav
        cell = ws.cell(row=row_idx, column=6, value=item.get("cav", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Ciclo
        cell = ws.cell(row=row_idx, column=7, value=item.get("ciclo", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Tiempo Trabajo
        cell = ws.cell(row=row_idx, column=8, value=fmt_time(item.get("tiempo_trabajo")))
        cell.font = Font(name='Calibri', size=10, color=AZUL_OSCURO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Meta Plan
        cell = ws.cell(row=row_idx, column=9, value=item.get("meta_plan", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_right
        cell.number_format = '#,##0'

        # Producción Total
        cell = ws.cell(row=row_idx, column=10, value=item.get("produccion_total", ""))
        cell.font = Font(name='Calibri', size=10, bold=True, color=VERDE)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_right
        cell.number_format = '#,##0'

        # % Prod
        pct = item.get("percent_prod", 0)
        cell = ws.cell(row=row_idx, column=11, value=f"{pct}%")
        # Color según porcentaje
        if pct >= 100:
            pct_color = VERDE
        elif pct >= 80:
            pct_color = AMARILLO
        else:
            pct_color = ROJO
        cell.font = Font(name='Calibri', size=10, bold=True, color=pct_color)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Motivo Paro
        cell = ws.cell(row=row_idx, column=12, value=item.get("motivo_paro", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_left

        # Motivo Mantenimiento
        cell = ws.cell(row=row_idx, column=13, value=item.get("motivo_mantenimiento", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_left

        # Tiempo Paro
        cell = ws.cell(row=row_idx, column=14, value=fmt_time(item.get("tiempo_paro")))
        cell.font = Font(name='Calibri', size=10, color=ROJO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Comentarios
        cell = ws.cell(row=row_idx, column=15, value=item.get("comentarios", ""))
        cell.font = Font(name='Calibri', size=9, color=NEGRO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_left

        # C/M (Parte Anterior)
        cell = ws.cell(row=row_idx, column=16, value=item.get("parte_anterior", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

    # Ajustar anchos de columna
    column_widths = {
        'A': 18,  # Lote
        'B': 12,  # Fecha
        'C': 8,   # Turno
        'D': 14,  # No. Parte
        'E': 10,  # Máquina
        'F': 6,   # Cav
        'G': 8,   # Ciclo
        'H': 14,  # Tiempo Trabajo
        'I': 12,  # Meta Plan
        'J': 16,  # Producción Total
        'K': 8,   # % Prod
        'L': 20,  # Motivo Paro
        'M': 20,  # Motivo Mantenimiento
        'N': 12,  # Tiempo Paro
        'O': 25,  # Comentarios
        'P': 12,  # C/M
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Congelar paneles (encabezados visibles al scroll)
    ws.freeze_panes = 'A5'

    # Filtros automáticos
    ws.auto_filter.ref = f"A4:P{4 + len(reporte_data)}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from fastapi.responses import StreamingResponse
    filename = f"reporte_inyeccion_{fecha}{'_' + turno if turno else ''}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/reporte-individual/{id}")
async def reporte_individual(
    id: int,
    db: AsyncSession = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """
    Reporte individual de una orden de inyección.
    Retorna datos para generar Excel con bitácora de producción.
    """
    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    orden = res.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    res_prod = await db.execute(select(Producto).where(Producto.sku == orden.numero_parte))
    producto = res_prod.scalar_one_or_none()

    res_avances = await db.execute(
        select(RegistroAvanceInyeccion).where(
            RegistroAvanceInyeccion.plan_inyeccion_id == orden.id
        ).order_by(RegistroAvanceInyeccion.timestamp)
    )
    avances = res_avances.scalars().all()

    if avances:
        tiempos_ciclo = [a.tiempo_ciclo for a in avances]
        from collections import Counter
        ciclo_real = Counter(tiempos_ciclo).most_common(1)[0][0]
    else:
        ciclo_real = None

    ciclo_teorico = None
    if producto and producto.caracteristicas_inyeccion:
        ciclo_teorico = producto.caracteristicas_inyeccion.get("ciclo")

    descripcion = producto.descripcion if producto else ""
    modelo = producto.modelo if producto else ""

    # FIX: Usar _get_turno_real en lugar de comparación UTC directa
    turno = _get_turno_real(orden.hora_inicio)

    return {
        "fecha": orden.hora_inicio.strftime("%Y-%m-%d") if orden.hora_inicio else "",
        "turno": turno,
        "numero_parte": orden.numero_parte,
        "descripcion": descripcion,
        "modelo": modelo,
        "maquina": orden.maquina,
        "contador_total": sum(a.produccion_total for a in avances),
        "ciclo_teorico": ciclo_teorico,
        "ciclo_real": ciclo_real,
        "total_prod": orden.piezas_producidas,
        "cav": orden.cav,
        "avances": [
            {
                "timestamp": a.timestamp.isoformat() if a.timestamp else None,
                "tiempo_ciclo": a.tiempo_ciclo,
                "contador_hora": a.contador_hora,
                "produccion_total": a.produccion_total,
            }
            for a in avances
        ],
    }


@router.get("/reporte-individual/{id}/excel")
async def reporte_individual_excel(
    id: int,
    db: AsyncSession = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """
    Descarga Excel del reporte individual de una orden de inyección.
    Estilo profesional con colores azules (igual que reporte general).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from datetime import datetime, timedelta
    from collections import Counter

    res = await db.execute(select(PlanInyeccion).where(PlanInyeccion.id == id))
    orden = res.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    res_prod = await db.execute(select(Producto).where(Producto.sku == orden.numero_parte))
    producto = res_prod.scalar_one_or_none()

    res_avances = await db.execute(
        select(RegistroAvanceInyeccion).where(
            RegistroAvanceInyeccion.plan_inyeccion_id == orden.id
        ).order_by(RegistroAvanceInyeccion.timestamp)
    )
    avances = res_avances.scalars().all()

    # Calcular ciclo real (moda de tiempos de ciclo)
    if avances:
        tiempos_ciclo = [a.tiempo_ciclo for a in avances]
        ciclo_real = Counter(tiempos_ciclo).most_common(1)[0][0]
    else:
        ciclo_real = None

    ciclo_teorico = None
    if producto and producto.caracteristicas_inyeccion:
        ciclo_teorico = producto.caracteristicas_inyeccion.get("ciclo")

    # Calcular tiempo de trabajo total
    tiempo_trabajo = _calc_tiempo_trabajo(orden)

    # Calcular paros finalizados
    paros_finalizados = [p for p in (orden.paros or []) 
                        if p.get("status") == "Finalizado" 
                        and p.get("duracion_segundos", 0) > 0]
    tiempo_paro_total = sum(p.get("duracion_segundos", 0) for p in paros_finalizados)

    # Calcular producción total
    produccion_total = sum(a.produccion_total for a in avances)

    # Calcular % producción
    percent_prod = round((orden.piezas_producidas / orden.plan_piezas * 100), 1) if orden.plan_piezas > 0 else 0

    # Determinar turno
    turno_real = _get_turno_real(orden.hora_inicio)
    fecha_str = orden.hora_inicio.strftime("%Y-%m-%d") if orden.hora_inicio else ""
    lote = await _construir_lote(db, orden.numero_parte, turno_real, fecha_str, orden.maquina)

    # ═════════════════════════════════════════════════════════════════
    # CREAR EXCEL CON ESTILOS PROFESIONALES (mismo tema que reporte general)
    # ═════════════════════════════════════════════════════════════════

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Individual"

    # Colores del tema (IGUALES al reporte general)
    AZUL_OSCURO = "1F4E78"
    AZUL_MEDIO = "2E75B6"
    AZUL_CLARO = "D6E3F8"
    GRIS_CLARO = "F2F2F2"
    BLANCO = "FFFFFF"
    NEGRO = "000000"
    VERDE = "27AE60"
    ROJO = "E74C3C"
    AMARILLO = "F1C40F"

    # Bordes
    thin_border = Border(
        left=Side(style='thin', color=AZUL_OSCURO),
        right=Side(style='thin', color=AZUL_OSCURO),
        top=Side(style='thin', color=AZUL_OSCURO),
        bottom=Side(style='thin', color=AZUL_OSCURO)
    )

    # Fuentes comunes
    header_font = Font(name='Calibri', size=10, bold=True, color=BLANCO)
    header_fill = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=10, color=NEGRO)
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    def fmt_time(segundos):
        if not segundos:
            return "00:00:00"
        h = segundos // 3600
        m = (segundos % 3600) // 60
        s = segundos % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    # ═════════════════════════════════════════════════════════════════
    # TÍTULO PRINCIPAL
    # ═════════════════════════════════════════════════════════════════
    ws.merge_cells('A1:F1')
    titulo_cell = ws['A1']
    titulo_cell.value = "REPORTE INDIVIDUAL DE INYECCIÓN"
    titulo_cell.font = Font(name='Calibri', size=18, bold=True, color=BLANCO)
    titulo_cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # ═════════════════════════════════════════════════════════════════
    # SUBTÍTULO (No. Parte y Máquina)
    # ═════════════════════════════════════════════════════════════════
    ws.merge_cells('A2:F2')
    subtitulo_cell = ws['A2']
    subtitulo_cell.value = f"Parte: {orden.numero_parte}  |  Máquina: {orden.maquina}  |  Lote: {lote}"
    subtitulo_cell.font = Font(name='Calibri', size=11, bold=True, color=AZUL_OSCURO)
    subtitulo_cell.fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
    subtitulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # ═════════════════════════════════════════════════════════════════
    # DATOS GENERALES (fila 4 en adelante)
    # ═════════════════════════════════════════════════════════════════
    datos_generales = [
        ["Fecha", orden.hora_inicio.strftime("%Y-%m-%d") if orden.hora_inicio else "—", "Turno", turno, "Cav", orden.cav],
        ["Descripción", producto.descripcion if producto else "—", "", "", "", ""],
        ["Modelo", producto.modelo if producto else "—", "", "", "", ""],
        ["Ciclo Teórico", ciclo_teorico if ciclo_teorico else "—", "Ciclo Real", ciclo_real if ciclo_real else "—", "", ""],
        ["Meta Plan", orden.plan_piezas, "Producción Total", produccion_total, "% Prod", f"{percent_prod}%"],
        ["Tiempo Trabajo", fmt_time(tiempo_trabajo), "Tiempo Paro", fmt_time(tiempo_paro_total), "", ""],
    ]

    # Ajustar anchos de columna para datos generales
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    for row_idx, fila in enumerate(datos_generales, start=4):
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor if valor != "" else None)
            cell.border = thin_border
            
            if col_idx % 2 == 1:  # Columnas de etiquetas (A, C, E)
                cell.font = Font(name='Calibri', size=10, bold=True, color=BLANCO)
                cell.fill = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
                cell.alignment = data_alignment_center
            else:  # Columnas de valores (B, D, F)
                cell.font = data_font
                cell.fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
                cell.alignment = data_alignment_left if col_idx == 2 else data_alignment_center

    # Merge para campos largos (Descripción y Modelo)
    ws.merge_cells('B5:F5')  # Descripción
    ws.merge_cells('B6:F6')  # Modelo

    # ═════════════════════════════════════════════════════════════════
    # SECCIÓN: BITÁCORA DE PRODUCCIÓN (fila 11)
    # ═════════════════════════════════════════════════════════════════
    ws.merge_cells('A11:F11')
    seccion_cell = ws['A11']
    seccion_cell.value = "BITÁCORA DE PRODUCCIÓN"
    seccion_cell.font = Font(name='Calibri', size=14, bold=True, color=BLANCO)
    seccion_cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    seccion_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 30

    # Encabezados de bitácora (fila 13)
    headers_bitacora = ["Hora", "Tiempo Ciclo", "Contador por Hora", "Contador Acumulado", "Cantidad Total", "Producción Total"]

    for col_idx, header in enumerate(headers_bitacora, start=1):
        cell = ws.cell(row=13, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment

    ws.row_dimensions[13].height = 30

    # ═════════════════════════════════════════════════════════════════
    # DATOS DE AVANCES
    # ═════════════════════════════════════════════════════════════════
    # Convertir hora de inicio a GMT-6 para determinar turno real
    hora_inicio_local = _utc_to_local(orden.hora_inicio) if orden.hora_inicio else None
    turno = _get_turno_real(orden.hora_inicio)
    
    horas_turno = [
        ("07:30", "08:30"), ("08:30", "09:30"), ("09:30", "10:30"), ("10:30", "11:30"),
        ("11:30", "12:30"), ("12:30", "13:30"), ("13:30", "14:30"), ("14:30", "15:30"),
        ("15:30", "16:30"), ("16:30", "17:30"), ("17:30", "18:30"), ("18:30", "19:30")
    ]

    # Si es turno noche, ajustar horas
    if turno == "NOCHE":
        horas_turno = [
            ("19:30", "20:30"), ("20:30", "21:30"), ("21:30", "22:30"), ("22:30", "23:30"),
            ("23:30", "00:30"), ("00:30", "01:30"), ("01:30", "02:30"), ("02:30", "03:30"),
            ("03:30", "04:30"), ("04:30", "05:30"), ("05:30", "06:30"), ("06:30", "07:30")
        ]

    # Crear filas base para cada hora del turno
    datos_hora = {i: {"tiempo_ciclo": None, "contador_hora": 0, "contador_acumulado": 0, "cantidad_total": 0} 
                  for i in range(len(horas_turno))}

    # Asignar avances a la hora correspondiente
    for avance in avances:
        if not avance.timestamp:
            continue
        
        avance_local = _utc_to_local(avance.timestamp)
        hora = avance_local.hour
        minuto = avance_local.minute
        
        # Determinar índice de la hora
        idx = None
        for i, (inicio, fin) in enumerate(horas_turno):
            h_inicio = int(inicio.split(':')[0])
            m_inicio = int(inicio.split(':')[1])
            h_fin = int(fin.split(':')[0])
            m_fin = int(fin.split(':')[1])

            # Convertir todo a minutos para comparación precisa
            inicio_min = h_inicio * 60 + m_inicio
            fin_min = h_fin * 60 + m_fin
            avance_min = hora * 60 + minuto
            
            # Manejar cruce de medianoche (turno noche)
            if fin_min < inicio_min:  # Cruza medianoche
                if avance_min >= inicio_min or avance_min < fin_min:
                    idx = i
                    break
            else:
                if avance_min >= inicio_min and avance_min < fin_min:
                    idx = i
                    break
        
        if idx is not None:
            datos_hora[idx]["tiempo_ciclo"] = avance.tiempo_ciclo
            datos_hora[idx]["contador_hora"] = avance.contador_hora
            datos_hora[idx]["cantidad_total"] = avance.produccion_total

    # Calcular acumulados
    acum = 0
    for i in range(len(horas_turno)):
        acum += datos_hora[i]["contador_hora"]
        datos_hora[i]["contador_acumulado"] = acum

    # Escribir datos en el Excel
    produccion_acumulada = 0
    for row_idx, (i, (hora_inicio, hora_fin)) in enumerate(zip(range(len(horas_turno)), horas_turno), start=14):
        dato = datos_hora[i]
        
        # Alternar colores de fondo
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
        else:
            row_fill = PatternFill(start_color=BLANCO, end_color=BLANCO, fill_type="solid")

        # Hora
        cell = ws.cell(row=row_idx, column=1, value=f"{hora_inicio} - {hora_fin}")
        cell.font = Font(name='Calibri', size=10, bold=True, color=AZUL_OSCURO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Tiempo Ciclo
        cell = ws.cell(row=row_idx, column=2, value=dato["tiempo_ciclo"] if dato["tiempo_ciclo"] else "—")
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center

        # Contador por Hora
        cell = ws.cell(row=row_idx, column=3, value=dato["contador_hora"] if dato["contador_hora"] > 0 else "—")
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center
        if dato["contador_hora"] > 0:
            cell.number_format = '#,##0'

        # Contador Acumulado
        cell = ws.cell(row=row_idx, column=4, value=dato["contador_acumulado"] if dato["contador_acumulado"] > 0 else "—")
        cell.font = Font(name='Calibri', size=10, bold=True, color=AZUL_OSCURO)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center
        if dato["contador_acumulado"] > 0:
            cell.number_format = '#,##0'

        # Cantidad Total (contador × cav)
        cantidad = dato["contador_hora"] * (orden.cav or 1)
        cell = ws.cell(row=row_idx, column=5, value=cantidad if cantidad > 0 else "—")
        cell.font = Font(name='Calibri', size=10, bold=True, color=VERDE)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = data_alignment_center
        if cantidad > 0:
            cell.number_format = '#,##0'

        # Producción Total (acumulado de Cantidad Total)
        if cantidad > 0:
            produccion_acumulada += cantidad
            cell = ws.cell(row=row_idx, column=6, value=produccion_acumulada)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = data_alignment_center
            cell.number_format = '#,##0'
        else:
            cell = ws.cell(row=row_idx, column=6, value="—")
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = data_alignment_center

    # ═════════════════════════════════════════════════════════════════
    # SECCIÓN: RESUMEN DE PAROS (si hay paros)
    # ═════════════════════════════════════════════════════════════════
    if paros_finalizados:
        fila_paro = 14 + len(horas_turno) + 2  # 2 espacios después de la bitácora
        
        ws.merge_cells(f'A{fila_paro}:F{fila_paro}')
        seccion_paro = ws[f'A{fila_paro}']
        seccion_paro.value = "RESUMEN DE PAROS"
        seccion_paro.font = Font(name='Calibri', size=14, bold=True, color=BLANCO)
        seccion_paro.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
        seccion_paro.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[fila_paro].height = 30

        # Encabezados de paros
        fila_header_paro = fila_paro + 2
        headers_paro = ["Motivo", "Motivo Mantenimiento", "Inicio", "Fin", "Duración", "Comentarios"]
        
        for col_idx, header in enumerate(headers_paro, start=1):
            cell = ws.cell(row=fila_header_paro, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment

        # Datos de paros
        for idx, paro in enumerate(paros_finalizados):
            row_idx = fila_header_paro + 1 + idx
            if row_idx % 2 == 0:
                row_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
            else:
                row_fill = PatternFill(start_color=BLANCO, end_color=BLANCO, fill_type="solid")

            ws.cell(row=row_idx, column=1, value=paro.get("motivo", "—")).font = data_font
            ws.cell(row=row_idx, column=1).fill = row_fill
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = data_alignment_left

            ws.cell(row=row_idx, column=2, value=paro.get("motivo_mantenimiento") or "—").font = data_font
            ws.cell(row=row_idx, column=2).fill = row_fill
            ws.cell(row=row_idx, column=2).border = thin_border
            ws.cell(row=row_idx, column=2).alignment = data_alignment_left

            inicio_str = paro.get("inicio", "—")
            if inicio_str and inicio_str != "—":
                try:
                    dt = datetime.fromisoformat(inicio_str.replace('Z', '+00:00'))
                    inicio_str = dt.strftime("%H:%M:%S")
                except:
                    pass
            ws.cell(row=row_idx, column=3, value=inicio_str).font = data_font
            ws.cell(row=row_idx, column=3).fill = row_fill
            ws.cell(row=row_idx, column=3).border = thin_border
            ws.cell(row=row_idx, column=3).alignment = data_alignment_center

            fin_str = paro.get("fin", "—")
            if fin_str and fin_str != "—":
                try:
                    dt = datetime.fromisoformat(fin_str.replace('Z', '+00:00'))
                    fin_str = dt.strftime("%H:%M:%S")
                except:
                    pass
            ws.cell(row=row_idx, column=4, value=fin_str).font = data_font
            ws.cell(row=row_idx, column=4).fill = row_fill
            ws.cell(row=row_idx, column=4).border = thin_border
            ws.cell(row=row_idx, column=4).alignment = data_alignment_center

            duracion = paro.get("duracion_segundos", 0)
            ws.cell(row=row_idx, column=5, value=fmt_time(duracion)).font = Font(name='Calibri', size=10, color=ROJO)
            ws.cell(row=row_idx, column=5).fill = row_fill
            ws.cell(row=row_idx, column=5).border = thin_border
            ws.cell(row=row_idx, column=5).alignment = data_alignment_center

            ws.cell(row=row_idx, column=6, value=paro.get("comentarios") or "—").font = data_font
            ws.cell(row=row_idx, column=6).fill = row_fill
            ws.cell(row=row_idx, column=6).border = thin_border
            ws.cell(row=row_idx, column=6).alignment = data_alignment_left

    # Congelar paneles
    ws.freeze_panes = 'A14'

    # Filtros automáticos en bitácora
    ws.auto_filter.ref = f"A13:F{13 + len(horas_turno)}"

    # Guardar y retornar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from fastapi.responses import StreamingResponse
    filename = f"reporte_individual_{orden.numero_parte}_{orden.maquina}_{orden.hora_inicio.strftime('%Y%m%d') if orden.hora_inicio else 'ND'}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )