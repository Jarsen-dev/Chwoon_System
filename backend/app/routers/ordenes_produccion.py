"""
Router para Órdenes de Producción unificadas:
PRE-EXPANSION, INYECCION, ASSY
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, or_
from typing import List, Optional
from datetime import datetime, timedelta, timezone
import math
import io

from app.database import AsyncSessionLocal
from app.models.orden_compra import OrdenCompra as OCModel, OrdenCompraItem as OCItemModel
from app.models.lote_inventario import LoteInventario, MovimientoLote
from app.models.orden_produccion import OrdenProduccion
from app.models.suministro_silo import SuministroSilo
from app.models.producto import Producto
from app.models.ubicacion import Ubicacion
from app.models.usuario import Usuario
from app.core.deps import get_current_user
from app.schemas.ordenes_produccion import (
    OrdenProduccionResponse,
    OrdenUnificadaResponse,
    IniciarPreExpansionRequest,
    RegistroParciallPreExpRequest,
    RegistroDatosProcesoRequest,
    FinalizarPreExpansionRequest,
    IniciarInyeccionRequest,
    RegistrarPiezaRequest,
    FinalizarInyeccionRequest,
    IniciarAssyRequest,
    FinalizarAssyRequest,
    RegistrarParoRequest,
    CrearSuministroRequest,
    SuministroSiloResponse,
    EstadoSiloResponse,
)

router = APIRouter(prefix="/ordenes-produccion", tags=["ordenes-produccion"])

TZ_LOCAL = timezone(timedelta(hours=-6))


def _get_turno_from_dt(dt: datetime) -> str:
    """Calcula turno DIA/NOCHE desde datetime UTC naive"""
    if dt.tzinfo:
        local = dt.astimezone(TZ_LOCAL)
    else:
        local = dt.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
    h, m = local.hour, local.minute
    total_min = h * 60 + m
    if 450 <= total_min < 1170:  # 07:30 → 19:30
        return "DIA"
    return "NOCHE"


def _get_fecha_turno(dt: datetime) -> str:
    """Fecha del turno (noche usa fecha de inicio del turno)"""
    if dt.tzinfo:
        local = dt.astimezone(TZ_LOCAL)
    else:
        local = dt.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
    h, m = local.hour, local.minute
    total_min = h * 60 + m
    if total_min < 450:  # Antes de 07:30 → turno noche del día anterior
        local = local - timedelta(days=1)
    return local.strftime("%Y-%m-%d")


async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


# ── Helpers ──────────────────────────────────────────────────────

async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


def _ahora_local() -> datetime:
    return datetime.now(TZ_LOCAL)


def _require_prod_role(user: Usuario):
    if user.rol.value not in ("admin", "supervisor", "operador"):
        raise HTTPException(status_code=403, detail="Sin permisos para órdenes de producción")


async def _generar_op_id(db: AsyncSession, sku: str, clase: str) -> str:
    ahora = _ahora_local()
    inicio_dia_local = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    inicio_dia_utc = inicio_dia_local.astimezone(timezone.utc).replace(tzinfo=None)

    result = await db.execute(
        select(func.count(OrdenProduccion.id)).where(
            OrdenProduccion.fecha_inicio >= inicio_dia_utc
        )
    )
    num_hoy = (result.scalar() or 0) + 1
    numero_trabajo = f"{num_hoy:02d}"
    sufijo_sku = sku[-4:] if len(sku) >= 4 else sku
    fecha_str = ahora.strftime("%d%m%y")

    prefijo = ""
    if clase == "PRE-EXPANSION":
        prefijo = "PE-"
    elif clase == "INYECCION":
        prefijo = "INY-"

    return f"{prefijo}{numero_trabajo}{fecha_str}{sufijo_sku}"


async def _get_producto(db: AsyncSession, sku: str) -> Producto:
    result = await db.execute(select(Producto).where(Producto.sku == sku))
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail=f"Producto '{sku}' no encontrado")
    return producto


async def _consultar_stock_aprobado(db: AsyncSession, sku: str) -> float:
    result = await db.execute(
        select(func.coalesce(func.sum(LoteInventario.cantidad_actual), 0)).where(
            and_(
                LoteInventario.sku_producto == sku,
                LoteInventario.estado_calidad == "Aprobado",
                LoteInventario.cantidad_actual > 0,
            )
        )
    )
    return float(result.scalar() or 0)


async def _consumir_stock_fifo(
    db: AsyncSession,
    sku: str,
    cantidad: float,
    detalles: dict,
    ubicacion_priorizada: str = None,
    excluir_ubicacion_ids: set = None,
) -> list:
    """Consume stock FIFO. Retorna lista de lotes consumidos.
    excluir_ubicacion_ids: set de ubicacion IDs a ignorar (ej: silos)
    """
    if excluir_ubicacion_ids is None:
        excluir_ubicacion_ids = set()

    prioritized_id = None
    if ubicacion_priorizada:
        res = await db.execute(select(Ubicacion).where(Ubicacion.nombre == ubicacion_priorizada))
        ub = res.scalar_one_or_none()
        if ub:
            prioritized_id = ub.id

    # Lotes prioritarios
    lotes_prio = []
    if prioritized_id and prioritized_id not in excluir_ubicacion_ids:
        res = await db.execute(
            select(LoteInventario).where(
                and_(
                    LoteInventario.sku_producto == sku,
                    LoteInventario.estado_calidad == "Aprobado",
                    LoteInventario.ubicacion_id == prioritized_id,
                    LoteInventario.cantidad_actual > 0,
                )
            ).order_by(LoteInventario.fecha_recepcion)
        )
        lotes_prio = list(res.scalars().all())

    ids_prio = {l.id for l in lotes_prio}

    # Lotes generales (excluyendo silos)
    res = await db.execute(
        select(LoteInventario).where(
            and_(
                LoteInventario.sku_producto == sku,
                LoteInventario.estado_calidad == "Aprobado",
                LoteInventario.cantidad_actual > 0,
            )
        ).order_by(LoteInventario.fecha_recepcion)
    )
    all_lotes = res.scalars().all()

    # Filtrar: excluir los que están en ubicaciones de silos y los ya priorizados
    lotes_generales = [
        l for l in all_lotes
        if l.id not in ids_prio
        and (l.ubicacion_id is None or l.ubicacion_id not in excluir_ubicacion_ids)
    ]

    todos = lotes_prio + lotes_generales

    restante = cantidad
    plan = []
    for lote in todos:
        if restante <= 0:
            break
        tomar = min(lote.cantidad_actual, restante)
        plan.append({
            "lote_id": lote.lote_id, "sku_producto": sku,
            "cantidad_consumida": tomar, "oc_origen": lote.oc_origen or "N/A",
        })
        lote.cantidad_actual -= tomar
        db.add(MovimientoLote(
            lote_id=lote.lote_id, fecha=datetime.utcnow(),
            tipo="CONSUMO_PRODUCCION", cantidad=-tomar, detalles=detalles,
        ))
        restante -= tomar

    if restante > 0.001:
        raise HTTPException(
            status_code=400,
            detail=f"Stock insuficiente para {sku}. Faltan {restante:.2f}"
        )
    return plan


async def _crear_lote_produccion(
    db: AsyncSession, sku: str, cantidad: float, op_origen: str,
    ubicacion_nombre: str = None,
) -> str:
    ahora = _ahora_local()
    sufijo = sku[-4:] if len(sku) >= 4 else sku
    lote_id = f"PROD-{ahora.strftime('%d%m%y%H%M%S')}-{sufijo}"

    ubicacion_id = None
    if ubicacion_nombre:
        res = await db.execute(select(Ubicacion).where(Ubicacion.nombre == ubicacion_nombre))
        ub = res.scalar_one_or_none()
        if ub:
            ubicacion_id = ub.id

    lote = LoteInventario(
        lote_id=lote_id, sku_producto=sku, cantidad_actual=cantidad,
        cantidad_inicial=cantidad, ubicacion_id=ubicacion_id,
        fecha_recepcion=datetime.utcnow(), op_origen=op_origen,
        estado_calidad="Aprobado",
    )
    db.add(lote)

    detalles = {"op_origen": op_origen, "sku": sku}
    if ubicacion_nombre:
        detalles["ubicacion_destino"] = ubicacion_nombre

    db.add(MovimientoLote(
        lote_id=lote_id, fecha=datetime.utcnow(),
        tipo="PRODUCCION", cantidad=cantidad, detalles=detalles,
    ))
    return lote_id


async def _get_silos_ubicaciones(db: AsyncSession):
    """Retorna dict con listas de silos principales y AUX"""
    todas = await db.execute(select(Ubicacion))
    ubicaciones = todas.scalars().all()

    silos_padre = None
    for u in ubicaciones:
        if u.nombre.upper() == "SILOS":
            silos_padre = u
            break

    if not silos_padre:
        return [], []

    hijas = [u for u in ubicaciones if u.parent_id == silos_padre.id]
    principales = [u for u in hijas if "AUX" not in u.nombre.upper()]
    auxiliares = [u for u in hijas if "AUX" in u.nombre.upper()]

    return principales, auxiliares

async def _get_silo_ubicacion_ids(db: AsyncSession) -> set:
    """Retorna set de IDs de todas las ubicaciones bajo SILOS (principales + AUX)"""
    todas = await db.execute(select(Ubicacion))
    ubicaciones = todas.scalars().all()

    silos_padre = None
    for u in ubicaciones:
        if u.nombre.upper() == "SILOS":
            silos_padre = u
            break

    if not silos_padre:
        return set()

    # IDs de todas las sub-ubicaciones de SILOS
    ids = {u.id for u in ubicaciones if u.parent_id == silos_padre.id}
    ids.add(silos_padre.id)
    return ids


# ══════════════════════════════════════════════════════════════════
# VISTA UNIFICADA (sin cambios)
# ══════════════════════════════════════════════════════════════════

@router.get("/", response_model=List[OrdenProduccionResponse])
@router.get("", response_model=List[OrdenProduccionResponse])
async def listar_ordenes(
    clase: Optional[str] = None, status: Optional[str] = None,
    activas: Optional[bool] = None, limite: int = 200,
    db: AsyncSession = Depends(get_db),
):
    query = select(OrdenProduccion)
    if clase:
        query = query.where(OrdenProduccion.clase_produccion == clase.upper())
    if status:
        query = query.where(OrdenProduccion.status == status)
    if activas is True:
        query = query.where(OrdenProduccion.status != "Finalizado")
    elif activas is False:
        query = query.where(OrdenProduccion.status == "Finalizado")
    query = query.order_by(OrdenProduccion.fecha_inicio.desc()).limit(limite)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/unificadas")
@router.get("/unificadas/")
async def listar_ordenes_unificadas(
    activas: bool = True, limite: int = 200,
    db: AsyncSession = Depends(get_db),
):
    query = select(OrdenProduccion)
    if activas:
        query = query.where(OrdenProduccion.status != "Finalizado")
    else:
        query = query.where(OrdenProduccion.status == "Finalizado")
    query = query.order_by(OrdenProduccion.fecha_inicio.desc()).limit(limite)
    result = await db.execute(query)
    ordenes = result.scalars().all()

    unificadas = []
    for op in ordenes:
        if op.clase_produccion == "PRE-EXPANSION":
            progreso = f"{op.cantidad_producida:.2f} / {op.cantidad_a_producir:.2f} kg"
        else:
            progreso = f"{int(op.cantidad_producida)} / {int(op.cantidad_a_producir)}"
        unificadas.append({
            "id": op.op_id, "tipo": op.clase_produccion,
            "sku": op.sku_producto, "nombre": op.nombre_producto or "",
            "progreso": progreso, "status": op.status,
            "fecha": op.fecha_inicio, "linea": op.linea_produccion,
            "operador": op.operador,
        })
    return unificadas


@router.get("/estado-silos")
@router.get("/estado-silos/")
async def obtener_estado_silos(db: AsyncSession = Depends(get_db)):
    """Retorna el estado actual de todos los silos (principales y AUX)"""
    principales, auxiliares = await _get_silos_ubicaciones(db)
    ahora_utc = datetime.utcnow()
    resultado = []

    for silo in principales:
        # Buscar lotes con stock en este silo
        res = await db.execute(
            select(LoteInventario).where(
                and_(
                    LoteInventario.ubicacion_id == silo.id,
                    LoteInventario.cantidad_actual > 0,
                    LoteInventario.estado_calidad == "Aprobado",
                )
            ).order_by(LoteInventario.fecha_recepcion.desc())
        )
        lotes = res.scalars().all()
        kg_total = sum(l.cantidad_actual for l in lotes)

        if lotes and kg_total > 0:
            # Usar el lote MÁS RECIENTE para datos del silo
            lote_principal = lotes[0]
            op_data = None
            if lote_principal.op_origen:
                res_op = await db.execute(
                    select(OrdenProduccion).where(
                        OrdenProduccion.op_id == lote_principal.op_origen
                    )
                )
                op_data = res_op.scalar_one_or_none()

            # Tiempo de reposo: desde la finalización del lote MÁS RECIENTE
            # Se resetea automáticamente cuando un nuevo lote se finaliza
            # porque lote_principal cambia al nuevo
            tiempo_reposo_seg = 0
            hora_fin = None
            if op_data and op_data.hora_finalizacion:
                hora_fin = op_data.hora_finalizacion
                tiempo_reposo_seg = (ahora_utc - hora_fin).total_seconds()
            elif op_data and op_data.fecha_fin:
                hora_fin = op_data.fecha_fin
                tiempo_reposo_seg = (ahora_utc - hora_fin).total_seconds()
            elif lote_principal.fecha_recepcion:
                # Fallback: usar fecha de recepción del lote
                hora_fin = lote_principal.fecha_recepcion
                tiempo_reposo_seg = (ahora_utc - hora_fin).total_seconds()

            resultado.append({
                "nombre_silo": silo.nombre,
                "es_aux": False,
                "vacio": False,
                "sku_resina": lote_principal.sku_producto,
                "nombre_resina": op_data.nombre_producto if op_data else "",
                "grado": op_data.grado if op_data else None,
                "densidad": op_data.densidad if op_data else None,
                "kg_totales": round(kg_total, 2),
                "fecha_entrada": lote_principal.fecha_recepcion,
                "hora_finalizacion_lote": hora_fin,
                "op_id_origen": lote_principal.op_origen,
                "tiempo_reposo_segundos": max(0, tiempo_reposo_seg),
                "tiempo_reposo_horas": round(max(0, tiempo_reposo_seg) / 3600, 2),
                "suministro": None,
                "silo_fuente": None,
            })
        else:
            # Silo vacío — tiempo de reposo = 0 (reseteado)
            resultado.append({
                "nombre_silo": silo.nombre,
                "es_aux": False,
                "vacio": True,
                "sku_resina": None, "nombre_resina": None, "grado": None,
                "densidad": None, "kg_totales": 0,
                "fecha_entrada": None, "hora_finalizacion_lote": None,
                "op_id_origen": None,
                "tiempo_reposo_segundos": 0, "tiempo_reposo_horas": 0,
                "suministro": None, "silo_fuente": None,
            })

    # AUX — obtener último suministro activo
    for aux in auxiliares:
        # Verificar si tiene stock propio (lotes en la ubicación del AUX)
        res_lotes = await db.execute(
            select(LoteInventario).where(
                and_(
                    LoteInventario.ubicacion_id == aux.id,
                    LoteInventario.cantidad_actual > 0,
                )
            )
        )
        lotes_aux = res_lotes.scalars().all()
        kg_aux = sum(l.cantidad_actual for l in lotes_aux)

        # Buscar último suministro registrado para este AUX
        res_sum = await db.execute(
            select(SuministroSilo).where(
                SuministroSilo.aux_destino == aux.nombre
            ).order_by(SuministroSilo.fecha_suministro.desc()).limit(1)
        )
        ultimo_suministro = res_sum.scalar_one_or_none()

        # El AUX está activo si tiene stock (independiente del estado del silo origen)
        if kg_aux > 0 and ultimo_suministro:
            resultado.append({
                "nombre_silo": aux.nombre,
                "es_aux": True,
                "vacio": False,
                "sku_resina": ultimo_suministro.sku_resina,
                "nombre_resina": ultimo_suministro.nombre_resina,
                "grado": ultimo_suministro.grado,
                "densidad": ultimo_suministro.densidad,
                "kg_totales": round(kg_aux, 2),
                "fecha_entrada": ultimo_suministro.fecha_suministro,
                "hora_finalizacion_lote": None,
                "op_id_origen": ultimo_suministro.silo_origen_op_id,
                "tiempo_reposo_segundos": ultimo_suministro.tiempo_reposo_horas * 3600,
                "tiempo_reposo_horas": ultimo_suministro.tiempo_reposo_horas,
                "suministro": {
                    "id": ultimo_suministro.id,
                    "suministro_id": ultimo_suministro.suministro_id,
                    "silo_origen": ultimo_suministro.silo_origen,
                    "silo_origen_op_id": ultimo_suministro.silo_origen_op_id,
                    "aux_destino": ultimo_suministro.aux_destino,
                    "sku_resina": ultimo_suministro.sku_resina,
                    "nombre_resina": ultimo_suministro.nombre_resina,
                    "grado": ultimo_suministro.grado,
                    "densidad": ultimo_suministro.densidad,
                    "kg_suministrados": ultimo_suministro.kg_suministrados,
                    "tiempo_reposo_horas": ultimo_suministro.tiempo_reposo_horas,
                    "maquinas_inyeccion": ultimo_suministro.maquinas_inyeccion or [],
                    "fecha_suministro": ultimo_suministro.fecha_suministro,
                    "creado_por": ultimo_suministro.creado_por,
                },
                "silo_fuente": ultimo_suministro.silo_origen,
            })
        elif kg_aux > 0 and not ultimo_suministro:
            # Tiene stock pero sin registro de suministro (caso raro/legacy)
            lote_ref = lotes_aux[0]
            resultado.append({
                "nombre_silo": aux.nombre,
                "es_aux": True,
                "vacio": False,
                "sku_resina": lote_ref.sku_producto,
                "nombre_resina": None,
                "grado": None,
                "densidad": None,
                "kg_totales": round(kg_aux, 2),
                "fecha_entrada": lote_ref.fecha_recepcion,
                "hora_finalizacion_lote": None,
                "op_id_origen": lote_ref.op_origen,
                "tiempo_reposo_segundos": 0,
                "tiempo_reposo_horas": 0,
                "suministro": None,
                "silo_fuente": None,
            })
        else:
            resultado.append({
                "nombre_silo": aux.nombre,
                "es_aux": True,
                "vacio": True,
                "sku_resina": None, "nombre_resina": None, "grado": None,
                "densidad": None, "kg_totales": 0,
                "fecha_entrada": None, "hora_finalizacion_lote": None,
                "op_id_origen": None,
                "tiempo_reposo_segundos": 0, "tiempo_reposo_horas": 0,
                "suministro": None, "silo_fuente": None,
            })

    return resultado


@router.get("/estado-silos/excel")
@router.get("/estado-silos/excel/")
async def descargar_estado_silos_excel(db: AsyncSession = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    principales, auxiliares = await _get_silos_ubicaciones(db)
    ahora_utc = datetime.utcnow()
    ahora_local = _ahora_local()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Silos"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    aux_fill = PatternFill(start_color="0097A7", end_color="0097A7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:N1")
    ws["A1"] = f"Estado EPS SILO — {ahora_local.strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Silo", "Tipo", "Estado", "SKU Resina", "Grado", "Densidad g/cm³",
        "Kg Totales", "Fecha Entrada", "Tiempo Reposo (hrs)",
        "Silo Fuente", "Kg Suministrados", "Reposo Silo Fuente (hrs)", "Máquinas", "Operador"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 4

    # Silos principales
    for silo in sorted(principales, key=lambda s: s.nombre):
        res = await db.execute(
            select(LoteInventario).where(
                and_(
                    LoteInventario.ubicacion_id == silo.id,
                    LoteInventario.cantidad_actual > 0,
                    LoteInventario.estado_calidad == "Aprobado",
                )
            ).order_by(LoteInventario.fecha_recepcion.desc())
        )
        lotes = res.scalars().all()
        kg = sum(l.cantidad_actual for l in lotes)

        op_data = None
        if lotes and lotes[0].op_origen:
            res_op = await db.execute(
                select(OrdenProduccion).where(OrdenProduccion.op_id == lotes[0].op_origen)
            )
            op_data = res_op.scalar_one_or_none()

        tiempo_reposo = 0
        if op_data and (op_data.hora_finalizacion or op_data.fecha_fin):
            ref = op_data.hora_finalizacion or op_data.fecha_fin
            tiempo_reposo = round((ahora_utc - ref).total_seconds() / 3600, 2)

        vals = [
            silo.nombre, "Principal",
            "Con material" if kg > 0 else "Vacío",
            lotes[0].sku_producto if lotes else "—",
            op_data.grado if op_data else "—",
            op_data.densidad if op_data else "—",
            round(kg, 2),
            lotes[0].fecha_recepcion.strftime("%Y-%m-%d %H:%M") if lotes else "—",
            tiempo_reposo if kg > 0 else 0,
            "N/A", "N/A", "N/A", "N/A",
            op_data.creado_por if op_data else "—",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
        row += 1

    # AUX
    for aux in sorted(auxiliares, key=lambda s: s.nombre):
        res_lotes = await db.execute(
            select(LoteInventario).where(
                and_(LoteInventario.ubicacion_id == aux.id, LoteInventario.cantidad_actual > 0)
            )
        )
        lotes_aux = res_lotes.scalars().all()
        kg_aux = sum(l.cantidad_actual for l in lotes_aux)

        res_sum = await db.execute(
            select(SuministroSilo).where(
                SuministroSilo.aux_destino == aux.nombre
            ).order_by(SuministroSilo.fecha_suministro.desc()).limit(1)
        )
        ultimo = res_sum.scalar_one_or_none()

        vals = [
            aux.nombre, "AUX",
            "Con material" if kg_aux > 0 and ultimo else "Vacío",
            ultimo.sku_resina if ultimo and kg_aux > 0 else "—",
            ultimo.grado if ultimo and kg_aux > 0 else "—",
            "—",
            round(kg_aux, 2),
            ultimo.fecha_suministro.strftime("%Y-%m-%d %H:%M") if ultimo and kg_aux > 0 else "—",
            "—",
            ultimo.silo_origen if ultimo and kg_aux > 0 else "—",
            ultimo.kg_suministrados if ultimo and kg_aux > 0 else "—",
            ultimo.tiempo_reposo_horas if ultimo and kg_aux > 0 else "—",
            ", ".join(ultimo.maquinas_inyeccion or []) if ultimo and kg_aux > 0 else "—",
            ultimo.creado_por if ultimo and kg_aux > 0 else "—",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if kg_aux > 0 and ultimo:
                cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
        row += 1

    for col in range(1, 15):
        ws.column_dimensions[chr(64 + col) if col < 27 else 'N'].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"estado_silos_{ahora_local.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════
# SUMINISTRO SILO
# ══════════════════════════════════════════════════════════════════

@router.get("/suministros")
@router.get("/suministros/")
async def listar_suministros(
    limite: int = 100,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SuministroSilo).order_by(SuministroSilo.fecha_suministro.desc()).limit(limite)
    )
    return result.scalars().all()


@router.post("/suministros")
@router.post("/suministros/")
async def crear_suministro(
    data: CrearSuministroRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)
    ahora = _ahora_local()
    ahora_utc = datetime.utcnow()

    # Validar silo origen
    res_silo = await db.execute(
        select(Ubicacion).where(Ubicacion.nombre == data.silo_origen)
    )
    silo_ub = res_silo.scalar_one_or_none()
    if not silo_ub:
        raise HTTPException(status_code=404, detail=f"Silo '{data.silo_origen}' no encontrado")

    # Obtener lotes del silo origen
    res_lotes = await db.execute(
        select(LoteInventario).where(
            and_(
                LoteInventario.ubicacion_id == silo_ub.id,
                LoteInventario.cantidad_actual > 0,
                LoteInventario.estado_calidad == "Aprobado",
            )
        ).order_by(LoteInventario.fecha_recepcion)
    )
    lotes_origen = res_lotes.scalars().all()
    stock_disponible = sum(l.cantidad_actual for l in lotes_origen)

    if stock_disponible < data.kg_suministrados:
        raise HTTPException(
            status_code=400,
            detail=f"Stock insuficiente en {data.silo_origen}. Disponible: {stock_disponible:.2f} kg, solicitado: {data.kg_suministrados:.2f} kg"
        )

    # Validar AUX destino
    res_aux = await db.execute(
        select(Ubicacion).where(Ubicacion.nombre == data.aux_destino)
    )
    aux_ub = res_aux.scalar_one_or_none()
    if not aux_ub:
        raise HTTPException(status_code=404, detail=f"AUX '{data.aux_destino}' no encontrado")

    # Obtener datos de la resina del silo
    lote_ref = lotes_origen[0]
    op_data = None
    if lote_ref.op_origen:
        res_op = await db.execute(
            select(OrdenProduccion).where(OrdenProduccion.op_id == lote_ref.op_origen)
        )
        op_data = res_op.scalar_one_or_none()

    # Calcular tiempo de reposo del silo
    tiempo_reposo = 0
    if op_data and (op_data.hora_finalizacion or op_data.fecha_fin):
        ref_time = op_data.hora_finalizacion or op_data.fecha_fin
        tiempo_reposo = round((ahora_utc - ref_time).total_seconds() / 3600, 2)

    # Consumir FIFO del silo origen
    restante = data.kg_suministrados
    for lote in lotes_origen:
        if restante <= 0:
            break
        tomar = min(lote.cantidad_actual, restante)
        lote.cantidad_actual -= tomar
        db.add(MovimientoLote(
            lote_id=lote.lote_id, fecha=ahora_utc,
            tipo="SUMINISTRO_SILO", cantidad=-tomar,
            detalles={
                "silo_origen": data.silo_origen,
                "aux_destino": data.aux_destino,
                "kg_suministrados": data.kg_suministrados,
            },
        ))
        restante -= tomar

    # ── Calcular kg restantes DESPUÉS del consumo ──────────────────
    kg_restantes_silo = round(stock_disponible - data.kg_suministrados, 2)

    # Crear lote en el AUX
    sufijo = (lote_ref.sku_producto or "XXXX")[-4:]
    nuevo_lote_id = f"SUM-{ahora.strftime('%d%m%y%H%M%S')}-{sufijo}"

    lote_aux = LoteInventario(
        lote_id=nuevo_lote_id,
        sku_producto=lote_ref.sku_producto,
        cantidad_actual=data.kg_suministrados,
        cantidad_inicial=data.kg_suministrados,
        ubicacion_id=aux_ub.id,
        fecha_recepcion=ahora_utc,
        op_origen=lote_ref.op_origen,
        estado_calidad="Aprobado",
    )
    db.add(lote_aux)

    db.add(MovimientoLote(
        lote_id=nuevo_lote_id, fecha=ahora_utc,
        tipo="ENTRADA_SUMINISTRO", cantidad=data.kg_suministrados,
        detalles={
            "silo_origen": data.silo_origen,
            "aux_destino": data.aux_destino,
        },
    ))

    # Crear registro de suministro con kg_restantes
    suministro_id = f"SUM-{ahora.strftime('%d%m%y%H%M%S')}"
    suministro = SuministroSilo(
        suministro_id=suministro_id,
        silo_origen=data.silo_origen,
        silo_origen_op_id=lote_ref.op_origen,
        aux_destino=data.aux_destino,
        sku_resina=lote_ref.sku_producto,
        nombre_resina=op_data.nombre_producto if op_data else "",
        grado=op_data.grado if op_data else None,
        densidad=0,
        kg_suministrados=data.kg_suministrados,
        kg_restantes=kg_restantes_silo,
        tiempo_reposo_horas=tiempo_reposo,
        maquinas_inyeccion=data.maquinas_inyeccion or [],
        fecha_suministro=ahora_utc,
        creado_por=user.username,
    )
    db.add(suministro)

    await db.commit()

    return {
        "message": f"Suministro registrado: {data.kg_suministrados:.2f} kg de {data.silo_origen} → {data.aux_destino}",
        "suministro_id": suministro_id,
        "stock_restante_silo": kg_restantes_silo,
    }


# ══════════════════════════════════════════════════════════════════
# REPORTE PRE-EXPANSIÓN EXCEL
# ══════════════════════════════════════════════════════════════════

@router.get("/reporte-preexpansion/excel")
@router.get("/reporte-preexpansion/excel/")
async def descargar_reporte_preexpansion_excel(
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    ahora_local = _ahora_local()

    if fecha:
        try:
            fecha_filtro = datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido")
    else:
        fecha_filtro = ahora_local.replace(hour=0, minute=0, second=0, microsecond=0)

    # Rango amplio para cubrir turnos noche (que cruzan medianoche)
    inicio_utc = (fecha_filtro - timedelta(days=1)).replace(tzinfo=TZ_LOCAL).astimezone(timezone.utc).replace(tzinfo=None)
    fin_utc = (fecha_filtro + timedelta(days=2)).replace(tzinfo=TZ_LOCAL).astimezone(timezone.utc).replace(tzinfo=None)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(
                OrdenProduccion.clase_produccion == "PRE-EXPANSION",
                OrdenProduccion.fecha_inicio >= inicio_utc,
                OrdenProduccion.fecha_inicio < fin_utc,
            )
        ).order_by(OrdenProduccion.fecha_inicio)
    )
    todas = result.scalars().all()

    fecha_str = fecha_filtro.strftime("%Y-%m-%d")
    ordenes = []
    for op in todas:
        ft = _get_fecha_turno(op.fecha_inicio)
        t = _get_turno_from_dt(op.fecha_inicio)
        if ft == fecha_str and (not turno or t == turno.upper()):
            ordenes.append(op)

    result_sum = await db.execute(
        select(SuministroSilo).where(
            and_(
                SuministroSilo.fecha_suministro >= inicio_utc,
                SuministroSilo.fecha_suministro < fin_utc,
            )
        ).order_by(SuministroSilo.fecha_suministro)
    )
    all_suministros = result_sum.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Pre-Expansión"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:P1")
    ws["A1"] = "REPORTE DE PREEXPANSIÓN"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Fecha: {fecha_str}" + (f" | Turno: {turno.upper()}" if turno else "")
    ws["A2"].font = Font(bold=True, size=11)

    headers = [
        "Turno", "No. Costal", "No. Lote", "No. Silo", "Resina\nGrade",
        "Densidad\ng/cm³", "Pantalla\nPeso Kg", "Ciclo\nSec.",
        "Fecha\nEntrada", "Inicio\nHr", "Final\nHr",
        "Counter\nTiro", "Silo de\nSalida", "Kg\nEnviados",
        "Kg\nRestantes", "Usuario"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    row = 5
    op_ids_usados = set()

    for op in ordenes:
        t = _get_turno_from_dt(op.fecha_inicio)
        ahora_l = op.fecha_inicio.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL) if op.fecha_inicio else None

        hora_inicio = ""
        if op.hora_inicio_real:
            h = op.hora_inicio_real.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
            hora_inicio = h.strftime("%d/%m/%y %H:%M")
        elif ahora_l:
            hora_inicio = ahora_l.strftime("%d/%m/%y %H:%M")

        hora_fin = ""
        if op.hora_finalizacion:
            h = op.hora_finalizacion.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
            hora_fin = h.strftime("%d/%m/%y %H:%M")
        elif op.fecha_fin:
            h = op.fecha_fin.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
            hora_fin = h.strftime("%d/%m/%y %H:%M")

        fecha_entrada = ""
        if op.fecha_inicio:
            fe = op.fecha_inicio.replace(tzinfo=timezone.utc).astimezone(TZ_LOCAL)
            fecha_entrada = fe.strftime("%d/%m/%y")

        # Buscar suministros de esta OP
        sums_op = [s for s in all_suministros if s.silo_origen_op_id == op.op_id]

        if not sums_op:
            vals = [
                t, op.numero_costal or "", op.op_id,
                op.ubicacion_destino or "", op.grado or "",
                op.densidad or "", op.pantalla_peso or "", op.ciclo_seg or "",
                fecha_entrada, hora_inicio, hora_fin,
                op.counter_tiro or "", "", "", "", op.creado_por or "",
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = center
            row += 1
        else:
            for idx, s in enumerate(sums_op):
                op_ids_usados.add(op.op_id)
                if idx == 0:
                    vals = [
                        t, op.numero_costal or "", op.op_id,
                        op.ubicacion_destino or "", op.grado or "",
                        op.densidad or "", op.pantalla_peso or "", op.ciclo_seg or "",
                        fecha_entrada, hora_inicio, hora_fin,
                        op.counter_tiro or "", s.aux_destino,
                        s.kg_suministrados,
                        s.kg_restantes if s.kg_restantes is not None else "",
                        op.creado_por or "",
                    ]
                else:
                    vals = [
                        "", "", "↳ suministro", "", "", "", "", "",
                        "", "", "", "", s.aux_destino,
                        s.kg_suministrados,
                        s.kg_restantes if s.kg_restantes is not None else "",
                        "",
                    ]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = thin_border
                    cell.alignment = center
                row += 1

    # Suministros huérfanos (de OPs antiguas, no del día)
    for s in all_suministros:
        fs = _get_fecha_turno(s.fecha_suministro)
        st = _get_turno_from_dt(s.fecha_suministro)
        if fs != fecha_str:
            continue
        if turno and st != turno.upper():
            continue
        if s.silo_origen_op_id and s.silo_origen_op_id in op_ids_usados:
            continue
        vals = [
            st, "", s.silo_origen_op_id or "—",
            s.silo_origen, s.grado or "", "", "", "",
            "", "", "", "", s.aux_destino,
            s.kg_suministrados,
            s.kg_restantes if s.kg_restantes is not None else "",  # ← CORREGIDO
            s.creado_por or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.alignment = center
        row += 1

    widths = [8, 12, 22, 12, 10, 10, 10, 8, 12, 16, 16, 10, 12, 10, 10, 14]
    for i, w in enumerate(widths):
        col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
        ws.column_dimensions[col_letter].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"reporte_preexpansion_{fecha_str}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════
# DETALLE OP
# ══════════════════════════════════════════════════════════════════

@router.get("/{op_id}", response_model=OrdenProduccionResponse)
@router.get("/{op_id}/", response_model=OrdenProduccionResponse)
async def obtener_orden(op_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(OrdenProduccion).where(OrdenProduccion.op_id == op_id)
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden de producción no encontrada")
    return op


# ══════════════════════════════════════════════════════════════════
# PRE-EXPANSIÓN
# ══════════════════════════════════════════════════════════════════

@router.post("/pre-expansion/iniciar")
@router.post("/pre-expansion/iniciar/")
async def iniciar_pre_expansion(
    data: IniciarPreExpansionRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    if data.cantidad_usada <= 0:
        raise HTTPException(status_code=400, detail="La cantidad de materia prima a usar debe ser mayor a cero")

    # Consultar stock excluyendo lo que ya está en silos
    silo_ids = await _get_silo_ubicacion_ids(db)
    stock = await _consultar_stock_aprobado_sin_silos(db, data.sku_materia_prima, silo_ids)
    mensaje_stock = ""
    oc_generada = None

    if stock < data.cantidad_usada:
        faltante = data.cantidad_usada - stock
        ahora = _ahora_local()
        oc_id = f"OC-PROD-{ahora.strftime('%Y%m%d%H%M%S')}"

        nombre_mp = data.sku_materia_prima
        prod_mp = await db.execute(select(Producto).where(Producto.sku == data.sku_materia_prima))
        prod_mp_obj = prod_mp.scalar_one_or_none()
        if prod_mp_obj and prod_mp_obj.nombre:
            nombre_mp = prod_mp_obj.nombre

        nueva_oc = OCModel(
            oc_id=oc_id, id_proveedor="POR-ASIGNAR", nombre_proveedor="Por asignar",
            status="Pendiente Aprobación", origen="PRODUCCION",
            notas=f"Generada automáticamente desde Pre-Expansión. "
                  f"Materia prima: {data.sku_materia_prima}. "
                  f"Stock actual: {stock:.2f} kg, requerido: {data.cantidad_usada:.2f} kg, "
                  f"faltante: {faltante:.2f} kg. Operador: {data.operador or 'N/A'}.",
            creado_por=f"SISTEMA (Pre-Expansión - {user.username})",
        )
        db.add(nueva_oc)
        await db.flush()

        item_oc = OCItemModel(
            orden_compra_id=nueva_oc.id, sku_producto=data.sku_materia_prima,
            nombre_producto=nombre_mp, cantidad_requerida=faltante,
            cantidad_recibida=0, precio_unitario=0, moneda="MXN",
        )
        db.add(item_oc)
        oc_generada = oc_id
        mensaje_stock = (
            f" ⚠️ Stock insuficiente de {data.sku_materia_prima} "
            f"(disponible: {stock:.2f}, requerido: {data.cantidad_usada:.2f}). "
            f"Se generó la orden de compra {oc_id} pendiente de aprobación por Compras."
        )

    op_id = await _generar_op_id(db, data.sku_producto_resina, "PRE-EXPANSION")
    ahora_utc = datetime.utcnow()

    # Si cantidad_a_producir es 0, usar cantidad_usada como referencia
    cant_producir = data.cantidad_a_producir if data.cantidad_a_producir > 0 else data.cantidad_usada

    orden = OrdenProduccion(
        op_id=op_id,
        clase_produccion="PRE-EXPANSION",
        sku_producto=data.sku_producto_resina,
        nombre_producto="",
        sku_materia_prima=data.sku_materia_prima,
        grado=data.grado,
        numero_costal=data.numero_costal,
        cantidad_a_producir=cant_producir,
        cantidad_usada_requerida=data.cantidad_usada,
        operador=data.operador,
        ubicacion_destino=data.ubicacion_destino or "PISO",
        status="En Proceso",
        fecha_inicio=ahora_utc,
        hora_inicio_real=ahora_utc,
        creado_por=user.username,
    )
    db.add(orden)

    producto = await db.execute(select(Producto).where(Producto.sku == data.sku_producto_resina))
    prod = producto.scalar_one_or_none()
    if prod:
        orden.nombre_producto = prod.nombre or ""

    await db.commit()
    await db.refresh(orden)

    return {
        "message": f"Lote de pre-expansión {op_id} iniciado.{mensaje_stock}",
        "op_id": op_id,
        "oc_generada": oc_generada,
    }


async def _consultar_stock_aprobado_sin_silos(
    db: AsyncSession, sku: str, excluir_ids: set
) -> float:
    """Suma stock aprobado excluyendo ubicaciones de silos"""
    result = await db.execute(
        select(LoteInventario).where(
            and_(
                LoteInventario.sku_producto == sku,
                LoteInventario.estado_calidad == "Aprobado",
                LoteInventario.cantidad_actual > 0,
            )
        )
    )
    lotes = result.scalars().all()
    total = 0.0
    for l in lotes:
        if l.ubicacion_id is None or l.ubicacion_id not in excluir_ids:
            total += l.cantidad_actual
    return total


@router.post("/pre-expansion/{op_id}/datos-proceso")
@router.post("/pre-expansion/{op_id}/datos-proceso/")
async def registrar_datos_proceso(
    op_id: str,
    data: RegistroDatosProcesoRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Registra densidad, pantalla peso y ciclo (se llama al primer parcial)"""
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(OrdenProduccion.op_id == op_id, OrdenProduccion.clase_produccion == "PRE-EXPANSION")
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    op.densidad = data.densidad
    op.pantalla_peso = data.pantalla_peso
    op.ciclo_seg = data.ciclo_seg

    await db.commit()
    return {"message": "Datos de proceso registrados", "op_id": op_id}


@router.post("/pre-expansion/{op_id}/produccion-parcial")
@router.post("/pre-expansion/{op_id}/produccion-parcial/")
async def registrar_produccion_parcial(
    op_id: str,
    data: RegistroParciallPreExpRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(OrdenProduccion.op_id == op_id, OrdenProduccion.clase_produccion == "PRE-EXPANSION")
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Lote de pre-expansión no encontrado")

    if op.status == "Finalizado":
        raise HTTPException(status_code=400, detail="El lote ya está finalizado")

    if data.cantidad_parcial_producida <= 0:
        raise HTTPException(status_code=400, detail="La cantidad debe ser mayor a cero")

    if op.cantidad_a_producir == 0:
        raise HTTPException(status_code=400, detail="La cantidad a producir no puede ser cero")

    cantidad_a_consumir = math.ceil(
        (data.cantidad_parcial_producida * op.cantidad_usada_requerida) / op.cantidad_a_producir
    )

    # Obtener IDs de ubicaciones SILO para excluirlas del consumo
    silo_ids = await _get_silo_ubicacion_ids(db)

    # Consumir materia prima FIFO (excluyendo lotes en silos)
    plan_consumo = await _consumir_stock_fifo(
        db, op.sku_materia_prima, cantidad_a_consumir,
        {"op_id": op_id, "tipo": "PRE-EXPANSION"},
        excluir_ubicacion_ids=silo_ids,
    )

    registro = {
        "fecha": datetime.utcnow().isoformat(),
        "cantidad_producida": data.cantidad_parcial_producida,
        "cantidad_consumida": cantidad_a_consumir,
        "lotes_materia_prima_usados": plan_consumo,
    }

    parciales = list(op.registros_parciales or [])
    parciales.append(registro)
    op.registros_parciales = parciales
    op.cantidad_producida += data.cantidad_parcial_producida
    op.cantidad_total_consumida += cantidad_a_consumir

    # Indicar si es el primer parcial (frontend muestra modal datos de proceso)
    es_primer_parcial = len(parciales) == 1

    await db.commit()

    return {
        "message": f"Producción parcial registrada: {data.cantidad_parcial_producida} kg",
        "op_id": op_id,
        "cantidad_total_producida": op.cantidad_producida,
        "cantidad_total_consumida": op.cantidad_total_consumida,
        "es_primer_parcial": es_primer_parcial,
    }


@router.post("/pre-expansion/{op_id}/finalizar")
@router.post("/pre-expansion/{op_id}/finalizar/")
async def finalizar_pre_expansion(
    op_id: str,
    data: FinalizarPreExpansionRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(OrdenProduccion.op_id == op_id, OrdenProduccion.clase_produccion == "PRE-EXPANSION")
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Lote de pre-expansión no encontrado")

    destino = data.ubicacion_destino_final or op.ubicacion_destino
    if not destino:
        raise HTTPException(status_code=400, detail="No se especificó ubicación de destino")

    res_ub = await db.execute(select(Ubicacion).where(Ubicacion.nombre == destino))
    ubicacion = res_ub.scalar_one_or_none()
    if not ubicacion:
        raise HTTPException(
            status_code=400,
            detail=f"La ubicación '{destino}' no existe en el sistema."
        )

    ahora_utc = datetime.utcnow()

    # Establecer cantidad producida final (del modal de finalización)
    if data.cantidad_producida > 0:
        op.cantidad_producida = data.cantidad_producida
        op.cantidad_a_producir = data.cantidad_producida  # Sincronizar

    lote_inv_id = None
    if op.cantidad_producida > 0:
        lote_inv_id = await _crear_lote_produccion(
            db, op.sku_producto, op.cantidad_producida, op_id,
            ubicacion_nombre=destino,
        )

    op.status = "Finalizado"
    op.fecha_fin = ahora_utc
    op.hora_finalizacion = ahora_utc
    op.lote_inventario_generado = lote_inv_id
    op.ubicacion_destino = destino
    op.silo_destino = destino
    op.counter_tiro = data.counter_tiro

    await db.commit()

    return {
        "message": f"Lote {op_id} finalizado. {op.cantidad_producida:.2f} kg ubicados en {destino}.",
        "op_id": op_id,
        "lote_inventario_generado": lote_inv_id,
    }


# ══════════════════════════════════════════════════════════════════
# INYECCIÓN
# ══════════════════════════════════════════════════════════════════

@router.post("/inyeccion/iniciar")
@router.post("/inyeccion/iniciar/")
async def iniciar_inyeccion(
    data: IniciarInyeccionRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    producto = await _get_producto(db, data.sku_producto)
    bom = producto.bom or []

    # Buscar componente tipo RESINA en el BOM
    componente_resina = None
    for comp in bom:
        sku_comp = comp.get("sku_componente")
        if sku_comp:
            res_prod = await db.execute(
                select(Producto).where(Producto.sku == sku_comp)
            )
            prod_comp = res_prod.scalar_one_or_none()
            if prod_comp and (prod_comp.tipo or "").upper() == "RESINA":
                componente_resina = comp
                break

    if not componente_resina:
        raise HTTPException(
            status_code=400,
            detail="No se encontró componente tipo 'RESINA' en el BOM del producto"
        )

    sku_resina = componente_resina["sku_componente"]
    cantidad_por_unidad = componente_resina.get("cantidad", 0)
    cantidad_resina_necesaria = math.ceil(data.cantidad_a_producir * cantidad_por_unidad)

    op_id = await _generar_op_id(db, data.sku_producto, "INYECCION")

    # Verificar stock
    stock = await _consultar_stock_aprobado(db, sku_resina)

    orden = OrdenProduccion(
        op_id=op_id,
        clase_produccion="INYECCION",
        sku_producto=data.sku_producto,
        nombre_producto=producto.nombre or "",
        linea_produccion=data.linea_produccion,
        cantidad_a_producir=data.cantidad_a_producir,
        cantidad_carrito=data.cantidad_carrito,
        operador=data.operador,
        fecha_inicio=datetime.utcnow(),
        creado_por=user.username,
    )

    if stock >= cantidad_resina_necesaria:
        consumo = await _consumir_stock_fifo(
            db, sku_resina, cantidad_resina_necesaria,
            {"op_id": op_id, "tipo": "INYECCION"},
            data.linea_produccion,
        )
        orden.status = "En Proceso"
        orden.material_consumido = consumo
        db.add(orden)
        await db.commit()
        await db.refresh(orden)

        return {
            "message": f"Orden {op_id} iniciada. Se consumieron {cantidad_resina_necesaria} kg de {sku_resina}.",
            "op_id": op_id,
            "status": "En Proceso",
        }
    else:
        orden.status = "Pendiente Material"
        orden.material_consumido = []
        db.add(orden)
        await db.commit()
        await db.refresh(orden)

        faltante = cantidad_resina_necesaria - stock
        return {
            "message": f"Orden {op_id} creada 'Pendiente Material'. Faltan {faltante:.2f} kg de {sku_resina}.",
            "op_id": op_id,
            "status": "Pendiente Material",
            "faltante": faltante,
            "sku_resina": sku_resina,
        }


@router.post("/inyeccion/{op_id}/pieza")
@router.post("/inyeccion/{op_id}/pieza/")
async def registrar_pieza_inyeccion(
    op_id: str,
    data: RegistrarPiezaRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(
                OrdenProduccion.op_id == op_id,
                OrdenProduccion.clase_produccion == "INYECCION",
            )
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden de inyección no encontrada")

    if op.status == "Finalizado":
        raise HTTPException(status_code=400, detail="La orden ya está finalizada")

    cantidad_antes = op.cantidad_producida
    op.cantidad_producida += data.cantidad

    carrito_completado = False
    numero_carrito = 0
    if op.cantidad_carrito > 0 and op.cantidad_producida > 0:
        carritos_antes = int(cantidad_antes // op.cantidad_carrito)
        carritos_despues = int(op.cantidad_producida // op.cantidad_carrito)
        if carritos_despues > carritos_antes:
            carrito_completado = True
            numero_carrito = carritos_despues

    await db.commit()

    return {
        "message": f"Pieza registrada. Total: {int(op.cantidad_producida)}",
        "cantidad_producida": int(op.cantidad_producida),
        "carrito_completado": carrito_completado,
        "numero_carrito": numero_carrito,
    }


@router.post("/inyeccion/{op_id}/finalizar")
@router.post("/inyeccion/{op_id}/finalizar/")
async def finalizar_inyeccion(
    op_id: str,
    data: FinalizarInyeccionRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(
                OrdenProduccion.op_id == op_id,
                OrdenProduccion.clase_produccion == "INYECCION",
            )
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden de inyección no encontrada")

    # Crear lote de producto terminado
    lote_inv_id = None
    if op.cantidad_producida > 0:
        lote_inv_id = await _crear_lote_produccion(
            db, op.sku_producto, op.cantidad_producida, op_id
        )

    op.status = "Finalizado"
    op.fecha_fin = datetime.utcnow()
    op.scrap_reportado = data.scrap_data
    op.lote_inventario_generado = lote_inv_id

    await db.commit()

    return {
        "message": f"Orden {op_id} finalizada",
        "op_id": op_id,
        "lote_inventario_generado": lote_inv_id,
    }


# ══════════════════════════════════════════════════════════════════
# ENSAMBLE (ASSY)
# ══════════════════════════════════════════════════════════════════

@router.post("/assy/iniciar")
@router.post("/assy/iniciar/")
async def iniciar_assy(
    data: IniciarAssyRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    producto = await _get_producto(db, data.sku_producto)
    bom = producto.bom or []
    if not bom:
        raise HTTPException(
            status_code=400,
            detail=f"El producto {data.sku_producto} no tiene BOM definido"
        )

    # Fase 1: Validar stock de todos los componentes
    componentes_faltantes = []
    plan_de_consumo = []

    for comp in bom:
        sku_comp = comp.get("sku_componente")
        cantidad_total = math.ceil(data.cantidad_a_producir * comp.get("cantidad", 0))
        if not sku_comp or cantidad_total <= 0:
            continue

        stock = await _consultar_stock_aprobado(db, sku_comp)
        if stock < cantidad_total:
            faltante = cantidad_total - stock
            componentes_faltantes.append({
                "sku": sku_comp,
                "requerido": cantidad_total,
                "disponible": stock,
                "faltante": faltante,
            })
        else:
            plan_de_consumo.append({"sku": sku_comp, "cantidad": cantidad_total})

    # Fase 2: Crear orden
    op_id = await _generar_op_id(db, data.sku_producto, "ASSY")

    orden = OrdenProduccion(
        op_id=op_id,
        clase_produccion="ASSY",
        sku_producto=data.sku_producto,
        nombre_producto=producto.nombre or "",
        linea_produccion=data.linea_produccion,
        cantidad_a_producir=data.cantidad_a_producir,
        cantidad_carrito=data.cantidad_carrito,
        uph_esperado=data.uph_esperado,
        metodo_conteo=data.metodo_conteo,
        operador=data.operador,
        fecha_inicio=datetime.utcnow(),
        creado_por=user.username,
    )

    if componentes_faltantes:
        # No hay suficiente stock
        orden.status = "Pendiente Material"
        orden.material_consumido = []
        db.add(orden)
        await db.commit()
        await db.refresh(orden)

        return {
            "message": f"Orden {op_id} creada 'Pendiente Material'",
            "op_id": op_id,
            "status": "Pendiente Material",
            "componentes_faltantes": componentes_faltantes,
        }
    else:
        # Consumir todo
        material_consumido = []
        for item in plan_de_consumo:
            consumo = await _consumir_stock_fifo(
                db, item["sku"], item["cantidad"],
                {"op_id": op_id, "tipo": "ASSY"},
                data.linea_produccion,
            )
            material_consumido.extend(consumo)

        orden.status = "En Proceso"
        orden.material_consumido = material_consumido
        db.add(orden)
        await db.commit()
        await db.refresh(orden)

        return {
            "message": f"Orden {op_id} iniciada. Material consumido del inventario.",
            "op_id": op_id,
            "status": "En Proceso",
        }


@router.post("/assy/{op_id}/pieza")
@router.post("/assy/{op_id}/pieza/")
async def registrar_pieza_assy(
    op_id: str,
    data: RegistrarPiezaRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(
                OrdenProduccion.op_id == op_id,
                OrdenProduccion.clase_produccion == "ASSY",
            )
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden de ensamble no encontrada")

    if op.status == "Finalizado":
        raise HTTPException(status_code=400, detail="La orden ya está finalizada")

    cantidad_antes = op.cantidad_producida
    op.cantidad_producida += data.cantidad

    carrito_completado = False
    numero_carrito = 0
    if op.cantidad_carrito > 0 and op.cantidad_producida > 0:
        carritos_antes = int(cantidad_antes // op.cantidad_carrito)
        carritos_despues = int(op.cantidad_producida // op.cantidad_carrito)
        if carritos_despues > carritos_antes:
            carrito_completado = True
            numero_carrito = carritos_despues

    await db.commit()

    return {
        "message": f"Pieza registrada. Total: {int(op.cantidad_producida)}",
        "cantidad_producida": int(op.cantidad_producida),
        "carrito_completado": carrito_completado,
        "numero_carrito": numero_carrito,
    }


@router.post("/assy/{op_id}/finalizar")
@router.post("/assy/{op_id}/finalizar/")
async def finalizar_assy(
    op_id: str,
    data: FinalizarAssyRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(
            and_(
                OrdenProduccion.op_id == op_id,
                OrdenProduccion.clase_produccion == "ASSY",
            )
        )
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden de ensamble no encontrada")

    lote_inv_id = None
    if op.cantidad_producida > 0:
        lote_inv_id = await _crear_lote_produccion(
            db, op.sku_producto, op.cantidad_producida, op_id
        )

    op.status = "Finalizado"
    op.fecha_fin = datetime.utcnow()
    op.scrap_reportado = data.scrap_data
    op.lote_inventario_generado = lote_inv_id

    await db.commit()

    return {
        "message": f"Orden {op_id} finalizada",
        "op_id": op_id,
        "lote_inventario_generado": lote_inv_id,
    }


# ══════════════════════════════════════════════════════════════════
# SURTIR MATERIAL PENDIENTE (cualquier clase)
# ══════════════════════════════════════════════════════════════════

@router.post("/{op_id}/surtir-material")
@router.post("/{op_id}/surtir-material/")
async def surtir_material_pendiente(
    op_id: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(OrdenProduccion.op_id == op_id)
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    if op.status != "Pendiente Material":
        raise HTTPException(status_code=400, detail="La orden no está pendiente de material")

    producto = await _get_producto(db, op.sku_producto)
    bom = producto.bom or []
    if not bom:
        raise HTTPException(status_code=400, detail="Producto sin BOM")

    # Verificar disponibilidad
    plan_de_consumo = []
    for comp in bom:
        sku_comp = comp.get("sku_componente")
        cantidad_total = math.ceil(op.cantidad_a_producir * comp.get("cantidad", 0))
        if not sku_comp or cantidad_total <= 0:
            continue

        stock = await _consultar_stock_aprobado(db, sku_comp)
        if stock < cantidad_total:
            raise HTTPException(
                status_code=400,
                detail=f"Aún falta material: {sku_comp} (disponible: {stock:.2f}, requerido: {cantidad_total:.2f})"
            )
        plan_de_consumo.append({"sku": sku_comp, "cantidad": cantidad_total})

    # Consumir todo
    material_consumido = []
    for item in plan_de_consumo:
        consumo = await _consumir_stock_fifo(
            db, item["sku"], item["cantidad"],
            {"op_id": op_id, "tipo": op.clase_produccion},
            op.linea_produccion,
        )
        material_consumido.extend(consumo)

    op.status = "En Proceso"
    op.material_consumido = material_consumido

    await db.commit()

    return {
        "message": f"Material surtido. Orden {op_id} ahora 'En Proceso'.",
        "op_id": op_id,
        "status": "En Proceso",
    }


# ══════════════════════════════════════════════════════════════════
# PAROS
# ══════════════════════════════════════════════════════════════════

@router.post("/{op_id}/paro/iniciar")
@router.post("/{op_id}/paro/iniciar/")
async def iniciar_paro(
    op_id: str,
    data: RegistrarParoRequest,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(OrdenProduccion.op_id == op_id)
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    paros = list(op.paros or [])

    # Verificar que no haya paro activo
    paro_activo = next((p for p in paros if p.get("status") == "Activo"), None)
    if paro_activo:
        raise HTTPException(status_code=400, detail="Ya hay un paro activo en esta orden")

    import uuid
    nuevo_paro = {
        "id": str(uuid.uuid4()),
        "motivo": data.motivo,
        "inicio": datetime.utcnow().isoformat(),
        "fin": None,
        "duracion_segundos": 0,
        "status": "Activo",
    }
    paros.append(nuevo_paro)
    op.paros = paros

    await db.commit()

    return {"message": "Paro iniciado", "paro_id": nuevo_paro["id"]}


@router.post("/{op_id}/paro/finalizar")
@router.post("/{op_id}/paro/finalizar/")
async def finalizar_paro(
    op_id: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_prod_role(user)

    result = await db.execute(
        select(OrdenProduccion).where(OrdenProduccion.op_id == op_id)
    )
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    paros = list(op.paros or [])
    paro_activo = next((p for p in paros if p.get("status") == "Activo"), None)

    if not paro_activo:
        raise HTTPException(status_code=400, detail="No hay paro activo en esta orden")

    ahora = datetime.utcnow()
    paro_activo["status"] = "Finalizado"
    paro_activo["fin"] = ahora.isoformat()

    try:
        inicio = datetime.fromisoformat(paro_activo["inicio"])
        paro_activo["duracion_segundos"] = (ahora - inicio).total_seconds()
    except Exception:
        paro_activo["duracion_segundos"] = 0

    op.paros = paros

    await db.commit()

    return {
        "message": "Paro finalizado",
        "duracion_segundos": paro_activo["duracion_segundos"],
    }