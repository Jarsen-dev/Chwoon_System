from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
from typing import List, Optional
from io import BytesIO
from sqlalchemy import select, func
from datetime import datetime, timedelta, timezone

from app.database import AsyncSessionLocal
from app.models.registro_secado  import RegistroSecado
from app.models.inventario        import InventarioPlanta
from app.core.security            import decode_token

# Zona horaria UTC-6 (CST México)
TZ_LOCAL = timezone(timedelta(hours=-6))

def ahora_local() -> datetime:
    return datetime.now(TZ_LOCAL)

router = APIRouter(prefix="/secado", tags=["secado"])


# ── Helpers ──────────────────────────────────────────────────────────

def get_turno_actual() -> str:
    now = ahora_local()
    total_min = now.hour * 60 + now.minute
    return "DIA" if 450 <= total_min < 1170 else "NOCHE"


def get_fecha_turno() -> str:
    now = ahora_local()
    total_min = now.hour * 60 + now.minute
    if total_min < 450:
        return (now - timedelta(days=1)).strftime("%Y-%m-%d")
    return now.strftime("%Y-%m-%d")


def calcular_tiempo(entrada: str, salida: str) -> str:
    try:
        hE, mE, sE = map(int, entrada.split(':'))
        hS, mS, sS = map(int, salida.split(':'))
        diff = (hS * 3600 + mS * 60 + sS) - (hE * 3600 + mE * 60 + sE)
        if diff <= 0:
            return "—"
        h = diff // 3600
        m = (diff % 3600) // 60
        s = diff % 60
        if h > 0:
            return f"{h}h {m}min {s}s"
        if m > 0:
            return f"{m}min {s}s"
        return f"{s}s"
    except Exception:
        return "—"


# ── GET /secado/registros/ ───────────────────────────────────────────

@router.get("/registros/")
@router.get("/registros")
async def obtener_registros(
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
):
    fecha_q = fecha or get_fecha_turno()
    turno_q = turno or get_turno_actual()

    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(RegistroSecado)
            .where(RegistroSecado.fecha == fecha_q)
            .where(RegistroSecado.turno == turno_q)
            .order_by(RegistroSecado.created_at.asc())   # ← asc para calcular orden
        )
        registros = result.scalars().all()

        # ── Recalcular total acumulado por parte ──────────────────────
        # contador[numero_parte] = cuántos han salido
        contador: dict[str, int] = {}
        # qty_unitario[numero_parte] = piezas por carrito
        qty_unit: dict[str, int] = {}

        for r in registros:
            if r.numero_parte not in qty_unit:
                # Inferir qty unitario del primer registro
                # Si tiene qty_total y es el carrito 1 → qty_unit = qty_total
                # Si no, dejamos 0 hasta tener referencia
                qty_unit[r.numero_parte] = 0

        # Primer paso: obtener qty unitario real del inventario
        partes_unicas = list({r.numero_parte for r in registros})
        if partes_unicas:
            result_inv = await db.execute(
                select(InventarioPlanta)
                .where(InventarioPlanta.codigo.in_(partes_unicas))
            )
            for inv in result_inv.scalars().all():
                qty_unit[inv.codigo] = int(inv.qtu or 0)

        # Segundo paso: calcular acumulado por orden de salida
        acumulado_salida: dict[str, int] = {}   # parte → total actual al salir

        # Ordenar solo los salidos por created_at para asignar orden correcto
        salidos_por_parte: dict[str, list] = {}
        for r in registros:
            if r.estado == "salido":
                if r.numero_parte not in salidos_por_parte:
                    salidos_por_parte[r.numero_parte] = []
                salidos_por_parte[r.numero_parte].append(r)

        # Asignar total acumulado en orden de salida
        total_map: dict[int, int] = {}   # id → qty_total calculado
        for parte, regs_salidos in salidos_por_parte.items():
            regs_salidos.sort(key=lambda x: x.created_at or x.id)
            qty_u = qty_unit.get(parte, 0)
            for idx, reg in enumerate(regs_salidos, start=1):
                total_map[reg.id] = idx * qty_u

        # ── Construir respuesta en orden DESC (más reciente primero) ──
        registros_desc = sorted(
            registros,
            key=lambda x: x.created_at or x.id,
            reverse=True,
        )

        return [
            {
                "id":               r.id,
                "numero_parte":     r.numero_parte,
                "descripcion":      r.descripcion      or "",
                "maquina":          r.maquina          or "—",
                "carrito":          r.carrito,
                "hora_entrada":     r.hora_entrada,
                "hora_salida":      r.hora_salida,
                "tiempo_en_camara": r.tiempo_en_camara,
                "qty_total":        total_map.get(r.id, r.qty_total),
                "estado":           r.estado,
                "usuario":          r.usuario          or "—",
            }
            for r in registros_desc
        ]

# ── POST /secado/escanear/ ───────────────────────────────────────────

@router.post("/escanear/")
@router.post("/escanear")
async def escanear(body: dict):
    codigo:    str = body.get("codigo", "").strip().upper()
    token_str: str = body.get("token",  "").strip()        # ← NUEVO

    if not codigo:
        raise HTTPException(status_code=400, detail="Código vacío")

    # ── Decodificar usuario ──────────────────────────────────────────
    usuario = "desconocido"
    if token_str:
        try:
            payload = decode_token(token_str)
            if payload:
                usuario = payload.get("sub") or "desconocido"
        except Exception:
            usuario = "desconocido"

    # ── Parsear código ───────────────────────────────────────────────
    if '_' in codigo:
        partes         = codigo.split('_')
        numero_parte   = partes[0]
        numero_carrito = partes[-1] if len(partes) > 1 else '?'
    elif '?' in codigo:
        partes         = codigo.split('?')
        numero_parte   = partes[0]
        numero_carrito = partes[3] if len(partes) > 3 else '?'
    else:
        numero_parte   = codigo
        numero_carrito = '?'

    turno      = get_turno_actual()
    fecha      = get_fecha_turno()
    hora_ahora = ahora_local().strftime("%H:%M:%S")

    async with AsyncSessionLocal() as db:

        # ── Buscar descripción en inventario ─────────────────────────
        result_inv = await db.execute(
            select(InventarioPlanta).where(
                InventarioPlanta.codigo == numero_parte
            )
        )
        inventario  = result_inv.scalar_one_or_none()
        descripcion = inventario.descripcion if inventario else ""
        maquina     = inventario.linea       if inventario else ""
        qty_total   = int(inventario.qtu or 0) if inventario else 0 

        # ── Verificar si ya completó ciclo ───────────────────────────
        result_completo = await db.execute(
            select(RegistroSecado).where(
                RegistroSecado.numero_parte == numero_parte,
                RegistroSecado.carrito      == numero_carrito,
                RegistroSecado.fecha        == fecha,
                RegistroSecado.turno        == turno,
                RegistroSecado.estado       == "salido"
            )
        )
        if result_completo.scalar_one_or_none():
            return {
                "tipo":    "ERROR",
                "mensaje": (
                    f"⚠️ Parte {numero_parte} · Carrito #{numero_carrito} "
                    f"ya completó su ciclo en este turno."
                )
            }

        # ── Buscar si está "dentro" ───────────────────────────────────
        result_dentro = await db.execute(
            select(RegistroSecado).where(
                RegistroSecado.numero_parte == numero_parte,
                RegistroSecado.carrito      == numero_carrito,
                RegistroSecado.fecha        == fecha,
                RegistroSecado.turno        == turno,
                RegistroSecado.estado       == "dentro"
            )
        )
        registro_dentro = result_dentro.scalar_one_or_none()

        if not registro_dentro:
            # Primera vez → ENTRADA
            nuevo = RegistroSecado(
                fecha        = fecha,
                turno        = turno,
                numero_parte = numero_parte,
                descripcion  = descripcion,
                maquina      = maquina, 
                carrito      = numero_carrito,
                hora_entrada = hora_ahora,
                qty_total    = qty_total,
                estado       = "dentro",
                usuario      = usuario,
                created_at   = ahora_local().replace(tzinfo=None)
            )
            db.add(nuevo)
            await db.commit()
            await db.refresh(nuevo)
            return {
                "tipo":    "ENTRADA",
                "mensaje": (
                    f"Parte {numero_parte} · Carrito #{numero_carrito} "
                    f"→ Entrada: {hora_ahora}"
                ),
                "registro": {
                    "id":               nuevo.id,
                    "numero_parte":     nuevo.numero_parte,
                    "descripcion":      nuevo.descripcion  or "",
                    "carrito":          nuevo.carrito,
                    "hora_entrada":     nuevo.hora_entrada,
                    "hora_salida":      None,
                    "tiempo_en_camara": None,
                    "estado":           nuevo.estado,
                    "usuario":          nuevo.usuario or "—",
                }
            }
        else:
            # ── SALIDA ───────────────────────────────────────────────
            tiempo = calcular_tiempo(registro_dentro.hora_entrada, hora_ahora)

            # ── Calcular total acumulado al salir ─────────────────────
            # Contar cuántos carritos de esta parte han SALIDO ya en este turno
            result_salidos = await db.execute(
                select(func.count(RegistroSecado.id))
                .where(RegistroSecado.numero_parte == numero_parte)
                .where(RegistroSecado.fecha        == fecha)
                .where(RegistroSecado.turno        == turno)
                .where(RegistroSecado.estado       == "salido")
            )
            carritos_salidos_antes = result_salidos.scalar() or 0
            # Este carrito es el siguiente → +1
            numero_carrito_orden   = carritos_salidos_antes + 1
            qty_unitario           = registro_dentro.qty_total or 0
            total_acumulado        = numero_carrito_orden * qty_unitario

            registro_dentro.hora_salida      = hora_ahora
            registro_dentro.tiempo_en_camara = tiempo
            registro_dentro.estado           = "salido"
            registro_dentro.qty_total        = total_acumulado   # ← acumulado

            # Completar maquina si faltaba
            if not registro_dentro.maquina:
                registro_dentro.maquina = maquina

            await db.commit()
            await db.refresh(registro_dentro)

            return {
                "tipo":    "SALIDA",
                "mensaje": (
                    f"Parte {numero_parte} · Carrito #{numero_carrito} "
                    f"→ Tiempo: {tiempo} · Total acumulado: {total_acumulado} pzs"
                ),
                "registro": {
                    "id":               registro_dentro.id,
                    "numero_parte":     registro_dentro.numero_parte,
                    "descripcion":      registro_dentro.descripcion  or "",
                    "maquina":          registro_dentro.maquina      or "—",
                    "carrito":          registro_dentro.carrito,
                    "hora_entrada":     registro_dentro.hora_entrada,
                    "hora_salida":      hora_ahora,
                    "tiempo_en_camara": tiempo,
                    "qty_total":        total_acumulado,
                    "estado":           "salido",
                    "usuario":          registro_dentro.usuario      or "—",
                }
            }


# ── GET /secado/registros/excel ──────────────────────────────────────
# Reemplaza el endpoint completo:

@router.get("/registros/excel")
@router.get("/registros/excel/")
async def descargar_excel_secado(
    fecha: Optional[str] = None,
    turno: Optional[str] = None,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Agrega 'openpyxl' a requirements.txt"
        )

    fecha_q = fecha or get_fecha_turno()
    turno_q = turno or get_turno_actual()

    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(RegistroSecado)
            .where(RegistroSecado.fecha == fecha_q)
            .where(RegistroSecado.turno == turno_q)
            .order_by(RegistroSecado.created_at.asc())
        )
        registros = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuarto de Secado"

    # ── Colores ──────────────────────────────────────────────────────
    COLOR_HEADER = "1E40AF"
    COLOR_FILA1  = "FFFFFF"
    COLOR_FILA2  = "EFF6FF"
    COLOR_DENTRO = "FEF3C7"
    COLOR_SALIDO = "DCFCE7"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font    = Font(name="Calibri", size=9)
    body_align_c = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin",   color="CBD5E1"),
        right=Side(style="thin",  color="CBD5E1"),
        top=Side(style="thin",    color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ── Encabezados (11 columnas ahora) ──────────────────────────────
    encabezados = [
        "Máquina",          # ← NUEVO (posición 1)
        "N° Parte",
        "Descripción",
        "Carrito",
        "Hora Entrada",
        "Hora Salida",
        "Tiempo en Cámara",
        "Total Piezas",     # ← NUEVO (antes de Estado)
        "Estado",
        "Usuario",
        "Turno",
    ]
    num_cols  = len(encabezados)
    col_letra = openpyxl.utils.get_column_letter(num_cols)

    # ── Título ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_letra}1")
    t           = ws["A1"]
    t.value     = "Control de Producción — Cuarto de Secado"
    t.font      = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_letra}2")
    i           = ws["A2"]
    i.value     = (
        f"Fecha: {fecha_q}   |   Turno: {turno_q}   |   "
        f"Total registros: {len(registros)}   |   "
        f"Generado: {ahora_local().strftime('%d/%m/%Y %H:%M:%S')}"
    )
    i.font      = Font(name="Calibri", size=9, color="64748B")
    i.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.append([])   # fila 3 vacía

    # ── Fila de encabezados ──────────────────────────────────────────
    ws.append(encabezados)
    fila_header = ws.max_row
    for col_idx in range(1, num_cols + 1):
        cell           = ws.cell(row=fila_header, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[fila_header].height = 30

    # ── Datos ────────────────────────────────────────────────────────
    for i, reg in enumerate(registros):
        estado_label = "🌡️ Dentro" if reg.estado == "dentro" else "✅ Salido"

        fila = [
            reg.maquina          or "—",          # Máquina
            reg.numero_parte,                      # N° Parte
            reg.descripcion      or "—",           # Descripción
            f"#{reg.carrito}",                     # Carrito
            reg.hora_entrada,                      # Hora Entrada
            reg.hora_salida      or "—",           # Hora Salida
            reg.tiempo_en_camara or "—",           # Tiempo en Cámara
            reg.qty_total        if reg.qty_total is not None else "—",  # Total Piezas
            estado_label,                          # Estado
            reg.usuario          or "—",           # Usuario
            reg.turno,                             # Turno
        ]
        ws.append(fila)

        fila_num = ws.max_row
        color_bg = (
            COLOR_DENTRO if reg.estado == "dentro"
            else COLOR_SALIDO if reg.estado == "salido"
            else (COLOR_FILA2 if i % 2 else COLOR_FILA1)
        )
        fill = PatternFill("solid", fgColor=color_bg)

        for col_idx in range(1, num_cols + 1):
            cell           = ws.cell(row=fila_num, column=col_idx)
            cell.font      = body_font
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = body_align_c

        ws.row_dimensions[fila_num].height = 18

    # ── Anchos ───────────────────────────────────────────────────────
    anchos = [18, 18, 35, 12, 14, 14, 20, 14, 14, 16, 10]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = ancho

    # ── Totales ───────────────────────────────────────────────────────
    ws.append([])
    total_dentro     = sum(1 for r in registros if r.estado == "dentro")
    total_salido     = sum(1 for r in registros if r.estado == "salido")
    total_piezas_sum = sum(r.qty_total or 0 for r in registros if r.estado == "salido")

    ws.append([
        f"Total: {len(registros)} registros",
        "", "", "", "", "",
        f"Dentro: {total_dentro}   |   Salidos: {total_salido}",
        f"Piezas salidas: {total_piezas_sum}",
        "", "", "",
    ])
    fila_total = ws.max_row
    for col_idx in range(1, num_cols + 1):
        cell           = ws.cell(row=fila_total, column=col_idx)
        cell.font      = Font(name="Calibri", bold=True, size=9, color=COLOR_HEADER)
        cell.alignment = body_align_c

    # ── Stream ────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"secado_{fecha_q}_{turno_q}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )