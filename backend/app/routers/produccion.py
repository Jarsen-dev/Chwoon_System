from fastapi import APIRouter, Depends, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Dict
from datetime import datetime, timedelta
from io import BytesIO

import pytz

from app.database import AsyncSessionLocal
from app.models.registro_produccion import RegistroProduccion
from app.models.plan_produccion     import PlanProduccion
from app.models.anomalia            import Anomalia
from app.models.inventario          import InventarioPlanta
from app.models.producto            import Producto
from app.models.registro_paro       import RegistroParo
from app.schemas.produccion import (
    RegistroProduccion     as RegistroSchema,
    RegistroProduccionCreate,
    PlanProduccion         as PlanSchema,
    PlanProduccionCreate,
    Anomalia               as AnomaliaSchema,
    AnomaliaCreate,
    RegistroParo           as RegistroParoSchema,
    RegistroParoCreate,
)

router = APIRouter(prefix="/produccion", tags=["produccion"])

TZ_LOCAL = pytz.timezone("America/Mexico_City")


# ── Helpers ───────────────────────────────────────────────────────────

async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


def _ahora_local() -> datetime:
    """Hora actual en America/Mexico_City (naive, para strftime)."""
    return datetime.now(TZ_LOCAL).replace(tzinfo=None)


def obtener_turno() -> str:
    ahora = _ahora_local().time()
    inicio = datetime.strptime("07:30", "%H:%M").time()
    fin    = datetime.strptime("19:30", "%H:%M").time()
    return "DIA" if inicio <= ahora < fin else "NOCHE"


def get_fecha_turno() -> str:
    ahora     = _ahora_local()
    total_min = ahora.hour * 60 + ahora.minute
    if total_min < 450:   # antes de las 07:30 → turno NOCHE empezó ayer
        return (ahora - timedelta(days=1)).strftime("%Y-%m-%d")
    return ahora.strftime("%Y-%m-%d")


# ==================== EXCEL ====================

@router.get("/registros/excel")
@router.get("/registros/excel/")
async def descargar_excel_registros(
    fecha: str = None,
    turno: str = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Descarga Excel con registros del escáner de producción.
    Ejemplo: /produccion/registros/excel?fecha=2025-01-15&turno=DIA
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Agrega 'openpyxl' a requirements.txt",
        )

    fecha_q = fecha or get_fecha_turno()
    turno_q = turno or obtener_turno()

    # ── Query ─────────────────────────────────────────────────────────
    query = (
        select(RegistroProduccion)
        .where(RegistroProduccion.fecha == fecha_q)
        .where(RegistroProduccion.turno == turno_q)
        .order_by(RegistroProduccion.hora.asc())
    )
    result    = await db.execute(query)
    registros = result.scalars().all()

    # ── Plan ──────────────────────────────────────────────────────────
    result_plan = await db.execute(select(PlanProduccion))
    mapa_plan   = {p.numero_parte: p.meta_piezas for p in result_plan.scalars().all()}

    # ── Workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros Producción"

    COLOR_HEADER = "1E40AF"
    COLOR_FILA1  = "FFFFFF"
    COLOR_FILA2  = "EFF6FF"
    COLOR_META   = "DCFCE7"
    COLOR_ALERTA = "FEF3C7"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font    = Font(name="Calibri", size=9)
    body_align_c = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left   = Side(style="thin", color="CBD5E1"),
        right  = Side(style="thin", color="CBD5E1"),
        top    = Side(style="thin", color="CBD5E1"),
        bottom = Side(style="thin", color="CBD5E1"),
    )

    encabezados = [
        "Hora", "Máquina", "N° Parte", "Descripción",
        "Carrito", "QTY", "Total Acum.", "Meta Plan", "Faltan", "Usuario", "Turno",
    ]
    num_cols  = len(encabezados)
    col_letra = openpyxl.utils.get_column_letter(num_cols)

    # ── Título ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_letra}1")
    t           = ws["A1"]
    t.value     = "Control de Producción — Registros de Escáner"
    t.font      = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_letra}2")
    inf           = ws["A2"]
    inf.value     = (
        f"Fecha: {fecha_q}   |   Turno: {turno_q}   |   "
        f"Total registros: {len(registros)}   |   "
        f"Generado: {_ahora_local().strftime('%d/%m/%Y %H:%M:%S')} CST"
    )
    inf.font      = Font(name="Calibri", size=9, color="64748B")
    inf.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.append([])   # fila 3 vacía

    # ── Encabezados ───────────────────────────────────────────────────
    ws.append(encabezados)
    fila_header = ws.max_row
    for col_idx in range(1, num_cols + 1):
        cell           = ws.cell(row=fila_header, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[fila_header].height = 30

    # ── Datos ─────────────────────────────────────────────────────────
    for i, reg in enumerate(registros):
        meta   = mapa_plan.get(reg.numero_parte)
        faltan = max(0, meta - reg.total_acumulado) if meta is not None else None

        ws.append([
            reg.hora,
            reg.maquina         or "—",
            reg.numero_parte,
            reg.descripcion     or "—",
            f"#{reg.carrito_numero}",
            reg.qty_bolsa,
            reg.total_acumulado,
            meta   if meta   is not None else "N/A",
            faltan if faltan is not None else "N/A",
            reg.usuario         or "—",
            reg.turno         or "—",
        ])

        fila_num = ws.max_row

        if meta is not None and reg.total_acumulado >= meta:
            color_bg = COLOR_META
        elif faltan is not None and meta and faltan <= meta * 0.1:
            color_bg = COLOR_ALERTA
        else:
            color_bg = COLOR_FILA2 if i % 2 else COLOR_FILA1

        fill = PatternFill("solid", fgColor=color_bg)
        for col_idx in range(1, num_cols + 1):
            cell           = ws.cell(row=fila_num, column=col_idx)
            cell.font      = body_font
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = body_align_c
        ws.row_dimensions[fila_num].height = 18

    # ── Anchos ────────────────────────────────────────────────────────
    for col_idx, ancho in enumerate([12, 18, 18, 35, 10, 10, 14, 12, 10, 16, 10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = ancho

    # ── Totales ───────────────────────────────────────────────────────
    ws.append([])
    total_piezas = sum(r.qty_bolsa or 0 for r in registros)
    ws.append([
        f"Total: {len(registros)} registros",
        "", "", "", "",
        f"Piezas producidas: {total_piezas}",
        "", "", "", "", "",
    ])
    fila_total = ws.max_row
    for col_idx in range(1, num_cols + 1):
        cell           = ws.cell(row=fila_total, column=col_idx)
        cell.font      = Font(name="Calibri", bold=True, size=9, color=COLOR_HEADER)
        cell.alignment = body_align_c

    # ── Stream ────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"produccion_{fecha_q}_{turno_q}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":            f"attachment; filename={filename}",
            "Access-Control-Expose-Headers":  "Content-Disposition",
        },
    )


# ==================== REGISTROS ====================

@router.get("/registros/", response_model=List[RegistroSchema])
@router.get("/registros",  response_model=List[RegistroSchema])
async def listar_registros(
    fecha: str = None,
    turno: str = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(RegistroProduccion)
    if fecha:
        query = query.where(RegistroProduccion.fecha == fecha)
    if turno:
        query = query.where(RegistroProduccion.turno == turno)
    query = query.order_by(
        RegistroProduccion.fecha.desc(),
        RegistroProduccion.hora.desc(),
    )
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/registros/", response_model=RegistroSchema)
@router.post("/registros",  response_model=RegistroSchema)
async def crear_registro(
    registro: RegistroProduccionCreate,
    db: AsyncSession = Depends(get_db),
):
    db_registro = RegistroProduccion(**registro.model_dump())
    db.add(db_registro)
    await db.commit()
    await db.refresh(db_registro)
    return db_registro


# ==================== WEBSOCKET ====================

class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket]        = []
        self.ultimo_escaneo:     Dict[str, datetime]    = {}
        self.historial_tiempos:  Dict[str, List[float]] = {}
        self.tiempos_maquina:    Dict[str, datetime]    = {}
        self.gaps_maquina:       Dict[str, List[float]] = {}

    async def connect(self, websocket: WebSocket):
        await websocket.accept(headers=[(b"access-control-allow-origin", b"*")])
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: Dict):
        for conn in self.active_connections:
            try:
                await conn.send_json(message)
            except Exception:
                pass


manager = ConnectionManager()


async def _calcular_conteo(
    db: AsyncSession,
    numero_parte: str,
    fecha_str: str,
    turno: str,
) -> int:
    result = await db.execute(
        select(func.count(RegistroProduccion.id))
        .where(RegistroProduccion.numero_parte == numero_parte)
        .where(RegistroProduccion.fecha        == fecha_str)
        .where(RegistroProduccion.turno        == turno)
    )
    return (result.scalar() or 0) + 1


async def _calcular_total_acumulado(
    db: AsyncSession,
    numero_parte: str,
    fecha_str: str,
    turno: str,
    qty_bolsa: int,
) -> int:
    result = await db.execute(
        select(func.max(RegistroProduccion.total_acumulado))
        .where(RegistroProduccion.numero_parte == numero_parte)
        .where(RegistroProduccion.fecha        == fecha_str)
        .where(RegistroProduccion.turno        == turno)
    )
    return (result.scalar() or 0) + qty_bolsa


async def _detectar_anomalia_ia(
    db: AsyncSession,
    numero_parte: str,
    diff_segundos: float,
    fecha_str: str,
    hora_str: str,
) -> dict | None:
    try:
        import numpy as np
        from sklearn.ensemble import IsolationForest

        historial = manager.historial_tiempos.get(numero_parte, [])
        if len(historial) < 5:
            return None

        X            = np.array(historial).reshape(-1, 1)
        modelo       = IsolationForest(contamination=0.1, random_state=42)
        modelo.fit(X)
        prediccion   = modelo.predict([[diff_segundos]])
        media_normal = float(np.mean(X))

        if prediccion[0] == -1 and diff_segundos < (media_normal * 0.3):
            motivo = (
                f"IA Detectó Anomalía: Escaneo inusualmente rápido "
                f"({int(diff_segundos)} seg). Tiempo normal: {int(media_normal)} seg."
            )
            db.add(Anomalia(
                fecha        = fecha_str,
                hora         = hora_str,
                numero_parte = numero_parte,
                motivo       = motivo,
                tipo         = "FRAUDE",
                created_at   = datetime.utcnow(),
            ))
            await db.commit()
            return {"tipo": "FRAUDE", "motivo": motivo}

    except Exception as e:
        print(f"Error IA anomalías: {e}")
    return None


async def _detectar_fatiga_maquina(
    db: AsyncSession,
    maquina: str,
    gap_segundos: float,
    fecha_str: str,
    hora_str: str,
) -> dict | None:
    try:
        import numpy as np

        gaps = manager.gaps_maquina.get(maquina, [])
        if len(gaps) < 5:
            return None

        y            = np.array(gaps)
        x            = np.arange(len(y))
        pendiente, _ = np.polyfit(x, y, 1)

        if pendiente > 5:
            motivo = (
                f"Lentitud progresiva detectada en {maquina}. "
                f"Posible atasco o fatiga mecánica. "
                f"Pendiente: +{pendiente:.1f} seg/ciclo"
            )
            db.add(Anomalia(
                fecha        = fecha_str,
                hora         = hora_str,
                numero_parte = "MANTENIMIENTO",
                motivo       = motivo,
                tipo         = "MANTENIMIENTO",
                created_at   = datetime.utcnow(),
            ))
            await db.commit()
            return {"tipo": "MANTENIMIENTO", "motivo": motivo, "maquina": maquina}

    except Exception as e:
        print(f"Error IA mantenimiento: {e}")
    return None


@router.websocket("/ws/scanner")
async def websocket_scanner(
    websocket: WebSocket,
    token: str = None,
):
    
    manager.active_connections.append(websocket)

    # ── Decodificar usuario del token ─────────────────────────────────
    usuario_ws = "desconocido"
    if token:
        try:
            from app.core.security import decode_token
            payload = decode_token(token)
            if payload:
                usuario_ws = payload.get("sub") or "desconocido"
            print(f"✅ WS conectado | usuario: {usuario_ws}")
        except Exception as e:
            print(f"Error decodificando token WS: {e}")
    else:
        print("⚠️  WS conectado SIN token")

    try:
        while True:
            data       = await websocket.receive_text()
            codigo_raw = data.strip()

            numero_carrito_qr = None

            if '?' in codigo_raw:
                partes_qr = codigo_raw.split('?')
                codigo = partes_qr[0].strip().upper()
                if len(partes_qr) >= 4:
                    try:
                        numero_carrito_qr = int(partes_qr[3].strip())
                    except (ValueError, IndexError):
                        numero_carrito_qr = None
            else:
                codigo = codigo_raw.strip().upper()

            if not codigo:
                continue

            async with AsyncSessionLocal() as db:

                # ── Buscar en productos (primero) o inventario (fallback) ──
                result = await db.execute(
                    select(Producto).where(Producto.sku == codigo)
                )
                producto_found = result.scalar_one_or_none()

                if producto_found:
                    inventario_codigo      = producto_found.sku
                    inventario_descripcion = producto_found.nombre or producto_found.descripcion or ""
                    inventario_linea       = producto_found.linea_produccion or "Sin máquina"
                    inventario_qtu         = producto_found.cantidad_carrito or 1
                else:
                    result = await db.execute(
                        select(InventarioPlanta).where(InventarioPlanta.codigo == codigo)
                    )
                    inv_item = result.scalar_one_or_none()

                    if not inv_item:
                        await websocket.send_json({
                            "type":    "error",
                            "message": f"Código '{codigo}' no encontrado",
                        })
                        continue

                    inventario_codigo      = inv_item.codigo
                    inventario_descripcion = inv_item.descripcion or ""
                    inventario_linea       = inv_item.linea or "Sin máquina"
                    inventario_qtu         = int(inv_item.qtu or 1)

                # ── Hora y turno LOCAL ────────────────────────────────
                ahora     = _ahora_local()
                fecha_str = get_fecha_turno()
                hora_str  = ahora.strftime("%H:%M:%S")
                turno     = obtener_turno()

                # ── Verifica duplicado ────────────────────────────────
                if numero_carrito_qr is not None:
                    result_dup = await db.execute(
                        select(RegistroProduccion).where(
                            RegistroProduccion.numero_parte   == codigo,
                            RegistroProduccion.carrito_numero == numero_carrito_qr,
                            RegistroProduccion.fecha          == fecha_str,
                            RegistroProduccion.turno          == turno,
                        )
                    )
                    if result_dup.scalar_one_or_none():
                        await websocket.send_json({
                            "type":    "error",
                            "message": (
                                f"⚠️ Parte {codigo} · Carrito #{numero_carrito_qr} "
                                f"ya fue escaneado en este turno ({turno})."
                            ),
                        })
                        continue

                qty_bolsa = inventario_qtu

                carrito_numero = (
                    numero_carrito_qr
                    if numero_carrito_qr is not None
                    else await _calcular_conteo(db, codigo, fecha_str, turno)
                )

                total_acumulado = await _calcular_total_acumulado(
                    db, codigo, fecha_str, turno, qty_bolsa
                )

                registro = RegistroProduccion(
                    fecha           = fecha_str,
                    hora            = hora_str,
                    turno           = turno,
                    maquina         = inventario_linea,
                    numero_parte    = inventario_codigo,
                    descripcion     = inventario_descripcion,
                    carrito_numero  = carrito_numero,
                    qty_bolsa       = qty_bolsa,
                    total_acumulado = total_acumulado,
                    usuario         = usuario_ws,
                    created_at      = datetime.utcnow(),
                )
                db.add(registro)
                await db.commit()

                # ── Actualizar estado del plan ─────────────────────────
                result_plan = await db.execute(
                    select(PlanProduccion).where(
                        PlanProduccion.numero_parte == codigo
                    )
                )
                plan_item = result_plan.scalar_one_or_none()
                if plan_item and plan_item.estado == "pendiente":
                    plan_item.estado = "en_proceso"
                    await db.commit()

                # ── Análisis IA ───────────────────────────────────────
                alertas = []

                if codigo in manager.ultimo_escaneo:
                    diff = (ahora - manager.ultimo_escaneo[codigo]).total_seconds()
                    if codigo not in manager.historial_tiempos:
                        manager.historial_tiempos[codigo] = []
                    manager.historial_tiempos[codigo].append(diff)
                    if len(manager.historial_tiempos[codigo]) > 50:
                        manager.historial_tiempos[codigo].pop(0)
                    alerta_fraude = await _detectar_anomalia_ia(
                        db, codigo, diff, fecha_str, hora_str
                    )
                    if alerta_fraude:
                        alertas.append(alerta_fraude)

                manager.ultimo_escaneo[codigo] = ahora

                maquina = inventario_linea
                if maquina != "Sin máquina":
                    if maquina in manager.tiempos_maquina:
                        gap = (ahora - manager.tiempos_maquina[maquina]).total_seconds()
                        if maquina not in manager.gaps_maquina:
                            manager.gaps_maquina[maquina] = []
                        manager.gaps_maquina[maquina].append(gap)
                        if len(manager.gaps_maquina[maquina]) > 8:
                            manager.gaps_maquina[maquina].pop(0)
                        alerta_maquina = await _detectar_fatiga_maquina(
                            db, maquina, gap, fecha_str, hora_str
                        )
                        if alerta_maquina:
                            alertas.append(alerta_maquina)
                    manager.tiempos_maquina[maquina] = ahora

                # ── Respuesta al cliente ──────────────────────────────
                await websocket.send_json({
                    "type": "scan_complete",
                    "registro": {
                        "hora":            hora_str,
                        "maquina":         registro.maquina,
                        "numero_parte":    registro.numero_parte,
                        "descripcion":     registro.descripcion  or "",
                        "carrito_numero":  registro.carrito_numero,
                        "qty_bolsa":       registro.qty_bolsa,
                        "total_acumulado": registro.total_acumulado,
                        "turno":           turno,
                        "fecha":           fecha_str,
                        "usuario":         usuario_ws,
                    },
                    "alertas": alertas,
                })

                await manager.broadcast({
                    "type": "update",
                    "data": "nuevo_escaneo",
                })

    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        print(f"Error WebSocket: {e}")
        manager.disconnect(websocket)


# ==================== PLAN ====================

@router.get("/plan/", response_model=List[PlanSchema])
@router.get("/plan",  response_model=List[PlanSchema])
async def obtener_plan(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(PlanProduccion))
    return result.scalars().all()


@router.post("/plan/", response_model=PlanSchema)
@router.post("/plan",  response_model=PlanSchema)
async def agregar_al_plan(
    plan: PlanProduccionCreate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PlanProduccion).where(PlanProduccion.numero_parte == plan.numero_parte)
    )
    existente = result.scalar_one_or_none()

    if existente:
        existente.meta_piezas    = plan.meta_piezas
        existente.turno_objetivo = plan.turno_objetivo
        await db.commit()
        await db.refresh(existente)
        return existente

    db_plan = PlanProduccion(**plan.model_dump(), created_at=datetime.utcnow())
    db.add(db_plan)
    await db.commit()
    await db.refresh(db_plan)
    return db_plan


@router.delete("/plan/{numero_parte}")
async def eliminar_del_plan(
    numero_parte: str,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PlanProduccion).where(PlanProduccion.numero_parte == numero_parte)
    )
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="No encontrado en plan")
    await db.delete(plan)
    await db.commit()
    return {"message": "Eliminado del plan"}


# ==================== PROYECCIÓN ====================

@router.get("/proyeccion/{turno}")
async def obtener_proyeccion(
    turno: str,
    db: AsyncSession = Depends(get_db),
):
    ahora       = _ahora_local()
    fecha_turno = get_fecha_turno()
    turno_upper = turno.upper()

    if turno_upper == "DIA":
        inicio_turno = ahora.replace(hour=7,  minute=30, second=0, microsecond=0)
        fin_turno    = ahora.replace(hour=19, minute=30, second=0, microsecond=0)
    else:
        hora_actual = ahora.time()
        if hora_actual >= datetime.strptime("19:30", "%H:%M").time():
            inicio_turno = ahora.replace(hour=19, minute=30, second=0, microsecond=0)
            fin_turno    = (ahora + timedelta(days=1)).replace(
                hour=7, minute=30, second=0, microsecond=0
            )
        else:
            inicio_turno = (ahora - timedelta(days=1)).replace(
                hour=19, minute=30, second=0, microsecond=0
            )
            fin_turno = ahora.replace(hour=7, minute=30, second=0, microsecond=0)

    tiempo_transcurrido_h = max(
        (ahora - inicio_turno).total_seconds() / 3600.0, 0.01
    )
    tiempo_total_h  = (fin_turno - inicio_turno).total_seconds() / 3600.0
    horas_restantes = max(0, tiempo_total_h - tiempo_transcurrido_h)

    result = await db.execute(
        select(
            RegistroProduccion.numero_parte,
            func.max(RegistroProduccion.total_acumulado).label("total"),
        )
        .where(RegistroProduccion.fecha == fecha_turno)
        .where(RegistroProduccion.turno == turno_upper)
        .group_by(RegistroProduccion.numero_parte)
    )
    produccion_actual = {row.numero_parte: row.total for row in result.all()}

    result_plan = await db.execute(select(PlanProduccion))
    planes      = result_plan.scalars().all()

    proyecciones     = []
    alertas_lentitud = []

    for plan in planes:
        producido      = produccion_actual.get(plan.numero_parte, 0)
        faltan         = max(0, plan.meta_piezas - producido)
        ritmo_por_hora = round(producido / tiempo_transcurrido_h, 1)

        if producido >= plan.meta_piezas and plan.meta_piezas > 0:
            tiempo_estimado = "✅ Completado"
        elif ritmo_por_hora > 0:
            horas_estimadas = faltan / ritmo_por_hora
            tiempo_estimado = f"{horas_estimadas:.1f} hrs"
            if horas_estimadas > horas_restantes:
                alertas_lentitud.append({
                    "numero_parte": plan.numero_parte,
                    "motivo": (
                        f"Producción Lenta: Ritmo actual ({int(ritmo_por_hora)} pz/h). "
                        f"Faltan {faltan} pz, tomará {horas_estimadas:.1f}h, "
                        f"pero al turno le quedan {horas_restantes:.1f}h."
                    ),
                    "tipo": "LENTITUD_PLAN",
                })
        else:
            tiempo_estimado = "Sin datos"

        proyecciones.append({
            "numero_parte":    plan.numero_parte,
            "producido":       producido,
            "ritmo_por_hora":  ritmo_por_hora,
            "meta_plan":       plan.meta_piezas,
            "faltan":          faltan,
            "tiempo_estimado": tiempo_estimado,
        })

    return {
        "turno":           turno_upper,
        "horas_restantes": round(horas_restantes, 1),
        "proyecciones":    proyecciones,
        "alertas_plan":    alertas_lentitud,
    }


# ==================== SALUD MÁQUINAS ====================

@router.get("/salud-maquinas/")
@router.get("/salud-maquinas")
async def salud_maquinas():
    try:
        import numpy as np
    except ImportError:
        return []

    resultado = []
    for maquina, gaps in manager.gaps_maquina.items():
        if len(gaps) < 5:
            continue

        y            = np.array(gaps)
        x            = np.arange(len(y))
        pendiente, _ = np.polyfit(x, y, 1)
        ultimo_gap   = int(gaps[-1])
        tendencia    = f"{pendiente:+.1f} seg/ciclo"

        if   pendiente > 5:  estado = "🔴 RIESGO (Atasco/Fatiga)"
        elif pendiente > 2:  estado = "🟡 Perdiendo Velocidad"
        elif pendiente < -2: estado = "🔵 Acelerando"
        else:                estado = "🟢 Estable"

        resultado.append({
            "maquina":               maquina,
            "ultimo_ciclo_segundos": ultimo_gap,
            "tendencia":             tendencia,
            "estado":                estado,
        })

    return resultado


# ==================== ANOMALÍAS ====================

@router.get("/anomalias/", response_model=List[AnomaliaSchema])
@router.get("/anomalias",  response_model=List[AnomaliaSchema])
async def listar_anomalias(
    limite: int = 100,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Anomalia)
        .order_by(Anomalia.fecha.desc(), Anomalia.hora.desc())
        .limit(limite)
    )
    return result.scalars().all()


@router.post("/anomalias/", response_model=AnomaliaSchema)
@router.post("/anomalias",  response_model=AnomaliaSchema)
async def registrar_anomalia(
    anomalia: AnomaliaCreate,
    db: AsyncSession = Depends(get_db),
):
    db_anomalia = Anomalia(**anomalia.model_dump(), created_at=datetime.utcnow())
    db.add(db_anomalia)
    await db.commit()
    await db.refresh(db_anomalia)
    return db_anomalia


# ==================== PAROS ====================

@router.get("/paros/", response_model=List[RegistroParoSchema])
@router.get("/paros",  response_model=List[RegistroParoSchema])
async def listar_paros(
    fecha: str = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(RegistroParo)
    if fecha:
        query = query.where(RegistroParo.fecha == fecha)
    query = query.order_by(RegistroParo.fecha.desc())
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/paros/", response_model=RegistroParoSchema)
@router.post("/paros",  response_model=RegistroParoSchema)
async def registrar_paro(
    paro: RegistroParoCreate,
    db: AsyncSession = Depends(get_db),
):
    db_paro = RegistroParo(**paro.model_dump(), created_at=datetime.utcnow())
    db.add(db_paro)
    await db.commit()
    await db.refresh(db_paro)
    return db_paro