import os
import io
import qrcode
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Path
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, and_, extract, cast, Date
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
from typing import List, Optional, Any
from datetime import datetime, timezone, timedelta, date
from fastapi.responses import StreamingResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter

from app.core.deps import get_current_user, get_db
from app.models.orden_compra import OrdenCompra, OrdenCompraItem, Proveedor, ProveedorMaterial, RecepcionCompra, ProveedorEvento
from app.services.proveedor_score import recalcular_score, registrar_evento
from app.models.orden_venta import OrdenVenta, OrdenVentaItem, EnvioVenta
from app.models.devolucion import Devolucion
from app.models.plan_ventas import PlanVentas
from app.models.psi_snapshot import PsiSnapshot
from app.models.empresa  import ConfiguracionEmpresa, ContactoEmpresa
from app.services.pdf_generator import generar_pdf_orden_compra
from app.core.deps import get_db, get_current_user, get_current_compras, get_current_finanzas
from app.schemas.finanzas import (
    OrdenCompraCreate, OrdenCompraUpdate, OrdenCompraResponse, ProveedorCreate, ProveedorResponse, ProveedorUpdate,
    RecepcionCompraCreate, RecepcionCompraResponse,
    OrdenVentaCreate, OrdenVentaUpdate, OrdenVentaResponse,
    EnvioVentaCreate,
    DevolucionCreate, DevolucionResponse, DisposicionDevolucionCreate,
    PlanVentasImport, PlanVentasResponse, AutorizarVentasMasivo,
    FinanzasDashboardResponse,
    AprobarOrdenCompraRequest,
    ProveedorScoreResponse, ProveedorEventoCreate, ProveedorEventoResponse, 
    EnvioLogisticoCreate, ValidarFinanzasRequest,
    CambiarEstadoRequest, DespacharRequest
)

TZ_LOCAL = timezone(timedelta(hours=-6))

router = APIRouter(prefix="/finanzas", tags=["Finanzas"])


# ─── Máquina de estados ───────────────────────────────────────────────────────
#
# Transiciones válidas y quién puede ejecutarlas.
# Cualquier intento de transición fuera de esta tabla devuelve 422.
#
TRANSICIONES: dict[str, dict[str, list[str]]] = {
    "En Preparación": {
        "desde":   ["Pendiente de Envío", "Stock Insuficiente"],
        "roles":   ["admin", "almacen", "logistica"],
    },
    "Lista para Carga": {
        "desde":   ["En Preparación"],
        "roles":   ["admin", "almacen"],
    },
    # Rollback explícito: Almacén detecta problema en el andén
    "En Preparación_rollback": {
        "desde":   ["Lista para Carga"],
        "roles":   ["admin", "almacen"],
    },
    "Cancelada": {
        # Puede cancelarse desde cualquier estado pre-despacho
        "desde":   [
            "Pendiente de Envío",
            "Stock Insuficiente",
            "En Preparación",
            "Lista para Carga",
        ],
        "roles":   ["admin", "finanzas", "ventas"],
    },
}
 
# Estados terminales — no se puede hacer nada desde ellos
ESTADOS_TERMINALES = {"Enviado", "Cancelada"}



# ========================
# HELPERS
# ========================
def ahora_local():
    return datetime.now(TZ_LOCAL)


def require_finanzas_role(current_user):
    """Admin y finanzas tienen acceso total (compras + ventas)."""
    if current_user.rol not in ("admin", "finanzas"):
        raise HTTPException(status_code=403, detail="Acceso denegado. Se requiere rol admin o finanzas.")
    return current_user


def require_compras_role(current_user):
    """Admin, finanzas y compras pueden acceder a endpoints de compras."""
    if current_user.rol not in ("admin", "finanzas", "compras"):
        raise HTTPException(status_code=403, detail="Acceso denegado. Se requiere rol admin, finanzas o compras.")
    return current_user


def require_ventas_role(current_user):
    """Admin, finanzas y ventas pueden acceder a endpoints de ventas."""
    if current_user.rol not in ("admin", "finanzas", "ventas"):
        raise HTTPException(status_code=403, detail="Acceso denegado. Se requiere rol admin, finanzas o ventas.")
    return current_user

async def _get_ov(db: AsyncSession, ov_id: str) -> OrdenVenta:
    result = await db.execute(
        select(OrdenVenta).where(OrdenVenta.ov_id == ov_id)
    )
    ov = result.scalar_one_or_none()
    if not ov:
        raise HTTPException(404, f"Orden de venta {ov_id!r} no encontrada")
    return ov
 
 
def _assert_transicion(ov: OrdenVenta, nuevo_estado: str, rol: str) -> None:
    """Valida que la transición sea legal. Lanza 422 si no lo es."""
 
    if ov.estado in ESTADOS_TERMINALES:
        raise HTTPException(
            422,
            f"La OV {ov.ov_id} está en estado terminal '{ov.estado}' "
            f"y no puede cambiar de estado.",
        )
 
    regla = TRANSICIONES.get(nuevo_estado) or TRANSICIONES.get(f"{nuevo_estado}_rollback")
    if not regla:
        raise HTTPException(422, f"Estado destino '{nuevo_estado}' no es válido.")
 
    if ov.estado not in regla["desde"]:
        raise HTTPException(
            422,
            f"No se puede pasar de '{ov.estado}' a '{nuevo_estado}'. "
            f"Estados permitidos como origen: {regla['desde']}",
        )
 
    if rol not in regla["roles"]:
        raise HTTPException(
            403,
            f"El rol '{rol}' no tiene permiso para mover a '{nuevo_estado}'. "
            f"Roles permitidos: {regla['roles']}",
        )


# ========================
# DASHBOARD
# ========================
@router.get("/dashboard", response_model=FinanzasDashboardResponse)
@router.get("/dashboard/", response_model=FinanzasDashboardResponse)
async def calcular_finanzas_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> FinanzasDashboardResponse:
    """
    Calcula todos los KPIs del dashboard de ventas/finanzas.
    Cada sección está aislada en su propio bloque — añade o quita sin tocar el resto.
    """
    hoy      = date.today()
    mes_ini  = hoy.replace(day=1)
    ahora_tz = datetime.now(TZ_LOCAL)
 
    # ── 1. OVs por estado ────────────────────────────────────────────────────
    ov_por_estado = (
        await db.execute(
            select(OrdenVenta.estado, func.count(OrdenVenta.id).label("cnt"))
            .group_by(OrdenVenta.estado)
        )
    ).all()
    estado_map: dict[str, int] = {row.estado: row.cnt for row in ov_por_estado}
 
    total_ov              = sum(estado_map.values())
    ov_pendientes         = estado_map.get("Pendiente de Envío", 0)
    ov_en_preparacion     = estado_map.get("En Preparación", 0)
    ov_lista_para_carga   = estado_map.get("Lista para Carga", 0)
    ov_enviadas           = estado_map.get("Enviado", 0) + estado_map.get("Embarque Parcial", 0)
    ov_stock_insuficiente = estado_map.get("Stock Insuficiente", 0)
 
    # ── 2. Devoluciones ───────────────────────────────────────────────────────
    # Importa Devolucion desde app.models si existe; ajusta el nombre si es diferente
    try:
        from app.models.devolucion import Devolucion
        dev_rows = (
            await db.execute(
                select(Devolucion.estado_inspeccion, func.count(Devolucion.id).label("cnt"))
                .group_by(Devolucion.estado_inspeccion)
            )
        ).all()
        dev_map               = {r.estado_inspeccion: r.cnt for r in dev_rows}
        total_devoluciones    = sum(dev_map.values())
        devoluciones_pendientes = dev_map.get("Pendiente", 0)
    except ImportError:
        total_devoluciones = devoluciones_pendientes = 0
 
    # ── 3. Plan de ventas activos ─────────────────────────────────────────────
    planes_activos_row = (
        await db.execute(
            select(func.count(PlanVentas.id))
            .where(PlanVentas.fecha_inicio_semana >= mes_ini)
        )
    ).scalar_one()
    planes_venta_activos = planes_activos_row or 0
 
    # ── 4. Valor ventas del mes ───────────────────────────────────────────────
    # Suma precio_unitario × cantidad_enviada de items en OVs enviadas este mes
    valor_mes_row = (
        await db.execute(
            select(
                func.coalesce(
                    func.sum(OrdenVentaItem.precio_unitario * OrdenVentaItem.cantidad_enviada),
                    0,
                )
            )
            .join(OrdenVenta, OrdenVenta.id == OrdenVentaItem.orden_venta_id)
            .where(
                and_(
                    OrdenVenta.estado.in_(["Enviado", "Embarque Parcial"]),
                    cast(OrdenVenta.fecha_actualizacion, Date) >= mes_ini,
                )
            )
        )
    ).scalar_one()
    valor_ventas_mes = float(valor_mes_row or 0)
 
    # ── 5. KPIs operativos del día ────────────────────────────────────────────
    programado_hoy    = 0
    embarcado_hoy     = 0
    skus_dif_negativa = 0
 
    # 5a. Programado hoy — buscar el plan de la semana activa y sumar el día de hoy
    dia_hoy = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"][hoy.weekday()]
 
    plan_activo = (
        await db.execute(
            select(PlanVentas)
            .where(PlanVentas.fecha_inicio_semana <= hoy)
            .order_by(PlanVentas.fecha_inicio_semana.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
 
    if plan_activo and plan_activo.items:
        for item in plan_activo.items:
            dia_data  = (item.get("dias") or {}).get(dia_hoy, {})
            programado_hoy += dia_data.get("plan", 0)
 
            # DIF negativa: stock_lg < plan acumulado hasta hoy
            stock_lg = item.get("stock_lg", 0) or 0
            ORDEN_DIAS = ["VIERNES", "LUNES", "MARTES", "MIERCOLES", "JUEVES"]
            acumulado = 0
            for dia in ORDEN_DIAS:
                acumulado += (item.get("dias") or {}).get(dia, {}).get("plan", 0)
                if dia == dia_hoy:
                    break
            if stock_lg - acumulado < 0:
                skus_dif_negativa += 1
 
    # 5b. Embarcado hoy — suma de envíos con fecha = hoy y status_salida = OK
    embarcado_row = (
        await db.execute(
            select(
                func.coalesce(
                    func.sum(
                        func.jsonb_array_length(EnvioVenta.items_enviados)
                        # ↑ Aproximación: si necesitas sumar cantidades exactas,
                        #   usa una subquery que expande el JSON.
                        #   Por ahora, cuenta el número de líneas enviadas hoy.
                    ),
                    0,
                )
            )
            .where(
                and_(
                    cast(EnvioVenta.fecha_envio, Date) == hoy,
                    EnvioVenta.status_salida == "OK",
                )
            )
        )
    ).scalar_one()
 
    # ── Cálculo exacto de embarcado_hoy (suma de cantidades en el JSON) ──────
    # El approach con jsonb_array_length solo cuenta líneas, no cantidades.
    # Para cantidades exactas hacemos una query más directa:
    envios_hoy = (
        await db.execute(
            select(EnvioVenta.items_enviados)
            .where(
                and_(
                    cast(EnvioVenta.fecha_envio, Date) == hoy,
                    EnvioVenta.status_salida == "OK",
                )
            )
        )
    ).scalars().all()
 
    for items_json in envios_hoy:
        if isinstance(items_json, list):
            for ie in items_json:
                embarcado_hoy += ie.get("cantidad", 0)
 
    # ── 6. PSI Coverage ──────────────────────────────────────────────────────
    # Primero intentar usar el snapshot oficial del PLAN EMBARQUE para hoy.
    psi_hoy = (
        await db.execute(
            select(PsiSnapshot).where(PsiSnapshot.fecha == hoy)
        )
    ).scalar_one_or_none()
    # La cobertura PSI se calcula como:
    #   coverage = stock_lg / demanda_diaria
    # agrupada por categoría (Ref = línea R1/R2, Oven = línea R3 u otro indicador
    # en tu Excel). Ajusta el filtro según cómo distingues Ref vs Oven en tus datos.
    #
    # D-Day  = hoy
    # D+1    = mañana
 
    coverage_ref_dday  = 0.0
    coverage_ref_d1    = 0.0
    coverage_oven_dday = 0.0
    coverage_oven_d1   = 0.0

    if psi_hoy:
        # Snapshot oficial del PLAN EMBARQUE — tiene prioridad sobre el cálculo
        coverage_ref_dday  = psi_hoy.coverage_ref_dday
        coverage_ref_d1    = psi_hoy.coverage_ref_d1
        coverage_oven_dday = psi_hoy.coverage_oven_dday
        coverage_oven_d1   = psi_hoy.coverage_oven_d1
    elif plan_activo and plan_activo.items:
        ORDEN_DIAS    = ["VIERNES", "LUNES", "MARTES", "MIERCOLES", "JUEVES"]
        dia_idx       = ORDEN_DIAS.index(dia_hoy) if dia_hoy in ORDEN_DIAS else -1
        dia_siguiente = ORDEN_DIAS[dia_idx + 1] if dia_idx >= 0 and dia_idx + 1 < len(ORDEN_DIAS) else None
 
        # Agrupa por categoría — ajusta "linea" o "id1" según tu estructura
        def _categoria(item: dict) -> str:
            linea = (item.get("linea") or "").upper()
            id1   = (item.get("id1")   or "").upper()
            # Heurística: si la descripción o id1 contiene "OVEN" es Oven; el resto es Ref
            if "OVEN" in id1 or "OVEN" in item.get("descripcion", "").upper():
                return "OVEN"
            return "REF"
 
        stock_ref   = stock_plan_ref_hoy   = stock_plan_ref_d1   = 0
        stock_oven  = stock_plan_oven_hoy  = stock_plan_oven_d1  = 0
 
        for item in plan_activo.items:
            cat      = _categoria(item)
            stock_lg = item.get("stock_lg", 0) or 0
            dias     = item.get("dias") or {}
 
            plan_hoy = dias.get(dia_hoy, {}).get("plan", 0) if dia_hoy else 0
            plan_d1  = dias.get(dia_siguiente, {}).get("plan", 0) if dia_siguiente else 0
 
            if cat == "REF":
                stock_ref          += stock_lg
                stock_plan_ref_hoy += plan_hoy
                stock_plan_ref_d1  += plan_d1
            else:
                stock_oven          += stock_lg
                stock_plan_oven_hoy += plan_hoy
                stock_plan_oven_d1  += plan_d1
 
        coverage_ref_dday  = round(stock_ref  / stock_plan_ref_hoy,  3) if stock_plan_ref_hoy  > 0 else 1.0
        coverage_ref_d1    = round(stock_ref  / stock_plan_ref_d1,   3) if stock_plan_ref_d1   > 0 else 1.0
        coverage_oven_dday = round(stock_oven / stock_plan_oven_hoy, 3) if stock_plan_oven_hoy > 0 else 1.0
        coverage_oven_d1   = round(stock_oven / stock_plan_oven_d1,  3) if stock_plan_oven_d1  > 0 else 1.0
 
        # Acota a máximo 2.0 (200%) para no inflar KPIs
        coverage_ref_dday  = min(coverage_ref_dday,  2.0)
        coverage_ref_d1    = min(coverage_ref_d1,    2.0)
        coverage_oven_dday = min(coverage_oven_dday, 2.0)
        coverage_oven_d1   = min(coverage_oven_d1,   2.0)
 
    # ── Ensamblar respuesta ───────────────────────────────────────────────────
    return FinanzasDashboardResponse(
        total_ov=total_ov,
        ov_pendientes=ov_pendientes,
        ov_en_preparacion=ov_en_preparacion,
        ov_lista_para_carga=ov_lista_para_carga,
        ov_enviadas=ov_enviadas,
        ov_stock_insuficiente=ov_stock_insuficiente,
        total_devoluciones=total_devoluciones,
        devoluciones_pendientes=devoluciones_pendientes,
        planes_venta_activos=planes_venta_activos,
        valor_ventas_mes=valor_ventas_mes,
        programado_hoy=programado_hoy,
        embarcado_hoy=embarcado_hoy,
        skus_dif_negativa=skus_dif_negativa,
        coverage_ref_dday=coverage_ref_dday,
        coverage_ref_d1=coverage_ref_d1,
        coverage_oven_dday=coverage_oven_dday,
        coverage_oven_d1=coverage_oven_d1,
    )


# ========================
# PLAN EMBARQUE — IMPORT
# ========================
@router.post("/plan-embarque/importar")
@router.post("/plan-embarque/importar/")
async def importar_plan_embarque(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Importa la hoja PSI RESUME del archivo PLAN EMBARQUE de LG.

    Estructura esperada (hoja 'PSI RESUME'):
      - Fila 10: Ref  → col C=Need D-Day, col D=Covered D-Day, col E=% D-Day,
                         col F=Need D+1,  col G=Covered D+1,   col H=% D+1
      - Fila 11: Oven → misma estructura

    Hace upsert en psi_snapshots por la fecha actual.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    content = await file.read()
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    if "PSI RESUME" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Hoja 'PSI RESUME' no encontrada en el archivo")

    ws = wb["PSI RESUME"]

    def _float(val) -> float:
        try:
            return float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _int(val) -> int:
        try:
            return int(val) if val is not None else 0
        except (TypeError, ValueError):
            return 0

    # Filas 10 y 11 (1-indexed en openpyxl)
    row_ref  = [c.value for c in ws[10]]
    row_oven = [c.value for c in ws[11]]

    # Índices (0-based): B=1, C=2, D=3, E=4, F=5, G=6, H=7
    snap_data = {
        "ref_need_dday":     _int(row_ref[2]),
        "ref_covered_dday":  _int(row_ref[3]),
        "coverage_ref_dday": min(_float(row_ref[4]), 2.0),
        "ref_need_d1":       _int(row_ref[5]),
        "ref_covered_d1":    _int(row_ref[6]),
        "coverage_ref_d1":   min(_float(row_ref[7]), 2.0),
        "oven_need_dday":    _int(row_oven[2]),
        "oven_covered_dday": _int(row_oven[3]),
        "coverage_oven_dday": min(_float(row_oven[4]), 2.0),
        "oven_need_d1":      _int(row_oven[5]),
        "oven_covered_d1":   _int(row_oven[6]),
        "coverage_oven_d1":  min(_float(row_oven[7]), 2.0),
    }

    hoy_local = datetime.now(TZ_LOCAL).date()

    existing = (
        await db.execute(select(PsiSnapshot).where(PsiSnapshot.fecha == hoy_local))
    ).scalar_one_or_none()

    raw_payload = {"row_ref": row_ref[:15], "row_oven": row_oven[:15]}

    if existing:
        for k, v in snap_data.items():
            setattr(existing, k, v)
        existing.importado_por    = current_user.nombre_usuario
        existing.fecha_importacion = datetime.now(TZ_LOCAL)
        existing.raw_data          = raw_payload
    else:
        snap = PsiSnapshot(
            fecha=hoy_local,
            importado_por=current_user.nombre_usuario,
            fecha_importacion=datetime.now(TZ_LOCAL),
            raw_data=raw_payload,
            **snap_data,
        )
        db.add(snap)

    await db.commit()

    return {
        "message": "PSI RESUME importado correctamente",
        "fecha":   str(hoy_local),
        **snap_data,
    }


# ========================
# ÓRDENES DE COMPRA — CRUD
# ========================
@router.get("/compras")
@router.get("/compras/")
async def listar_ordenes_compra(
    status: Optional[str] = None,
    limite: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    query = select(OrdenCompra).order_by(OrdenCompra.fecha_creacion.desc()).limit(limite)
    if status:
        query = query.where(OrdenCompra.status == status)

    result = await db.execute(query)
    ordenes = result.scalars().unique().all()

    response = []
    for orden in ordenes:
        items_result = await db.execute(
            select(OrdenCompraItem).where(OrdenCompraItem.orden_compra_id == orden.id)
        )
        items = items_result.scalars().all()

        response.append({
            "id": orden.id,
            "oc_id": orden.oc_id,
            "id_proveedor": orden.id_proveedor,
            "nombre_proveedor": orden.nombre_proveedor,
            "status": orden.status,
            "origen": orden.origen or "FINANZAS",
            "fecha_creacion": orden.fecha_creacion.isoformat() if orden.fecha_creacion else None,
            "fecha_actualizacion": orden.fecha_actualizacion.isoformat() if orden.fecha_actualizacion else None,
            "notas": orden.notas,
            "creado_por": orden.creado_por,
            "aprobado_por": orden.aprobado_por,
            "items": [
                {
                    "id": item.id,
                    "sku_producto": item.sku_producto,
                    "nombre_producto": item.nombre_producto,
                    "cantidad_requerida": item.cantidad_requerida,
                    "cantidad_recibida": item.cantidad_recibida,
                    "precio_unitario": item.precio_unitario,
                    "moneda": item.moneda,
                }
                for item in items
            ],
        })

    return response


@router.post("/compras")
@router.post("/compras/")
async def crear_orden_compra(
    data: OrdenCompraCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    ahora = ahora_local()
    oc_id = f"OC-{ahora.strftime('%Y%m%d%H%M%S')}"

    orden = OrdenCompra(
        oc_id=oc_id,
        id_proveedor=data.id_proveedor,
        nombre_proveedor=data.nombre_proveedor,
        status="Pendiente de Firma",
        origen="FINANZAS",
        notas=data.notas,
        creado_por=current_user.username,
        iva=data.iva,
    )
    db.add(orden)
    await db.flush()

    for item_data in data.items:
        item = OrdenCompraItem(
            orden_compra_id=orden.id,
            sku_producto=item_data.sku_producto,
            nombre_producto=item_data.nombre_producto,
            cantidad_requerida=item_data.cantidad_requerida,
            cantidad_recibida=0,
            precio_unitario=item_data.precio_unitario,
            moneda=item_data.moneda,
        )
        db.add(item)

    await db.commit()
    return {"message": f"Orden de compra {oc_id} creada", "oc_id": oc_id, "id": orden.id}

# ==========================================
# FIRMA DE COMPRAS
# ==========================================
@router.post("/compras/{oc_id}/firmar-compras")
async def firmar_orden_compras(
    oc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_compras)  # <-- Validación automática de rol Compras
):
    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
        
    if orden.status not in ["Pendiente de Firma", "Rechazada"]:
        raise HTTPException(status_code=400, detail=f"No se puede firmar una orden en estado: {orden.status}")

    # Estampar firma digital
    orden.firma_compras = current_user.username
    orden.fecha_firma_compras = func.now()
    orden.status = "Pendiente de Firma" # Aseguramos estado si venía de Rechazada
    orden.motivo_rechazo = None # Limpiamos motivos anteriores
    
    await db.commit()
    return {"message": "Firma de Compras registrada exitosamente"}


# ==========================================
# VALIDACIÓN DE FINANZAS
# ==========================================
@router.post("/compras/{oc_id}/validar-finanzas")
async def validar_orden_finanzas(
    oc_id: str,
    payload: ValidarFinanzasRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_finanzas)  # <-- Validación automática de rol Finanzas
):
    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
        
    if not orden.firma_compras:
        raise HTTPException(status_code=400, detail="La orden debe ser firmada por Compras primero")

    if payload.accion == "aprobar":
        orden.firma_finanzas = current_user.username
        orden.fecha_firma_finanzas = func.now()
        orden.status = "Autorizada"
        mensaje = "Orden Autorizada exitosamente"
        
    elif payload.accion == "rechazar":
        if not payload.motivo:
            raise HTTPException(status_code=400, detail="Debe proporcionar un motivo para el rechazo")
        
        orden.status = "Rechazada"
        orden.motivo_rechazo = payload.motivo
        # BORRAMOS la firma de compras para obligar a re-revisión
        orden.firma_compras = None
        orden.fecha_firma_compras = None
        mensaje = "Orden Rechazada. Se ha devuelto a Compras."
    else:
        raise HTTPException(status_code=400, detail="Acción no válida")

    await db.commit()
    return {"message": mensaje, "nuevo_status": orden.status}


@router.post("/compras/{oc_id}/aprobar")
@router.post("/compras/{oc_id}/aprobar/")
async def aprobar_orden_compra(
    oc_id: str,
    data: AprobarOrdenCompraRequest = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user) 

    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")

    if orden.status != "Pendiente Aprobación":
        raise HTTPException(
            status_code=400,
            detail=f"Solo se pueden aprobar órdenes en 'Pendiente Aprobación'. Status actual: {orden.status}"
        )

    if data:
        if data.id_proveedor is not None:
            orden.id_proveedor = data.id_proveedor
        if data.nombre_proveedor is not None:
            orden.nombre_proveedor = data.nombre_proveedor
        if data.notas is not None:
            orden.notas = data.notas

        if data.items is not None:
            await db.execute(
                delete(OrdenCompraItem).where(OrdenCompraItem.orden_compra_id == orden.id)
            )
            for item_data in data.items:
                item = OrdenCompraItem(
                    orden_compra_id=orden.id,
                    sku_producto=item_data.sku_producto,
                    nombre_producto=item_data.nombre_producto,
                    cantidad_requerida=item_data.cantidad_requerida,
                    cantidad_recibida=0,
                    precio_unitario=item_data.precio_unitario,
                    moneda=item_data.moneda,
                )
                db.add(item)

    orden.status = "Creada"
    orden.aprobado_por = current_user.username
    await db.commit()

    return {"message": f"Orden {oc_id} aprobada exitosamente. Status: Creada"}


@router.get("/compras/{oc_id}")
@router.get("/compras/{oc_id}/")
async def obtener_orden_compra(
    oc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")

    items_result = await db.execute(
        select(OrdenCompraItem).where(OrdenCompraItem.orden_compra_id == orden.id)
    )
    items = items_result.scalars().all()

    recepciones_result = await db.execute(
        select(RecepcionCompra).where(RecepcionCompra.orden_compra_id == orden.id)
        .order_by(RecepcionCompra.fecha_recepcion.desc())
    )
    recepciones = recepciones_result.scalars().all()

    return {
        "id": orden.id,
        "oc_id": orden.oc_id,
        "id_proveedor": orden.id_proveedor,
        "nombre_proveedor": orden.nombre_proveedor,
        "status": orden.status,
        "origen": orden.origen or "FINANZAS",
        "fecha_creacion": orden.fecha_creacion.isoformat() if orden.fecha_creacion else None,
        "notas": orden.notas,
        "creado_por": orden.creado_por,
        "aprobado_por": orden.aprobado_por,
        "items": [
            {
                "id": i.id,
                "sku_producto": i.sku_producto,
                "nombre_producto": i.nombre_producto,
                "cantidad_requerida": i.cantidad_requerida,
                "cantidad_recibida": i.cantidad_recibida,
                "precio_unitario": i.precio_unitario,
                "moneda": i.moneda,
            }
            for i in items
        ],
        "recepciones": [
            {
                "id": r.id,
                "recepcion_id": r.recepcion_id,
                "sku_producto": r.sku_producto,
                "cantidad_recibida": r.cantidad_recibida,
                "fecha_recepcion": r.fecha_recepcion.isoformat() if r.fecha_recepcion else None,
                "recibido_por": r.recibido_por,
                "notas": r.notas,
            }
            for r in recepciones
        ],
    }


@router.get("/compras/{oc_id}/pdf")
@router.get("/compras/{oc_id}/pdf/")
async def generar_pdf_orden_compra(
    oc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    # ==========================================
    # 1. EXTRACCIÓN Y PREPARACIÓN DE DATOS
    # ==========================================
    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")

    items_result = await db.execute(
        select(OrdenCompraItem).where(OrdenCompraItem.orden_compra_id == orden.id)
    )
    items = items_result.scalars().all()

    # Obtener datos del proveedor (receptor)
    # Validamos si el id_proveedor guardado es un entero o un UUID string (ej. 'PROV-C7BE10')
    if isinstance(orden.id_proveedor, int) or str(orden.id_proveedor).isdigit():
        prov_result = await db.execute(select(Proveedor).where(Proveedor.id == int(orden.id_proveedor)))
    else:
        prov_result = await db.execute(select(Proveedor).where(Proveedor.uuid == str(orden.id_proveedor)))
        
    proveedor = prov_result.scalar_one_or_none()

    # Obtener datos de la empresa (emisor)
    empresa_result = await db.execute(select(ConfiguracionEmpresa).limit(1))
    empresa = empresa_result.scalar_one_or_none()

    # Cálculos financieros
    subtotal = sum(i.cantidad_requerida * i.precio_unitario for i in items)
    
    # Manejo robusto del IVA
    raw_iva = orden.iva if orden.iva is not None else 0.0
    try:
        val_iva = float(raw_iva)
    except (ValueError, TypeError):
        val_iva = 0.0
        
    if val_iva >= 1.0:
        porcentaje_iva = val_iva
        monto_iva = subtotal * (porcentaje_iva / 100.0)
    else:
        porcentaje_iva = val_iva * 100.0
        monto_iva = subtotal * val_iva

    total = subtotal + monto_iva
    moneda_oc = items[0].moneda if items else "MXN"

    # ==========================================
    # 2. CONFIGURACIÓN DEL DOCUMENTO (PLATYPUS)
    # ==========================================
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=30, leftMargin=30,
        topMargin=30, bottomMargin=30,
        title=f"Orden de Compra {oc_id}",
        author="Cheong Woon Mexico"
    )
    elements = []
    styles = getSampleStyleSheet()

    # Estilos de texto personalizados
    title_style = ParagraphStyle(name="TitleStyle", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#333333"))
    normal_style = styles["Normal"]
    right_style = ParagraphStyle(name="RightStyle", parent=styles["Normal"], alignment=TA_RIGHT)

    # ==========================================
    # 3. ENCABEZADO (Logo y Folio)
    # ==========================================
    logo_path = os.path.join("static", "Logo.png")
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=1.5 * inch, height=0.75 * inch)
    else:
        logo = Paragraph("<b>Cheong Woon</b>", title_style)

    fecha_str = orden.fecha_creacion.strftime("%Y-%m-%d") if orden.fecha_creacion else "N/A"

    header_data = [
        [logo, Paragraph(f"<b>ORDEN DE COMPRA</b><br/>Folio: <b>{oc_id}</b><br/>Fecha: {fecha_str}", right_style)]
    ]
    
    header_table = Table(header_data, colWidths=[3 * inch, 4.5 * inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # ==========================================
    # 4. BLOQUE DE INFORMACIÓN (Emisor / Receptor)
    # ==========================================
    # Extraer datos exactos basados en el modelo ConfiguracionEmpresa
    if empresa:
        razon_social_empresa = empresa.nombre
        rfc_empresa = empresa.rfc or "N/A"
        
        # Armar la dirección completa omitiendo campos vacíos
        direccion_empresa = empresa.direccion if empresa.direccion else "N/A"
    else:
        razon_social_empresa = "Cheong Woon Mexico"
        rfc_empresa = "N/A"
        direccion_empresa = "N/A"

    # Datos del Proveedor
    razon_social_prov = proveedor.razon_social if proveedor else orden.nombre_proveedor
    rfc_prov = proveedor.rfc if proveedor else "N/A"

    # Usamos getattr por seguridad si la columna llega a estar vacía
    direccion_prov = getattr(proveedor, 'direccion', None) or "N/A"
    cond_pago_prov = proveedor.condiciones_pago if proveedor else "N/A"
    dias_credito = getattr(proveedor, 'dias_credito', None) or 0

    emisor_text = f"""
    <b>Facturar y Entregar a:</b><br/>
    Razón Social: {razon_social_empresa}<br/>
    RFC: {rfc_empresa}<br/>
    Dirección: {direccion_empresa}
    """

    receptor_text = f"""
    <b>Datos del Proveedor:</b><br/>
    Razón Social: {razon_social_prov}<br/>
    RFC: {rfc_prov}<br/>
    Dirección: {direccion_prov}<br/>
    Condiciones de Pago: {cond_pago_prov} ({dias_credito} días)
    """

    info_data = [[Paragraph(emisor_text, normal_style), Paragraph(receptor_text, normal_style)]]
    info_table = Table(info_data, colWidths=[3.75 * inch, 3.75 * inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f9f9f9")),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))

    # ==========================================
    # 5. TABLA DE PARTIDAS (ITEMS)
    # ==========================================
    items_data = [["Partida", "SKU", "Descripción", "Cantidad", "P. Unitario", "Importe"]]
    for idx, item in enumerate(items, 1):
        importe = item.cantidad_requerida * item.precio_unitario
        items_data.append([
            str(idx),
            item.sku_producto,
            Paragraph(item.nombre_producto, normal_style),
            str(item.cantidad_requerida),
            f"${item.precio_unitario:,.2f}",
            f"${importe:,.2f}"
        ])

    items_table = Table(items_data, colWidths=[0.6 * inch, 1.2 * inch, 2.7 * inch, 0.8 * inch, 1 * inch, 1.2 * inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Descripción alineada a la izquierda
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'), # Precios alineados a la derecha
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 15))

    # ==========================================
    # 6. RESUMEN FINANCIERO
    # ==========================================
    summary_data = [
        ["Subtotal:", f"${subtotal:,.2f} {moneda_oc}"],
        [f"IVA ({porcentaje_iva:g}%):", f"${monto_iva:,.2f} {moneda_oc}"],
        ["Total:", f"${total:,.2f} {moneda_oc}"]
    ]
    summary_table = Table(summary_data, colWidths=[1.5 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), # Total en negritas
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
    ]))

    # Envolvemos la tabla resumen en una tabla invisible para alinearla a la derecha
    container_data = [["", summary_table]]
    container_table = Table(container_data, colWidths=[4.5 * inch, 3 * inch])
    container_table.setStyle(TableStyle([('ALIGN', (1, 0), (1, 0), 'RIGHT')]))
    elements.append(container_table)
    elements.append(Spacer(1, 50)) # Espacio antes de las firmas

    # ==========================================
    # 7. PIE DE PÁGINA (QR y Firmas)
    # ==========================================
    # Generar código QR con el folio de la OC
    qr_img = qrcode.make(oc_id)
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)
    qr_rl = RLImage(qr_buffer, width=1.2 * inch, height=1.2 * inch)

    # Estilo centrado para las firmas
    center_style = ParagraphStyle(name="CenterStyle", parent=styles["Normal"], alignment=TA_CENTER)

    # Función para dibujar la firma o dejar la línea en blanco
    def formatear_firma(nombre, fecha):
        linea_blanco = "_________________________"
        if nombre and fecha:
            fecha_str = fecha.strftime("%Y-%m-%d %H:%M")
            # Agregamos un <br/> y la línea debajo del texto
            return Paragraph(
                f"<font size=8 color='#10b981'><b>Firmado digitalmente por:</b></font><br/>"
                f"<b>{nombre}</b><br/>"
                f"<font size=8>{fecha_str}</font><br/>"
                f"{linea_blanco}", 
                center_style
            )
        return linea_blanco

    # Extraemos las firmas del objeto orden
    firma_compras_rl = formatear_firma(getattr(orden, 'firma_compras', None), getattr(orden, 'fecha_firma_compras', None))
    firma_finanzas_rl = formatear_firma(getattr(orden, 'firma_finanzas', None), getattr(orden, 'fecha_firma_finanzas', None))

    firmas_data = [
        [qr_rl, firma_compras_rl, firma_finanzas_rl],
        ["", "Compras", "Finanzas"]
    ]
    firmas_table = Table(firmas_data, colWidths=[2.5 * inch, 2.5 * inch, 2.5 * inch])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),      # QR centrado/izq
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),   # Firmas centradas
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),  # Alineado a la base
    ]))
    elements.append(firmas_table)

    # ==========================================
    # 8. CONSTRUCCIÓN Y RETORNO DEL PDF
    # ==========================================
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={oc_id}.pdf"},
    )


@router.put("/compras/{oc_id}")
@router.put("/compras/{oc_id}/")
async def actualizar_orden_compra(
    oc_id: str,
    data: OrdenCompraUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")

    if data.nombre_proveedor is not None:
        orden.nombre_proveedor = data.nombre_proveedor
    if data.status is not None:
        orden.status = data.status
    if data.notas is not None:
        orden.notas = data.notas

    if data.items is not None:
        await db.execute(
            delete(OrdenCompraItem).where(OrdenCompraItem.orden_compra_id == orden.id)
        )
        for item_data in data.items:
            item = OrdenCompraItem(
                orden_compra_id=orden.id,
                sku_producto=item_data.sku_producto,
                nombre_producto=item_data.nombre_producto,
                cantidad_requerida=item_data.cantidad_requerida,
                cantidad_recibida=0,
                precio_unitario=item_data.precio_unitario,
                moneda=item_data.moneda,
            )
            db.add(item)

    await db.commit()
    return {"message": f"Orden {oc_id} actualizada"}


@router.delete("/compras/{oc_id}")
@router.delete("/compras/{oc_id}/")
async def eliminar_orden_compra(
    oc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")

    rec_count = (await db.execute(
        select(func.count(RecepcionCompra.id)).where(RecepcionCompra.orden_compra_id == orden.id)
    )).scalar() or 0

    if rec_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"No se puede eliminar. La OC {oc_id} ya tiene {rec_count} recepciones registradas."
        )

    await db.delete(orden)
    await db.commit()
    return {"message": f"Orden {oc_id} eliminada"}


@router.get("/recepciones")
@router.get("/recepciones/")
async def listar_recepciones(
    limite: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(
        select(RecepcionCompra).order_by(RecepcionCompra.fecha_recepcion.desc()).limit(limite)
    )
    recepciones = result.scalars().all()

    return [
        {
            "id": r.id,
            "recepcion_id": r.recepcion_id,
            "oc_id": r.oc_id,
            "sku_producto": r.sku_producto,
            "cantidad_recibida": r.cantidad_recibida,
            "fecha_recepcion": r.fecha_recepcion.isoformat() if r.fecha_recepcion else None,
            "recibido_por": r.recibido_por,
            "notas": r.notas,
        }
        for r in recepciones
    ]

@router.get("/lote/{lote_id}")
@router.get("/lote/{lote_id}/")
async def obtener_info_lote(
    lote_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    partes = lote_id.split("-")
    if len(partes) != 3:
        raise HTTPException(status_code=400, detail="Formato de Lote ID inválido. Esperado: YYYYMMDD-XXXX-N")

    fecha_str, sku_suffix, rec_count_str = partes

    result = await db.execute(
        select(RecepcionCompra).where(
            RecepcionCompra.sku_producto.ilike(f"%{sku_suffix}")
        ).order_by(RecepcionCompra.fecha_recepcion.desc())
    )
    recepciones = result.scalars().all()

    if not recepciones:
        raise HTTPException(status_code=404, detail=f"No se encontraron recepciones para el lote {lote_id}")

    recepciones_filtradas = [
        r for r in recepciones
        if r.fecha_recepcion and r.fecha_recepcion.strftime("%Y%m%d") == fecha_str
    ]

    recs_finales = recepciones_filtradas if recepciones_filtradas else recepciones

    if not recs_finales:
        raise HTTPException(status_code=404, detail=f"No se encontraron recepciones para el lote {lote_id}")

    rec_principal = recs_finales[0]
    sku_completo = rec_principal.sku_producto
    oc_id = rec_principal.oc_id

    oc_result = await db.execute(select(OrdenCompra).where(OrdenCompra.oc_id == oc_id))
    orden = oc_result.scalar_one_or_none()

    item = None
    if orden:
        item_result = await db.execute(
            select(OrdenCompraItem).where(
                and_(
                    OrdenCompraItem.orden_compra_id == orden.id,
                    OrdenCompraItem.sku_producto == sku_completo,
                )
            )
        )
        item = item_result.scalar_one_or_none()

    cantidad_total = sum(r.cantidad_recibida for r in recs_finales if r.sku_producto == sku_completo)

    return {
        "lote_id": lote_id,
        "sku_producto": sku_completo,
        "nombre_producto": item.nombre_producto if item else None,
        "oc_id": oc_id,
        "nombre_proveedor": orden.nombre_proveedor if orden else None,
        "cantidad_total_recibida": cantidad_total,
        "cantidad_requerida": item.cantidad_requerida if item else None,
        "precio_unitario": item.precio_unitario if item else None,
        "moneda": item.moneda if item else "MXN",
        "status_oc": orden.status if orden else None,
        "total_recepciones": len(recs_finales),
        "recepciones": [
            {
                "recepcion_id": r.recepcion_id,
                "cantidad_recibida": r.cantidad_recibida,
                "fecha_recepcion": r.fecha_recepcion.isoformat() if r.fecha_recepcion else None,
                "recibido_por": r.recibido_por,
                "notas": r.notas,
            }
            for r in recs_finales
            if r.sku_producto == sku_completo
        ],
    }


# ========================
# ÓRDENES DE VENTA — CRUD
# ========================
@router.get("/ventas")
@router.get("/ventas/")
async def listar_ordenes_venta(
    estado: Optional[str] = None,
    limite: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    query = select(OrdenVenta).order_by(OrdenVenta.fecha_creacion.desc()).limit(limite)
    if estado and estado != "Todos":
        query = query.where(OrdenVenta.estado == estado)

    result = await db.execute(query)
    ordenes = result.scalars().unique().all()

    response = []
    for orden in ordenes:
        items_result = await db.execute(
            select(OrdenVentaItem).where(OrdenVentaItem.orden_venta_id == orden.id)
        )
        items = items_result.scalars().all()

        total_items = len(items)
        valor_total = sum((i.cantidad * i.precio_unitario) for i in items)

        response.append({
            "id": orden.id,
            "ov_id": orden.ov_id,
            "cliente_id": orden.cliente_id,
            "nombre_cliente": orden.nombre_cliente,
            "estado": orden.estado,
            "fecha_creacion": orden.fecha_creacion.isoformat() if orden.fecha_creacion else None,
            "fecha_actualizacion": orden.fecha_actualizacion.isoformat() if orden.fecha_actualizacion else None,
            "notas": orden.notas,
            "creado_por": orden.creado_por,
            "total_items": total_items,
            "valor_total": round(valor_total, 2),
            "items": [
                {
                    "id": i.id,
                    "sku_producto": i.sku_producto,
                    "nombre_producto": i.nombre_producto,
                    "cantidad": i.cantidad,
                    "cantidad_enviada": i.cantidad_enviada,
                    "precio_unitario": i.precio_unitario,
                    "moneda": i.moneda,
                }
                for i in items
            ],
        })

    return response


@router.post("/ventas")
@router.post("/ventas/")
async def crear_orden_venta(
    data: OrdenVentaCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    ahora = ahora_local()
    ov_id = f"OV-{ahora.strftime('%Y%m%d%H%M%S')}"

    orden = OrdenVenta(
        ov_id=ov_id,
        cliente_id=data.cliente_id,
        nombre_cliente=data.nombre_cliente,
        estado="Pendiente de Envío",
        notas=data.notas,
        creado_por=current_user.username,
    )
    db.add(orden)
    await db.flush()

    for item_data in data.items:
        item = OrdenVentaItem(
            orden_venta_id=orden.id,
            sku_producto=item_data.sku_producto,
            nombre_producto=item_data.nombre_producto,
            cantidad=item_data.cantidad,
            cantidad_enviada=0,
            precio_unitario=item_data.precio_unitario,
            moneda=item_data.moneda,
        )
        db.add(item)

    await db.commit()
    return {"message": f"Orden de venta {ov_id} creada", "ov_id": ov_id, "id": orden.id}


@router.get("/ventas/{ov_id}")
@router.get("/ventas/{ov_id}/")
async def obtener_orden_venta(
    ov_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(select(OrdenVenta).where(OrdenVenta.ov_id == ov_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de venta no encontrada")

    items_result = await db.execute(
        select(OrdenVentaItem).where(OrdenVentaItem.orden_venta_id == orden.id)
    )
    items = items_result.scalars().all()

    envios_result = await db.execute(
        select(EnvioVenta).where(EnvioVenta.orden_venta_id == orden.id)
        .order_by(EnvioVenta.fecha_envio.desc())
    )
    envios = envios_result.scalars().all()

    return {
        "id": orden.id,
        "ov_id": orden.ov_id,
        "cliente_id": orden.cliente_id,
        "nombre_cliente": orden.nombre_cliente,
        "estado": orden.estado,
        "fecha_creacion": orden.fecha_creacion.isoformat() if orden.fecha_creacion else None,
        "notas": orden.notas,
        "creado_por": orden.creado_por,
        "items": [
            {
                "id": i.id,
                "sku_producto": i.sku_producto,
                "nombre_producto": i.nombre_producto,
                "cantidad": i.cantidad,
                "cantidad_enviada": i.cantidad_enviada,
                "precio_unitario": i.precio_unitario,
                "moneda": i.moneda,
            }
            for i in items
        ],
        "envios": [
            {
                "id": e.id,
                "envio_id": e.envio_id,
                "fecha_envio": e.fecha_envio.isoformat() if e.fecha_envio else None,
                "autorizado_por": e.autorizado_por,
                "items_enviados": e.items_enviados,
                "notas": e.notas,
            }
            for e in envios
        ],
    }


@router.put("/ventas/{ov_id}")
@router.put("/ventas/{ov_id}/")
async def actualizar_orden_venta(
    ov_id: str,
    data: OrdenVentaUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(select(OrdenVenta).where(OrdenVenta.ov_id == ov_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden de venta no encontrada")

    if data.nombre_cliente is not None:
        orden.nombre_cliente = data.nombre_cliente
    if data.estado is not None:
        orden.estado = data.estado
    if data.notas is not None:
        orden.notas = data.notas

    await db.commit()
    return {"message": f"Orden {ov_id} actualizada"}


@router.post("/ventas/{ov_id}/enviar")
@router.post("/ventas/{ov_id}/enviar/")
async def enviar_orden_venta(
    ov_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(select(OrdenVenta).where(OrdenVenta.ov_id == ov_id))
    orden = result.scalar_one_or_none()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    if orden.estado not in ("Pendiente de Envío",):
        raise HTTPException(status_code=400, detail=f"No se puede enviar. Estado actual: {orden.estado}")

    items_result = await db.execute(
        select(OrdenVentaItem).where(OrdenVentaItem.orden_venta_id == orden.id)
    )
    items = items_result.scalars().all()

    ahora = ahora_local()
    envio_id = f"ENV-{ahora.strftime('%Y%m%d%H%M%S')}"

    items_enviados = [{"sku_producto": i.sku_producto, "cantidad": i.cantidad} for i in items]

    envio = EnvioVenta(
        envio_id=envio_id,
        orden_venta_id=orden.id,
        ov_id=ov_id,
        autorizado_por=current_user.username,
        items_enviados=items_enviados,
    )
    db.add(envio)

    for item in items:
        item.cantidad_enviada = item.cantidad

    orden.estado = "Enviado"

    await db.commit()
    return {"message": f"Orden {ov_id} enviada", "envio_id": envio_id}


# ─── Endpoint: cambio de estado ───────────────────────────────────────────────
 
@router.patch("/ventas/{ov_id}/estado")
async def cambiar_estado_ov(
    ov_id: str = Path(...),
    body: CambiarEstadoRequest = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Mueve una OV al estado especificado.
    Valida la máquina de estados y los permisos por rol.
 
    Uso típico:
      PATCH /finanzas/ventas/OV-20260608-001/estado
      { "estado": "En Preparación" }           ← Almacén inicia armado
      { "estado": "Lista para Carga" }          ← Almacén valida cajas en andén
      { "estado": "En Preparación" }            ← Rollback: problema en andén
      { "estado": "Cancelada", "notas": "..." } ← Ventas cancela
    """
    ov = await _get_ov(db, ov_id)
    _assert_transicion(ov, body.estado, current_user.rol)
 
    estado_anterior = ov.estado
    ov.estado = body.estado
    ov.fecha_actualizacion = datetime.now(TZ_LOCAL)
 
    if body.notas:
        ov.notas = (ov.notas or "") + f"\n[{body.estado}] {body.notas}"
 
    await db.commit()
    await db.refresh(ov)
 
    return {
        "message":          f"OV {ov_id} movida de '{estado_anterior}' a '{body.estado}'",
        "ov_id":            ov.ov_id,
        "estado_anterior":  estado_anterior,
        "estado_nuevo":     ov.estado,
    }
 
 
# ─── Endpoint: despacho físico ────────────────────────────────────────────────
 
@router.post("/ventas/{ov_id}/despachar")
async def despachar_ov(
    ov_id: str = Path(...),
    body: DespacharRequest = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Registra la salida física del camión.
    Solo permitido desde estado 'Lista para Carga'.
 
    Lógica de status_salida:
      "OK"  → OV pasa a 'Enviado' (o 'Embarque Parcial' si cantidades < pedido)
      "NG"  → OV VUELVE a 'Lista para Carga' — el camión no salió,
               el envío queda registrado como fallido para trazabilidad
    """
    ov = await _get_ov(db, ov_id)
 
    if ov.estado != "Lista para Carga":
        raise HTTPException(
            422,
            f"Solo se puede despachar una OV en estado 'Lista para Carga'. "
            f"Estado actual: '{ov.estado}'",
        )
 
    if current_user.rol not in ["admin", "logistica", "ventas"]:
        raise HTTPException(403, "Solo logística o ventas pueden registrar el despacho.")
 
    # ── Determinar items enviados ──────────────────────────────────────────
    items_enviados = body.items_enviados or []
    if not items_enviados:
        # Si no se especifican items, asumimos envío completo de todos los items
        items_enviados = [
            {"sku_producto": item.sku_producto, "cantidad": item.cantidad - item.cantidad_enviada}
            for item in ov.items
            if item.cantidad > item.cantidad_enviada
        ]
 
    # ── Calcular si es envío parcial o completo ────────────────────────────
    es_completo = True
    for item_ov in ov.items:
        enviado_ahora = next(
            (ie["cantidad"] for ie in items_enviados if ie["sku_producto"] == item_ov.sku_producto),
            0,
        )
        nuevo_total_enviado = item_ov.cantidad_enviada + enviado_ahora
        if nuevo_total_enviado < item_ov.cantidad:
            es_completo = False
 
    # ── Crear registro de envío (siempre, incluso si NG) ──────────────────
    import uuid as _uuid
    envio_id = f"ENV-{datetime.now(TZ_LOCAL).strftime('%Y%m%d%H%M%S')}-{_uuid.uuid4().hex[:4].upper()}"
 
    nuevo_envio = EnvioVenta(
        envio_id=envio_id,
        orden_venta_id=ov.id,
        ov_id=ov.ov_id,
        autorizado_por=current_user.username,
        items_enviados=items_enviados,
        no_camion=body.no_camion,
        chofer=body.chofer,
        status_salida=body.status_salida,
        no_departure=body.no_departure,   # campo nuevo
        notas=f"status_salida={body.status_salida}",
    )
    db.add(nuevo_envio)
 
    # ── Actualizar cantidades y estado de la OV ────────────────────────────
    if body.status_salida == "OK":
        # Actualizar cantidad_enviada de cada item
        for item_ov in ov.items:
            enviado_ahora = next(
                (ie["cantidad"] for ie in items_enviados if ie["sku_producto"] == item_ov.sku_producto),
                0,
            )
            item_ov.cantidad_enviada += enviado_ahora
 
        # Actualizar cw_invoice en la OV
        if body.cw_invoice:
            ov.cw_invoice = body.cw_invoice
 
        # Nuevo estado según completitud
        ov.estado = "Enviado" if es_completo else "Embarque Parcial"
 
    else:
        # NG: el camión no salió — regresa a Lista para Carga para reintentar
        ov.estado = "Lista para Carga"
 
    ov.fecha_actualizacion = datetime.now(TZ_LOCAL)
    await db.commit()
 
    return {
        "message":       f"Despacho registrado. Status: {body.status_salida}",
        "envio_id":      envio_id,
        "estado_ov":     ov.estado,
        "no_departure":  body.no_departure,
    }


# ========================
# DEVOLUCIONES
# ========================
@router.get("/devoluciones")
@router.get("/devoluciones/")
async def listar_devoluciones(
    estado: Optional[str] = None,
    limite: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    query = select(Devolucion).order_by(Devolucion.fecha_devolucion.desc()).limit(limite)
    if estado:
        query = query.where(Devolucion.estado_inspeccion == estado)

    result = await db.execute(query)
    devoluciones = result.scalars().all()

    return [
        {
            "id": d.id,
            "devolucion_id": d.devolucion_id,
            "ov_id": d.ov_id,
            "sku_producto": d.sku_producto,
            "nombre_producto": d.nombre_producto,
            "cantidad_devuelta": d.cantidad_devuelta,
            "motivo": d.motivo,
            "lote_produccion_origen": d.lote_produccion_origen,
            "fecha_devolucion": d.fecha_devolucion.isoformat() if d.fecha_devolucion else None,
            "estado_inspeccion": d.estado_inspeccion,
            "disposicion_final": d.disposicion_final,
            "cantidad_scrap": d.cantidad_scrap,
            "cantidad_retrabajo": d.cantidad_retrabajo,
            "procesado_por": d.procesado_por,
            "creado_por": d.creado_por,
        }
        for d in devoluciones
    ]


@router.post("/devoluciones")
@router.post("/devoluciones/")
async def registrar_devolucion(
    data: DevolucionCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    ahora = ahora_local()
    fecha_str = ahora.strftime("%d%m%y%H%M")
    sku_suffix = data.sku_producto[-4:].upper()
    devolucion_id = f"DV{fecha_str}{sku_suffix}"

    devolucion = Devolucion(
        devolucion_id=devolucion_id,
        ov_id=data.ov_id,
        sku_producto=data.sku_producto,
        nombre_producto=data.nombre_producto,
        cantidad_devuelta=data.cantidad_devuelta,
        motivo=data.motivo,
        lote_produccion_origen=data.lote_produccion_origen,
        estado_inspeccion="Pendiente",
        creado_por=current_user.username,
    )
    db.add(devolucion)
    await db.commit()

    return {"message": f"Devolución {devolucion_id} registrada", "devolucion_id": devolucion_id}


@router.put("/devoluciones/{devolucion_id}/disposicion")
@router.put("/devoluciones/{devolucion_id}/disposicion/")
async def procesar_disposicion(
    devolucion_id: str,
    data: DisposicionDevolucionCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(
        select(Devolucion).where(Devolucion.devolucion_id == devolucion_id)
    )
    devolucion = result.scalar_one_or_none()
    if not devolucion:
        raise HTTPException(status_code=404, detail="Devolución no encontrada")

    if devolucion.estado_inspeccion == "Finalizado":
        raise HTTPException(status_code=400, detail="Esta devolución ya fue procesada")

    total_disposicion = data.cantidad_scrap + data.cantidad_retrabajo
    if total_disposicion > devolucion.cantidad_devuelta:
        raise HTTPException(
            status_code=400,
            detail=f"La suma scrap+retrabajo ({total_disposicion}) supera la cantidad devuelta ({devolucion.cantidad_devuelta})"
        )

    devolucion.cantidad_scrap = data.cantidad_scrap
    devolucion.cantidad_retrabajo = data.cantidad_retrabajo
    devolucion.disposicion_final = f"Scrap: {data.cantidad_scrap}, Retrabajo: {data.cantidad_retrabajo}"
    devolucion.estado_inspeccion = "Finalizado"
    devolucion.procesado_por = current_user.username

    await db.commit()
    return {"message": f"Devolución {devolucion_id} procesada", "disposicion": devolucion.disposicion_final}


# ========================
# PLAN DE VENTAS
# ========================
@router.get("/plan-ventas")
@router.get("/plan-ventas/")
async def listar_planes_ventas(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(
        select(PlanVentas).order_by(PlanVentas.fecha_inicio_semana.desc()).limit(20)
    )
    planes = result.scalars().all()

    return [
        {
            "id": p.id,
            "identificador_semana": p.identificador_semana,
            "fecha_inicio_semana": p.fecha_inicio_semana.isoformat(),
            "fecha_importacion": p.fecha_importacion.isoformat() if p.fecha_importacion else None,
            "items": p.items or [],
            "importado_por": p.importado_por,
            "total_skus": len(p.items) if p.items else 0,
        }
        for p in planes
    ]


@router.get("/plan-ventas/{identificador_semana}")
@router.get("/plan-ventas/{identificador_semana}/")
async def obtener_plan_ventas(
    identificador_semana: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(
        select(PlanVentas).where(PlanVentas.identificador_semana == identificador_semana)
    )
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan de ventas no encontrado")

    return {
        "id": plan.id,
        "identificador_semana": plan.identificador_semana,
        "fecha_inicio_semana": plan.fecha_inicio_semana.isoformat(),
        "fecha_importacion": plan.fecha_importacion.isoformat() if plan.fecha_importacion else None,
        "items": plan.items or [],
        "importado_por": plan.importado_por,
    }


# ------------------------------------------------------------------ #
# Mapeo timestamp → nombre de día en español (cómo vienen del Excel) #
# ------------------------------------------------------------------ #
DIA_MAP = {
    0: "LUNES",
    1: "MARTES",
    2: "MIERCOLES",
    3: "JUEVES",
    4: "VIERNES",
    5: "SABADO",
    6: "DOMINGO",
}
 
# Columnas mínimas que deben existir en la fila de headers (fila 4)
COLS_REQUERIDAS = {"NO. PARTE", "DESCRIPCION", "LINE", "INV. CW", "INV. LG"}
 
 
# ------------------------------------------------------------------ #
# Función principal de parseo                                          #
# ------------------------------------------------------------------ #
def parsear_cw_plan(contenido: bytes) -> list[dict[str, Any]]:
    """
    Lee la hoja 'CW PLAN' de un Excel y devuelve una lista de items
    con la estructura extendida de PlanVentas.items.
 
    Estructura de la hoja CW PLAN:
      Filas 0-3  → metadata (fechas, labels de sección) — se ignoran
      Fila 4     → headers reales (NO. PARTE, DESCRIPCION, LINE, ...)
      Filas 5+   → datos. Filas con NO. PARTE vacío son separadores.
 
    Columnas de días: son datetime objects en pandas. Se detectan
    automáticamente buscando columnas de tipo datetime64.
    """
    try:
        df_raw = pd.read_excel(
            io.BytesIO(contenido),
            sheet_name="CW PLAN",
            header=None,          # leemos sin header para manejar las filas de metadata
        )
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer la hoja 'CW PLAN': {e}")
 
    # ── 1. Extraer headers de la fila 4 (índice 4) ──────────────────
    header_row = df_raw.iloc[4]
 
    # Los headers de días son objetos datetime (pandas los parsea así)
    # Los demás son strings. Construimos un dict col_idx → nombre_col
    col_names: dict[int, str] = {}
    dia_cols:  dict[int, str] = {}   # col_idx → "LUNES" / "MARTES" / etc.
 
    for idx, val in enumerate(header_row):
        if pd.isna(val):
            continue
        if isinstance(val, (datetime, pd.Timestamp)):
            # Es una columna de día — la mapeamos por weekday()
            dia_nombre = DIA_MAP.get(pd.Timestamp(val).weekday())
            if dia_nombre:
                dia_cols[idx] = dia_nombre
                col_names[idx] = dia_nombre
        else:
            col_names[idx] = str(val).strip()
 
    # Verificar que las columnas esenciales existen
    nombres = set(col_names.values())
    faltantes = COLS_REQUERIDAS - nombres
    if faltantes:
        raise HTTPException(
            400,
            f"Faltan columnas en CW PLAN: {faltantes}. "
            f"Columnas encontradas: {nombres}"
        )
 
    # Índices de columnas clave
    def _idx(nombre: str) -> int:
        for k, v in col_names.items():
            if v == nombre:
                return k
        raise HTTPException(400, f"Columna '{nombre}' no encontrada")
 
    idx_sku        = _idx("NO. PARTE")
    idx_desc       = _idx("DESCRIPCION")
    idx_linea      = _idx("LINE")
    idx_stock_cw   = _idx("INV. CW")
    idx_stock_lg   = _idx("INV. LG")
 
    # Columnas opcionales (pueden no existir en versiones anteriores del Excel)
    idx_cw_line = next((k for k, v in col_names.items() if v == "CW LINE"), None)
    idx_model   = next((k for k, v in col_names.items() if v == "MODEL"), None)
    idx_id1     = next((k for k, v in col_names.items() if v == "ID 1"), None)
 
    # ── 2. Iterar filas de datos (fila 5 en adelante) ───────────────
    items: list[dict[str, Any]] = []
    data_rows = df_raw.iloc[5:]
 
    for _, row in data_rows.iterrows():
        sku = row.iloc[idx_sku]
 
        # Saltar separadores (filas vacías o con SKU nulo)
        if pd.isna(sku) or str(sku).strip() == "":
            continue
 
        sku = str(sku).strip()
 
        # Saltar filas de subtotal (tienen SKU numérico o "TOTAL")
        if sku.upper() in ("TOTAL", "SUBTOTAL") or sku.isdigit():
            continue
 
        def _val(idx: int | None, default=None):
            if idx is None:
                return default
            v = row.iloc[idx]
            return None if pd.isna(v) else v
 
        def _int(idx: int | None, default: int = 0) -> int:
            v = _val(idx)
            try:
                return int(v) if v is not None else default
            except (ValueError, TypeError):
                return default
 
        # ── Construir dict de días ──────────────────────────────────
        dias: dict[str, dict] = {}
        for col_idx, dia_nombre in dia_cols.items():
            plan_qty = _int(col_idx, 0)
            dias[dia_nombre] = {
                "plan":        plan_qty,
                "status":      "Pendiente",
                "ov_generada": None,
            }
 
        item: dict[str, Any] = {
            "sku":          sku,
            "descripcion":  str(_val(idx_desc) or "").strip(),
            "linea":        str(_val(idx_linea) or "").strip(),
            # ── Campos nuevos ──
            "cw_line":      str(_val(idx_cw_line) or "").strip() or None,
            "model":        str(_val(idx_model)   or "").strip() or None,
            "id1":          str(_val(idx_id1)     or "").strip() or None,
            # ── Stock ──
            "stock_actual": _int(idx_stock_cw, 0),   # INV. CW  (stock en planta CW)
            "stock_lg":     _int(idx_stock_lg, 0),   # INV. LG  (stock en planta LG)
            # ── Días ──
            "dias":         dias,
        }
        items.append(item)
 
    if not items:
        raise HTTPException(400, "No se encontraron SKUs válidos en la hoja CW PLAN")
 
    return items


@router.post("/plan-ventas/importar")
@router.post("/plan-ventas/importar/")
async def importar_plan_ventas(
    fecha_inicio_semana: date = Query(..., description="Lunes de la semana. Ej: 2026-06-09"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Importa la hoja 'CW PLAN' de un Excel y guarda/actualiza el PlanVentas
    de esa semana. Es idempotente: si ya existe el plan de esa semana,
    lo reemplaza (upsert por identificador_semana).
 
    Preserva los status "Autorizado" y las ov_generada de días que
    ya fueron procesados — no los sobreescribe con "Pendiente".
    """
    contenido = await file.read()
    nuevos_items = parsear_cw_plan(contenido)
 
    # Identificador de semana: YYYY-WW (ISO week number)
    iso_year, iso_week, _ = fecha_inicio_semana.isocalendar()
    identificador = f"{iso_year}-{iso_week:02d}"
 
    # Buscar si ya existe el plan de esa semana
    result = await db.execute(
        select(PlanVentas).where(PlanVentas.identificador_semana == identificador)
    )
    plan_existente: PlanVentas | None = result.scalar_one_or_none()
 
    if plan_existente:
        # ── Upsert: preservar status/ov_generada de días ya autorizados ──
        items_previos = {item["sku"]: item for item in (plan_existente.items or [])}
 
        for nuevo_item in nuevos_items:
            sku = nuevo_item["sku"]
            if sku in items_previos:
                prev = items_previos[sku]
                for dia, datos_dia in nuevo_item["dias"].items():
                    prev_dia = prev.get("dias", {}).get(dia, {})
                    # Solo preservar si ya fue autorizado o tiene OV generada
                    if prev_dia.get("status") == "Autorizado" or prev_dia.get("ov_generada"):
                        datos_dia["status"]      = prev_dia["status"]
                        datos_dia["ov_generada"] = prev_dia["ov_generada"]
                # Actualizar stock y campos nuevos (pueden haber cambiado)
                prev.update({
                    "stock_actual": nuevo_item["stock_actual"],
                    "stock_lg":     nuevo_item["stock_lg"],
                    "cw_line":      nuevo_item.get("cw_line"),
                    "model":        nuevo_item.get("model"),
                    "id1":          nuevo_item.get("id1"),
                    "dias":         nuevo_item["dias"],
                })
 
        # SQLAlchemy no detecta mutaciones en JSON automáticamente
        from sqlalchemy.orm.attributes import flag_modified
        plan_existente.items = nuevos_items
        flag_modified(plan_existente, "items")
        await db.commit()
        await db.refresh(plan_existente)
 
        return {
            "message":    f"Plan semana {identificador} actualizado",
            "total_skus": len(nuevos_items),
            "accion":     "actualizado",
        }
 
    # ── Crear nuevo plan ──
    nuevo_plan = PlanVentas(
        identificador_semana=identificador,
        fecha_inicio_semana=fecha_inicio_semana,
        items=nuevos_items,
        importado_por=current_user.username,
    )
    db.add(nuevo_plan)
    await db.commit()
    await db.refresh(nuevo_plan)
 
    return {
        "message":    f"Plan semana {identificador} creado",
        "total_skus": len(nuevos_items),
        "accion":     "creado",
    }


@router.post("/plan-ventas/autorizar")
@router.post("/plan-ventas/autorizar/")
async def autorizar_ventas_masivo(
    data: AutorizarVentasMasivo,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_ventas_role(current_user)

    result = await db.execute(
        select(PlanVentas).where(PlanVentas.identificador_semana == data.identificador_semana)
    )
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    resultados = []
    items_actualizados = plan.items or []
    items_map = {item["sku"]: item for item in items_actualizados}

    for venta in data.ventas:
        ahora = ahora_local()
        ov_id = f"OV-{ahora.strftime('%Y%m%d%H%M%S')}-{venta.sku[-4:]}"

        orden = OrdenVenta(
            ov_id=ov_id,
            cliente_id="PLAN_VENTAS",
            nombre_cliente=f"Autorización Plan {data.identificador_semana}",
            estado="Pendiente de Envío",
            creado_por=current_user.username,
        )
        db.add(orden)
        await db.flush()

        item_ov = OrdenVentaItem(
            orden_venta_id=orden.id,
            sku_producto=venta.sku,
            cantidad=venta.cantidad,
        )
        db.add(item_ov)

        if venta.sku in items_map:
            dias = items_map[venta.sku].get("dias", {})
            if venta.dia in dias:
                dias[venta.dia]["status"] = "Autorizado"
                dias[venta.dia]["ov_generada"] = ov_id

        resultados.append(f"SKU {venta.sku} ({venta.dia}): ✅ OV creada: {ov_id}")

    plan.items = list(items_map.values())
    await db.commit()

    return {"resultados": resultados}


@router.post("/ventas/{ov_id}/despachar")
async def despachar_orden_venta(
    ov_id: str,
    data: EnvioLogisticoCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    # Se permite a Logística y Ventas
    if current_user.rol not in ("admin", "finanzas", "ventas", "logistica"):
        raise HTTPException(status_code=403, detail="Sin permisos de despacho")

    result = await db.execute(select(OrdenVenta).where(OrdenVenta.ov_id == ov_id))
    orden = result.scalar_one_or_none()
    if not orden or orden.estado != "Lista para Carga":
        raise HTTPException(status_code=400, detail="Orden no lista para carga")

    if data.cw_invoice:
        orden.cw_invoice = data.cw_invoice

    items_result = await db.execute(select(OrdenVentaItem).where(OrdenVentaItem.orden_venta_id == orden.id))
    items = items_result.scalars().all()

    ahora = ahora_local()
    envio_id = f"ENV-{ahora.strftime('%Y%m%d%H%M%S')}"

    # Generar el registro físico
    envio = EnvioVenta(
        envio_id=envio_id,
        orden_venta_id=orden.id,
        ov_id=ov_id,
        autorizado_por=current_user.username,
        no_camion=data.no_camion,
        chofer=data.chofer,
        status_salida=data.status_salida,
        items_enviados=[{"sku": i.sku_producto, "qty": i.cantidad} for i in items]
    )
    db.add(envio)
    orden.estado = "Despachada"
    
    await db.commit()
    return {"message": "Despacho registrado exitosamente", "envio_id": envio_id}

# 2. Endpoint para Reporte Diario/Mensual Excel
@router.get("/ventas/reporte-diario/excel")
async def descargar_reporte_ventas(
    fecha: str = Query(..., description="YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db)
):
    # Filtramos envíos de ese día
    query = select(EnvioVenta, OrdenVenta).join(OrdenVenta).where(
        func.date(EnvioVenta.fecha_envio) == fecha
    )
    result = await db.execute(query)
    registros = result.all()

    datos_reporte = []
    for envio, orden in registros:
        for item in envio.items_enviados:
            datos_reporte.append({
                "NO. PARTE": item["sku"],
                "DESCRIPCION": "Extraer de maestro", # Aquí cruzarías con modelo Producto
                "LINE": "R1/R2", 
                "NO. EMB": envio.id,
                "CANTIDAD": item["qty"],
                "FECHA": envio.fecha_envio.strftime("%Y-%m-%d"),
                "NO. DEPARTURE": orden.ov_id,
                "CW INVOICE": orden.cw_invoice or "NO PO",
                "HRA DE SALIDA": envio.fecha_envio.strftime("%H:%M:%S"),
                "SALIDA STATUS": envio.status_salida,
                "CANTIDAD FINAL": item["qty"],
                "CAPTURISTA": envio.autorizado_por,
                "NO. CAMION": envio.no_camion,
                "CHOFER": envio.chofer,
                "COMENTARIO": orden.notas
            })

    df = pd.DataFrame(datos_reporte)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='VENTA REPORTE(D)', index=False)
        # Puedes añadir más sheets (BOM, RETURN) aquí.
    
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CW_Venta_Reporte_{fecha}.xlsx"}
    )


# ========================
# PROVEEDORES — CRUD
# ========================
@router.get("/proveedores", response_model=List[ProveedorResponse])
@router.get("/proveedores/", response_model=List[ProveedorResponse])
async def listar_proveedores(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)
    result = await db.execute(
        select(Proveedor).options(selectinload(Proveedor.materiales)).order_by(Proveedor.razon_social.asc())
    )
    return result.scalars().unique().all()


@router.post("/proveedores", response_model=ProveedorResponse)
@router.post("/proveedores/", response_model=ProveedorResponse)
async def crear_proveedor(
    data: ProveedorCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    # Validar RFC duplicado
    rfc_check = await db.execute(select(Proveedor).where(Proveedor.rfc == data.rfc.strip().upper()))
    if rfc_check.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Ya existe un proveedor con el RFC {data.rfc}")

    proveedor = Proveedor(
        razon_social=data.razon_social.strip(),
        rfc=data.rfc.strip().upper(),
        lead_time_dias=data.lead_time_dias,
        condiciones_pago=data.condiciones_pago,
        dias_credito=data.dias_credito,
        estatus_calidad=data.estatus_calidad,
        direccion=data.direccion,
        nombre_ventas=data.nombre_contacto,
        numero_contacto=data.numero_contacto,
        correo_contacto=data.correo_contacto,
        notas=data.notas,
    )
    db.add(proveedor)
    await db.flush()

    for mat in data.materiales:
        material = ProveedorMaterial(
            proveedor_id=proveedor.id,
            sku_material=mat.sku_material.strip().upper(),
            codigo_proveedor=mat.codigo_proveedor.strip() if mat.codigo_proveedor else None,
            costo_unitario=mat.costo_unitario,
            moneda=mat.moneda,
        )
        db.add(material)

    await db.commit()
    await recalcular_score(proveedor.id, db)

    # Refrescar con relación cargada para la respuesta
    result = await db.execute(
        select(Proveedor).options(selectinload(Proveedor.materiales)).where(Proveedor.id == proveedor.id)
    )
    proveedor = result.scalar_one()
    return proveedor


@router.put("/proveedores/{proveedor_id}", response_model=ProveedorResponse)
@router.put("/proveedores/{proveedor_id}/", response_model=ProveedorResponse)
async def actualizar_proveedor(
    proveedor_id: int,
    data: ProveedorUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    if data.razon_social is not None:
        proveedor.razon_social = data.razon_social.strip()
    if data.rfc is not None:
        proveedor.rfc = data.rfc.strip().upper()
    if data.lead_time_dias is not None:
        proveedor.lead_time_dias = data.lead_time_dias
    if data.condiciones_pago is not None:
        proveedor.condiciones_pago = data.condiciones_pago
    if data.dias_credito is not None:
        proveedor.dias_credito = data.dias_credito
    if data.estatus_calidad is not None:
        proveedor.estatus_calidad = data.estatus_calidad
    if data.direccion is not None:
        proveedor.direccion = data.direccion
    if data.nombre_contacto is not None:
        proveedor.nombre_ventas = data.nombre_contacto
    if data.numero_contacto is not None:
        proveedor.numero_contacto = data.numero_contacto
    if data.correo_contacto is not None:
        proveedor.correo_contacto = data.correo_contacto
    if data.notas is not None:
        proveedor.notas = data.notas

    if data.materiales is not None:
        # Eliminar materiales existentes
        await db.execute(
            delete(ProveedorMaterial).where(ProveedorMaterial.proveedor_id == proveedor_id)
        )
        # Insertar nuevos
        for mat in data.materiales:
            material = ProveedorMaterial(
                proveedor_id=proveedor_id,
                sku_material=mat.sku_material.strip().upper(),
                codigo_proveedor=mat.codigo_proveedor.strip() if mat.codigo_proveedor else None,
                costo_unitario=mat.costo_unitario,
                moneda=mat.moneda,
            )
            db.add(material)

    await db.commit()
    await recalcular_score(proveedor.id, db)

    result = await db.execute(
        select(Proveedor).options(selectinload(Proveedor.materiales)).where(Proveedor.id == proveedor.id)
    )
    proveedor = result.scalar_one()
    return proveedor


@router.delete("/proveedores/{proveedor_id}")
@router.delete("/proveedores/{proveedor_id}/")
async def eliminar_proveedor(
    proveedor_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    await db.delete(proveedor)
    await db.commit()
    return {"message": "Proveedor eliminado exitosamente"}


@router.get("/proveedores/buscar-por-sku/{sku}")
async def buscar_proveedores_por_sku(
    sku: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Busca qué proveedores surten un SKU específico (Para resolver las OCs POR-ASIGNAR)"""
    require_compras_role(current_user)
    
    result = await db.execute(
        select(ProveedorMaterial, Proveedor)
        .join(Proveedor, ProveedorMaterial.proveedor_id == Proveedor.id)
        .where(ProveedorMaterial.sku_material == sku.upper())
    )
    mapping = result.all()
    
    return [
        {
            "proveedor_id": prov.id,
            "razon_social": prov.razon_social,
            "uuid": prov.uuid,
            "costo_unitario": mat.costo_unitario,
            "moneda": mat.moneda,
            "lead_time_dias": prov.lead_time_dias,
            "estatus_calidad": prov.estatus_calidad
        }
        for mat, prov in mapping
    ]


# ========================
# PROVEEDORES — SCORE & EVENTOS
# ========================
@router.get("/proveedores/{proveedor_id}/score", response_model=ProveedorScoreResponse)
@router.get("/proveedores/{proveedor_id}/score/")
async def obtener_score_proveedor(
    proveedor_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    # Lazy recalc: si la última actualización fue hace más de 24 horas
    ahora = ahora_local()
    if proveedor.score_updated_at is None or (ahora - proveedor.score_updated_at).total_seconds() > 86400:
        score_data = await recalcular_score(proveedor_id, db)
    else:
        from app.services.proveedor_score import _recomendacion_estatus
        score_data = {
            "proveedor_id": proveedor.id,
            "razon_social": proveedor.razon_social,
            "score_calidad": proveedor.score_calidad or 100.0,
            "score_detalle": proveedor.score_detalle or {},
            "score_updated_at": proveedor.score_updated_at,
            "recomendacion_estatus": _recomendacion_estatus(proveedor.score_calidad or 100.0, proveedor.estatus_calidad),
        }

    return ProveedorScoreResponse(**score_data)


@router.get("/proveedores/{proveedor_id}/eventos", response_model=list[ProveedorEventoResponse])
@router.get("/proveedores/{proveedor_id}/eventos/")
async def listar_eventos_proveedor(
    proveedor_id: int,
    limite: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(
        select(ProveedorEvento)
        .where(ProveedorEvento.proveedor_id == proveedor_id)
        .order_by(ProveedorEvento.fecha.desc())
        .limit(limite)
    )
    return result.scalars().all()


@router.post("/proveedores/{proveedor_id}/eventos", response_model=ProveedorEventoResponse)
@router.post("/proveedores/{proveedor_id}/eventos/")
async def crear_evento_proveedor(
    proveedor_id: int,
    data: ProveedorEventoCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    evento = await registrar_evento(
        proveedor_id=proveedor_id,
        tipo_evento=data.tipo_evento,
        impacto=data.impacto,
        referencia_id=data.referencia_id,
        descripcion=data.descripcion,
        registrado_por=current_user.username,
        db=db,
    )
    return evento


@router.post("/proveedores/{proveedor_id}/recalcular")
@router.post("/proveedores/{proveedor_id}/recalcular/")
async def forzar_recalcular_score(
    proveedor_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    score_data = await recalcular_score(proveedor_id, db)
    return ProveedorScoreResponse(**score_data)


@router.get("/proveedores/ranking", response_model=List[ProveedorResponse])
@router.get("/proveedores/ranking/", response_model=List[ProveedorResponse])
async def ranking_proveedores(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_compras_role(current_user)

    result = await db.execute(
        select(Proveedor)
        .options(selectinload(Proveedor.materiales))
        .order_by(Proveedor.score_calidad.desc().nullslast())
    )
    proveedores = result.scalars().unique().all()
    return proveedores


# ========================
# LIMPIEZA DE DATOS
# ========================
@router.post("/limpiar/compras-completadas")
@router.post("/limpiar/compras-completadas/")
async def limpiar_compras_completadas(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_finanzas_role(current_user)
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede realizar limpiezas")

    result = await db.execute(
        delete(OrdenCompra).where(OrdenCompra.status == "Completada")
    )
    await db.commit()
    return {"message": f"{result.rowcount} órdenes de compra completadas eliminadas"}


@router.post("/limpiar/devoluciones-finalizadas")
@router.post("/limpiar/devoluciones-finalizadas/")
async def limpiar_devoluciones_finalizadas(
    dias: int = Query(90, ge=1),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    require_finanzas_role(current_user)
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede realizar limpiezas")

    fecha_limite = ahora_local() - timedelta(days=dias)
    result = await db.execute(
        delete(Devolucion).where(
            and_(
                Devolucion.estado_inspeccion == "Finalizado",
                Devolucion.fecha_devolucion < fecha_limite,
            )
        )
    )
    await db.commit()
    return {"message": f"{result.rowcount} devoluciones finalizadas eliminadas (>{dias} días)"}